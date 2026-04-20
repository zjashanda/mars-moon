#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Case-first runner for the Dooya curtain project."""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import json
import re
import shutil
import sys
import time
from collections import Counter, deque
from pathlib import Path
from typing import Any

import serial
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from sample.voiceTestLite import REGEX_CANDIDATES, SerialReader, play_audio, tts_generate

DEFAULT_CONFIG = ROOT / "generated" / "deviceInfo_dooya.json"
DEFAULT_CASES = ROOT / "generated" / "cases.json"
DEFAULT_CASE_WORKBOOK = ROOT / "generated" / "CSK5062_杜亚窗帘_测试用例.xlsx"
DEFAULT_SPEC = ROOT / "work" / "normalized_spec.json"
VOICE_ACTION_TYPES = {"wake", "say", "volume_walk", "assert_wake_repeats", "assert_no_wake", "phrase_check"}


RUNTIME_GAP_RULES = {
    "TC_CTRL_001": {
        "field": "ctrl_pair_protocol_runtime",
        "label": "配对遥控器协议",
        "expected": "55 AA 05 01 12 00 80 AF",
    },
    "TC_CTRL_007": {
        "field": "ctrl_clear_protocol_runtime",
        "label": "清除遥控器协议",
        "expected": "55 AA 05 01 16 00 18 AC",
    },
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_hex(value: str) -> str:
    text = str(value or "")
    tokens = re.findall(r"(?i)\b[0-9A-F]{2}\b", text)
    if tokens:
        return " ".join(token.upper() for token in tokens)
    cleaned = re.sub(r"[^0-9A-Fa-f]", "", text)
    if not cleaned:
        return ""
    if len(cleaned) % 2:
        cleaned = f"0{cleaned}"
    return " ".join(cleaned[index:index + 2].upper() for index in range(0, len(cleaned), 2))


def parse_tone_id(value: str) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith("TONE_ID_"):
        return text
    match = re.search(r"(\d+)", text)
    if match:
        return f"TONE_ID_{int(match.group(1))}"
    return text


def tone_id_number(value: str) -> str:
    text = parse_tone_id(value)
    match = re.search(r"(\d+)", text)
    return match.group(1) if match else ""


def decode_recognized(text: str, spell2zh: dict[str, str]) -> str:
    value = str(text or "").strip().strip('"').strip("'").replace('\\"', '"')
    if "\\u" in value:
        try:
            value = value.encode("utf-8").decode("unicode_escape")
        except UnicodeDecodeError:
            pass
    return spell2zh.get(value, value)


def copy_into(src: Path, dst: Path) -> None:
    if src.is_dir():
        shutil.copytree(src, dst, dirs_exist_ok=True)
    elif src.is_file():
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)


def path_is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except Exception:
        return False


def copy_tree_filtered(src: Path, dst: Path, ignored_roots: list[Path] | None = None) -> None:
    ignored = [item.resolve() for item in (ignored_roots or [])]

    def _ignore(current_dir: str, names: list[str]) -> set[str]:
        current = Path(current_dir).resolve()
        skipped: set[str] = set()
        for name in names:
            child = current / name
            try:
                resolved = child.resolve()
            except Exception:
                resolved = child
            if any(resolved == root or path_is_within(resolved, root) for root in ignored):
                skipped.add(name)
        return skipped

    shutil.copytree(src, dst, dirs_exist_ok=True, ignore=_ignore)


class ToolLogger:
    def __init__(self, log_dir: Path, verbose: bool = True) -> None:
        self.log_dir = log_dir
        self.verbose = verbose
        self.log_dir.mkdir(parents=True, exist_ok=True)
        self.path = self.log_dir / "tool.log"
        self._fp = self.path.open("a", encoding="utf-8")

    def _log(self, level: str, message: str) -> None:
        ts = dt.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        line = f"[{ts}][{level}] {message}"
        if self.verbose:
            print(line)
        self._fp.write(line + "\n")
        self._fp.flush()

    def info(self, message: str) -> None:
        self._log("INFO", message)

    def warn(self, message: str) -> None:
        self._log("WARN", message)

    def error(self, message: str) -> None:
        self._log("ERROR", message)

    def debug(self, message: str) -> None:
        self._log("DEBUG", message)

    def close(self) -> None:
        self._fp.close()


class ProtocolMonitor:
    def __init__(self, port: str, baudrate: int, frame_header: str, frame_length: int, log_path: Path, logger: ToolLogger) -> None:
        self.port = port
        self.baudrate = baudrate
        self.frame_header = bytes.fromhex(frame_header)
        self.frame_length = int(frame_length)
        self.log_path = log_path
        self.logger = logger
        self.serial = None
        self.buffer = bytearray()
        self.frames: list[str] = []
        self.partial_frames: list[str] = []
        self.stop_flag = False
        self.thread = None

    def connect(self) -> bool:
        try:
            self.serial = serial.Serial(self.port, self.baudrate, timeout=0.05)
            self.logger.info(f"协议串口连接成功: {self.port} @ {self.baudrate}")
            return True
        except Exception as exc:
            self.logger.error(f"协议串口连接失败: {self.port} @ {self.baudrate} | {exc}")
            return False

    def start(self) -> None:
        import threading

        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def _run(self) -> None:
        self.log_path.parent.mkdir(parents=True, exist_ok=True)
        with self.log_path.open("a", encoding="utf-8") as fp:
            while not self.stop_flag:
                try:
                    if not self.serial or not self.serial.is_open:
                        break
                    waiting = self.serial.in_waiting
                    data = self.serial.read(waiting or 1)
                    if not data:
                        continue
                    fp.write(f"[{dt.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] RX {normalize_hex(data.hex())}\n")
                    fp.flush()
                    self.buffer.extend(data)
                    self._extract_frames()
                except Exception as exc:
                    if self.stop_flag or not self.serial or not self.serial.is_open:
                        break
                    self.logger.debug(f"协议串口读取异常: {exc}")
                    break

    def _extract_frames(self) -> None:
        header = self.frame_header
        while True:
            if len(self.buffer) < len(header):
                return
            if self.buffer[: len(header)] != header:
                position = self.buffer.find(header)
                if position < 0:
                    self.buffer.clear()
                    return
                del self.buffer[:position]
            if len(self.buffer) < self.frame_length:
                if self.buffer.startswith(header):
                    partial = normalize_hex(self.buffer.hex())
                    if partial and partial not in self.partial_frames:
                        self.partial_frames.append(partial)
                return
            frame = bytes(self.buffer[: self.frame_length])
            del self.buffer[: self.frame_length]
            self.frames.append(normalize_hex(frame.hex()))

    def clear(self) -> None:
        self.frames = []
        self.partial_frames = []

    def get_frames(self) -> list[str]:
        return list(self.frames)

    def get_partial_frames(self) -> list[str]:
        return list(self.partial_frames)

    def write_hex(self, hex_value: str) -> None:
        payload = bytes.fromhex(normalize_hex(hex_value).replace(" ", ""))
        self.serial.write(payload)
        with self.log_path.open("a", encoding="utf-8") as fp:
            fp.write(f"[{dt.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] TX {normalize_hex(hex_value)}\n")

    def close(self) -> None:
        self.stop_flag = True
        if self.serial and self.serial.is_open:
            self.serial.close()
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=1.0)


class DooyaRunner:
    def __init__(self, config_path: Path, cases_path: Path, args: argparse.Namespace) -> None:
        self.config_path = config_path
        self.cases_path = cases_path
        self.args = args
        self.config = load_json(config_path)
        self.cases = load_json(cases_path)
        self.workspace_root = Path(self.config.get("workspaceRoot", ROOT)).resolve()
        self.spec_path = self._resolve_spec_path()
        self.spec = load_json(self.spec_path) if self.spec_path and self.spec_path.is_file() else {}
        self.requirement_catalog = list(self.spec.get("requirement_catalog", []))
        self.requirement_case_map = self._build_requirement_case_map(self.cases)
        self._apply_runtime_overrides()
        self.policy = self.config["executionPolicy"]
        if args.result_dir:
            self.result_dir = Path(args.result_dir).resolve()
        else:
            self.result_dir = ROOT / self.policy["resultRoot"] / dt.datetime.now().strftime(self.policy["resultDirTimeFormat"])
        self.result_dir.mkdir(parents=True, exist_ok=True)
        self.log = ToolLogger(self.result_dir, verbose=not args.quiet)
        self.wav_dir = ROOT / self.policy["wavRoot"]
        self.wav_dir.mkdir(parents=True, exist_ok=True)
        self.spell2zh = self.config.get("spell2zh", {})
        self.selected_cases = self._select_cases()
        self.reader: SerialReader | None = None
        self.proto: ProtocolMonitor | None = None
        self.last_action: dict[str, Any] | None = None
        self.capabilities: dict[str, dict[str, Any]] = {}
        self.runtime_gaps: list[dict[str, str]] = []
        self.serial_checks: dict[str, dict[str, Any]] = {}
        self.case_results_by_id: dict[str, dict[str, Any]] = {}
        self.unexpected_reboot_count = 0
        self.reboot_count_baseline = 0
        self.last_work_mode = ""
        self.consecutive_wake_failure_count = 0
        self.protocol_truncation_count = 0
        self.stop_reason = ""
        self._copy_inputs()

    def _apply_runtime_overrides(self) -> None:
        device_info = self.config.setdefault("deviceListInfo", {})
        log_cfg = device_info.setdefault("cskApLog", {})
        uart_cfg = device_info.setdefault("uart1", {})
        power_cfg = self.config.setdefault("powerControl", {})
        if self.args.log_port:
            log_cfg["port"] = self.args.log_port
        if self.args.log_baud:
            log_cfg["baudRate"] = self.args.log_baud
        if self.args.uart1_port:
            uart_cfg["port"] = self.args.uart1_port
        if self.args.uart1_baud:
            uart_cfg["baudRate"] = self.args.uart1_baud
        if self.args.uart1_frame_header:
            uart_cfg["frameHeader"] = self.args.uart1_frame_header
        if self.args.uart1_frame_length:
            uart_cfg["frameLength"] = self.args.uart1_frame_length
        if getattr(self.args, "ctrl_port", ""):
            power_cfg["port"] = self.args.ctrl_port
        if getattr(self.args, "ctrl_baud", 0):
            power_cfg["baudRate"] = self.args.ctrl_baud

    def _resolve_spec_path(self) -> Path | None:
        candidates = [
            self.workspace_root / "work" / "normalized_spec.json",
            self.config_path.resolve().parent.parent / "work" / "normalized_spec.json",
            DEFAULT_SPEC,
        ]
        seen: set[Path] = set()
        for candidate in candidates:
            resolved = candidate.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            if resolved.is_file():
                return resolved
        return None

    def _build_requirement_case_map(self, cases: list[dict[str, Any]]) -> dict[str, list[str]]:
        mapping: dict[str, list[str]] = {}
        for item in cases:
            for requirement_id in item.get("requirement_ids", []):
                requirement_text = str(requirement_id).strip()
                if not requirement_text:
                    continue
                bucket = mapping.setdefault(requirement_text, [])
                if item["case_id"] not in bucket:
                    bucket.append(item["case_id"])
        return mapping

    def _copy_inputs(self) -> None:
        shutil.copy2(self.config_path, self.result_dir / self.config_path.name)
        shutil.copy2(self.cases_path, self.result_dir / self.cases_path.name)
        if DEFAULT_CASE_WORKBOOK.is_file():
            shutil.copy2(DEFAULT_CASE_WORKBOOK, self.result_dir / "testCases.xlsx")
        if self.spec_path and self.spec_path.is_file():
            shutil.copy2(self.spec_path, self.result_dir / self.spec_path.name)

    def archive_workspace_artifacts(self) -> None:
        archive_root = self.result_dir / "artifacts"
        sources = {
            "generated": self.workspace_root / "generated",
            "work": self.workspace_root / "work",
            "firmware": self.workspace_root / "artifacts" / "firmware",
            "burn": self.workspace_root / "artifacts" / "burn",
        }
        for name, src in sources.items():
            if src.exists():
                dst = archive_root / name
                if src.is_dir():
                    copy_tree_filtered(src, dst, ignored_roots=[self.result_dir, archive_root])
                else:
                    copy_into(src, dst)

    def _select_cases(self) -> list[dict[str, Any]]:
        selected = list(self.cases)
        if self.args.case_id:
            wanted = [item.strip() for item in self.args.case_id.split(",") if item.strip()]
            by_id = {item["case_id"]: item for item in self.cases}
            selected = [by_id[item] for item in wanted if item in by_id]
        if self.args.module:
            wanted = {item.strip() for item in self.args.module.split(",") if item.strip()}
            selected = [item for item in selected if item["module"] in wanted]
        if self.args.priority:
            wanted = {item.strip() for item in self.args.priority.split(",") if item.strip()}
            selected = [item for item in selected if item["priority"] in wanted]
        if self.args.limit:
            selected = selected[: self.args.limit]
        return selected

    def required_words(self) -> list[str]:
        words = [self.config["wakeupWord"]]
        for case_item in self.selected_cases:
            for action in case_item["actions"]:
                if action["type"] in {"wake", "say", "assert_wake_repeats"}:
                    words.append(action["word"])
                if action["type"] == "volume_walk":
                    command_word = str(action.get("command_word") or action.get("word") or "").strip()
                    if command_word:
                        words.append(command_word)
                if action["type"] == "assert_no_wake":
                    words.extend(str(item).strip() for item in action.get("words", []) if str(item).strip())
                if action["type"] == "phrase_check":
                    wake_word = str(action.get("wake_word") or self.config["wakeupWord"]).strip()
                    if wake_word:
                        words.append(wake_word)
                    for item in action.get("items", []):
                        word = str(item.get("word", "")).strip()
                        if word:
                            words.append(word)
        ordered: list[str] = []
        seen: set[str] = set()
        for word in words:
            if word not in seen:
                seen.add(word)
                ordered.append(word)
        return ordered

    def configured_wake_timeout_s(self) -> float:
        requirement_defaults = self.spec.get("requirement", {}).get("defaults", {})
        config_defaults = self.config.get("requirementDefaults", {})
        value = (
            requirement_defaults.get("wake_timeout_s")
            or config_defaults.get("wake_timeout_s")
            or self.config.get("wakeTimeoutS")
            or 15
        )
        try:
            return float(value)
        except Exception:
            return 15.0

    def case_dependencies(self, case_item: dict[str, Any]) -> list[str]:
        deps: list[str] = []
        action_types = {action["type"] for action in case_item["actions"]}
        if action_types.intersection(VOICE_ACTION_TYPES):
            deps.append("voice")
        inject_actions = [action for action in case_item["actions"] if action["type"] == "inject_protocol"]
        inject_actions = [action for action in inject_actions if not bool(action.get("expect_no_response"))]
        if inject_actions and not all(bool(action.get("requires_active_window")) for action in inject_actions):
            deps.append("uart_inject")
        if "manual_power_cycle" in action_types:
            deps.append("manual_power")
        return deps

    def set_capability(self, name: str, supported: bool, reason: str, evidence: list[str] | None = None) -> None:
        self.capabilities[name] = {
            "supported": supported,
            "reason": reason,
            "evidence": evidence or [],
        }

    def ensure_audio(self) -> bool:
        missing: list[str] = []
        for word in self.required_words():
            path = self.wav_dir / f"{word}.mp3"
            if not path.is_file() or path.stat().st_size == 0:
                missing.append(word)
        if not missing:
            self.log.info(f"音频检查通过：{len(self.required_words())} 条")
            return True
        tts_cfg = self.config.get("ttsConfig", {})
        if not tts_cfg.get("app_id") or not tts_cfg.get("api_key"):
            self.log.error(f"缺少 TTS 配置，无法补齐音频：{missing}")
            return False
        for index, word in enumerate(missing, start=1):
            path = self.wav_dir / f"{word}.mp3"
            self.log.info(f"TTS [{index}/{len(missing)}] {word}")
            try:
                tts_generate(word, str(path), tts_cfg)
            except Exception as exc:
                self.log.error(f"TTS 失败：{word} | {exc}")
                return False
        return True

    def play_word(self, word: str) -> bool:
        path = self.wav_dir / f"{word}.mp3"
        self.log.info(f"播放音频：{word}")
        return play_audio(str(path), self.log)

    def connect(self) -> bool:
        return self._open_observers()

    def _open_observers(self) -> bool:
        log_cfg = self.config["deviceListInfo"]["cskApLog"]
        self.reader = SerialReader(
            log_cfg["port"],
            log_cfg["baudRate"],
            log_cfg.get("regex", {}),
            self.log,
            serial_log_dir=str(self.result_dir),
        )
        if not self.reader.connect():
            return False
        self.reader.start()
        uart_cfg = self.config["deviceListInfo"]["uart1"]
        self.proto = ProtocolMonitor(
            uart_cfg["port"],
            uart_cfg["baudRate"],
            uart_cfg.get("frameHeader", "55 AA"),
            uart_cfg.get("frameLength", 8),
            self.result_dir / self.policy["protocolLogName"],
            self.log,
        )
        if not self.proto.connect():
            return False
        self.proto.start()
        time.sleep(1.0)
        return True

    def configured_power_boot_wait_s(self) -> float:
        pretest_cfg = self.config.get("pretestConfig", {}) or {}
        boot_wait = pretest_cfg.get("bootWait")
        if boot_wait is not None:
            try:
                return max(float(boot_wait), 0.0)
            except (TypeError, ValueError):
                pass
        return max(float(self.policy.get("postBootReadyDelayS", 4.0)), 0.0)

    def reopen_observers_after_power_event(self, reason: str, wait_s: float = 0.0) -> bool:
        self.log.info(f"{reason}：关闭并重连日志/协议串口")
        if self.proto:
            self.proto.close()
            self.proto = None
        if self.reader:
            self.reader.close()
            self.reader = None
        if wait_s > 0:
            self.log.info(f"{reason}：等待设备重新枚举/启动 {wait_s:.1f}s")
            time.sleep(wait_s)
        return self._open_observers()

    def close(self) -> None:
        if self.proto:
            self.proto.close()
        if self.reader:
            self.reader.close()
        self.log.close()

    def clear_observation(self) -> None:
        if self.reader:
            self.reader.clear()
        if self.proto:
            self.proto.clear()

    def observed_asr(self) -> list[str]:
        if not self.reader:
            return []
        return [decode_recognized(value, self.spell2zh) for value in self.reader.get_all("asrKw")]

    def observed_tones(self) -> list[str]:
        if not self.reader:
            return []
        return [parse_tone_id(value) for value in self.reader.get_all("playId")]

    def observed_serial_protocols(self) -> list[str]:
        protocols: list[str] = []
        if self.proto:
            protocols.extend(self.proto.get_frames())
            protocols.extend(self.proto.get_partial_frames())
        return [item for item in protocols if item]

    def observed_log_protocols(self, direction: str | None = None) -> list[str]:
        protocols: list[str] = []
        if not self.reader:
            return protocols
        normalized = str(direction or "").strip().lower()
        if normalized in {"", "send", "tx"}:
            protocols.extend(normalize_hex(item) for item in self.reader.get_all("sendMsg"))
        if normalized in {"", "recv", "rx"}:
            protocols.extend(normalize_hex(item) for item in self.reader.get_all("recvMsg"))
        return [item for item in protocols if item]

    def observed_protocols(self) -> list[str]:
        # Protocol assertions use the dedicated protocol UART as the source of
        # truth. Firmware log-side send/recv markers are kept only as evidence.
        return self.observed_serial_protocols()

    def observed_log_values(self, key: str) -> list[str]:
        if not self.reader:
            return []
        values = [str(item).strip() for item in self.reader.get_all(key) if str(item).strip()]
        if key in {"asrKw", "wakeKw"}:
            return [decode_recognized(value, self.spell2zh) for value in values]
        synthetic_values = self._synthetic_log_values(key)
        for item in synthetic_values:
            if item not in values:
                values.append(item)
        return values

    def _recent_serial_text(self, last_n: int = 200) -> str:
        if not self.reader:
            return ""
        lines = [line.strip() for line in self.reader.get_recent_lines()[-last_n:] if line.strip()]
        if not lines:
            return ""
        # Some firmware logs split a single "refresh config ..." record across
        # multiple bursts. Join them back together so state values are not lost.
        return "".join(lines)

    def _synthetic_log_values(self, key: str) -> list[str]:
        combined = self._recent_serial_text()
        if not combined:
            return []
        patterns: list[str] = []
        if key == "wakeup":
            patterns = [r"wakeup\s*[:=]\s*(\d+)"]
        elif key == "workMode":
            patterns = [r"(?:work mode\s*:\s*|work_mode\s*[:=]\s*)(\d+)"]
        elif key == "curtainMode":
            patterns = [r"(?:curtain type\s*:\s*|curtain_type\s*[:=]\s*)(\d+)"]
        elif key == "configRefresh":
            patterns = [r"(refresh config .*?(?:save config success|TIME_OUT|MODE=\d+|$))"]
        if not patterns:
            return []
        values: list[str] = []
        for pattern in patterns:
            try:
                matches = list(re.finditer(pattern, combined))
            except re.error:
                continue
            for match in matches:
                try:
                    value = match.group(1).strip()
                except IndexError:
                    value = ""
                if value and value not in values:
                    values.append(value)
        return values

    def observed_recognized_values(self) -> list[str]:
        recognized = self.observed_log_values("asrKw")
        for value in self.observed_log_values("wakeKw"):
            if value not in recognized:
                recognized.append(value)
        return recognized

    def latest_regex_value(self, key: str) -> str:
        if not self.reader:
            return ""
        current = str(self.reader.get(key) or "").strip()
        if current:
            if key == "workMode":
                self.last_work_mode = current
            return current
        log_cfg = self.config.get("deviceListInfo", {}).get("cskApLog", {})
        patterns: list[str] = []
        primary = str(log_cfg.get("regex", {}).get(key, "")).strip()
        if primary:
            patterns.append(primary)
        for pattern in log_cfg.get("regexCandidates", {}).get(key, []) or []:
            candidate = str(pattern).strip()
            if candidate and candidate not in patterns:
                patterns.append(candidate)
        if not patterns:
            return ""
        for line in reversed(self.reader.get_recent_lines()):
            for pattern in patterns:
                try:
                    match = re.match(pattern, line)
                except re.error:
                    continue
                if match:
                    try:
                        value = match.group(1).strip()
                    except IndexError:
                        value = ""
                    if value:
                        if key == "workMode":
                            self.last_work_mode = value
                        return value
        synthetic_values = self._synthetic_log_values(key)
        if synthetic_values:
            value = synthetic_values[-1]
            if key == "workMode":
                self.last_work_mode = value
            return value
        if key == "workMode" and self.last_work_mode:
            return self.last_work_mode
        return ""

    def current_work_mode(self) -> str:
        return self.latest_regex_value("workMode")

    def wake_tones_by_work_mode(self) -> dict[str, list[str]]:
        raw = self.policy.get("wakeToneByWorkMode", {}) or {
            "0": ["TONE_ID_0"],
            "1": ["TONE_ID_15"],
        }
        normalized: dict[str, list[str]] = {}
        for mode, tones in raw.items():
            items = tones if isinstance(tones, list) else [tones]
            parsed = [parse_tone_id(item) for item in items if parse_tone_id(item)]
            if parsed:
                normalized[str(mode).strip()] = parsed
        return normalized

    def acceptable_wake_tones(self, explicit_tone: str = "", mode_hint: str = "") -> list[str]:
        tones: list[str] = []
        if explicit_tone:
            parsed = parse_tone_id(explicit_tone)
            if parsed:
                tones.append(parsed)
        tone_map = self.wake_tones_by_work_mode()
        mode = str(mode_hint or self.current_work_mode()).strip()
        if mode:
            candidate_tones = tone_map.get(mode, [])
        else:
            candidate_tones = []
        # Wake-stage success only depends on seeing a valid wake prompt; later
        # command-tone assertions still verify whether the device stayed in the expected mode.
        for items in tone_map.values():
            for tone in items:
                if tone not in candidate_tones:
                    candidate_tones.append(tone)
        for tone in candidate_tones:
            if tone not in tones:
                tones.append(tone)
        return tones

    def configured_wake_ready_markers(self) -> list[str]:
        raw = self.policy.get("wakeReadyMarkers", []) or ["wake up ready to asr"]
        items = raw if isinstance(raw, list) else [raw]
        return [str(item).strip() for item in items if str(item).strip()]

    def observe_wake_ready(self, after_clock: float | None = None) -> tuple[bool, float | None, list[str]]:
        ready_values = self.observed_log_values("wakeReady")
        ready_clock = self._find_latest_serial_clock(self.configured_wake_ready_markers(), after_clock=after_clock)
        evidence: list[str] = []
        for item in ready_values[-2:]:
            evidence.append(f"wakeReady={item}")
        if not evidence and ready_clock is not None:
            evidence.append("wakeReady=marker")
        return bool(ready_values) or ready_clock is not None, ready_clock, evidence

    def configured_playback_start_markers(self) -> list[str]:
        raw = self.policy.get("playbackStartMarkers", []) or ["play start"]
        items = raw if isinstance(raw, list) else [raw]
        return [str(item).strip() for item in items if str(item).strip()]

    def configured_playback_stop_markers(self) -> list[str]:
        raw = self.policy.get("playbackStopMarkers", []) or ["play stop"]
        items = raw if isinstance(raw, list) else [raw]
        return [str(item).strip() for item in items if str(item).strip()]

    def _latest_playback_start_clock(self, after_clock: float | None = None) -> float | None:
        return self._find_latest_serial_clock(self.configured_playback_start_markers(), after_clock=after_clock)

    def _latest_playback_stop_clock(self, after_clock: float | None = None) -> float | None:
        return self._find_latest_serial_clock(self.configured_playback_stop_markers(), after_clock=after_clock)

    def wait_for_response_playback_complete(
        self,
        after_clock: float | None,
        *,
        context: str = "",
        timeout_s: float | None = None,
    ) -> tuple[bool, list[str]]:
        if after_clock is None:
            return True, []
        wait_s = max(float(timeout_s if timeout_s is not None else self.policy.get("responsePlaybackWaitS", 8.0)), 0.5)
        post_gap_s = max(float(self.policy.get("postPlaybackGapS", 0.2)), 0.0)
        implicit_complete_s = max(
            float(
                self.policy.get(
                    "playbackImplicitCompleteS",
                    self.policy.get("successSettleS", 2.0),
                )
            ),
            0.5,
        )
        latest_start = self._latest_playback_start_clock(after_clock=after_clock)
        latest_stop = self._latest_playback_stop_clock(after_clock=after_clock)
        if latest_start is None:
            return True, []
        if latest_stop is not None and self._clock_is_after(latest_stop, latest_start):
            if post_gap_s > 0:
                time.sleep(post_gap_s)
            return True, [f"playStop@{latest_stop:.3f}"]
        elapsed_since_start = self._clock_diff(self._now_clock(), latest_start)
        if elapsed_since_start is not None and elapsed_since_start >= implicit_complete_s:
            if post_gap_s > 0:
                time.sleep(post_gap_s)
            return True, [f"playbackGap@{elapsed_since_start:.3f}s"]
        if context:
            self.log.info(f"{context}：等待响应播报结束")
        deadline = time.time() + wait_s
        while time.time() < deadline:
            time.sleep(0.2)
            latest_start = self._latest_playback_start_clock(after_clock=after_clock)
            latest_stop = self._latest_playback_stop_clock(after_clock=after_clock)
            if latest_start is None:
                if post_gap_s > 0:
                    time.sleep(post_gap_s)
                return True, []
            if latest_stop is not None and self._clock_is_after(latest_stop, latest_start):
                if post_gap_s > 0:
                    time.sleep(post_gap_s)
                return True, [f"playStop@{latest_stop:.3f}"]
            elapsed_since_start = self._clock_diff(self._now_clock(), latest_start)
            if elapsed_since_start is not None and elapsed_since_start >= implicit_complete_s:
                if post_gap_s > 0:
                    time.sleep(post_gap_s)
                return True, [f"playbackGap@{elapsed_since_start:.3f}s"]
        evidence = []
        if latest_start is not None:
            evidence.append(f"playStart@{latest_start:.3f}")
        if latest_stop is not None:
            evidence.append(f"playStop@{latest_stop:.3f}")
        return False, evidence

    def wait_for_last_action_playback_complete(self, context: str) -> tuple[bool, list[str]]:
        if not self.last_action:
            return True, []
        reference_clock = self.last_action.get("reference_clock")
        if reference_clock is None:
            return True, []
        action_type = str(self.last_action.get("action_type") or "").strip()
        if action_type not in VOICE_ACTION_TYPES and action_type != "wait":
            return True, []
        return self.wait_for_response_playback_complete(reference_clock, context=context)

    def _evaluate_wake_signal(
        self,
        *,
        word: str,
        expect_tone: str,
        require_any_tone: bool,
        expected_protocol: str,
        mode_snapshot: str,
        match_after_clock: float | None,
        require_protocol_when_configured: bool,
        require_ready_signal: bool,
    ) -> dict[str, Any]:
        wake_values = [decode_recognized(value, self.spell2zh) for value in self.reader.get_all("wakeKw")]
        asr_values = self.observed_asr()
        recognized_values = list(wake_values)
        for value in asr_values:
            if value not in recognized_values:
                recognized_values.append(value)
        tones = self.observed_tones()
        mode_states = self.observed_log_values("modeState")
        protocols = self.observed_protocols()
        ready_ok, ready_clock, ready_evidence = self.observe_wake_ready(match_after_clock)
        recognition_clock = self._find_latest_serial_clock(["Wakeup:", "keyword:"], after_clock=match_after_clock)
        acceptable_tones = self.acceptable_wake_tones(expect_tone, mode_snapshot)
        tone_ok = any(tone in tones for tone in acceptable_tones) if acceptable_tones else (bool(tones) if require_any_tone else True)
        mode_active_ok = "1" in mode_states
        if not tone_ok and mode_active_ok:
            tone_ok = True
        protocol_required = bool(expected_protocol) and require_protocol_when_configured
        protocol_ok = not protocol_required or expected_protocol in protocols
        ready_required = require_ready_signal and bool(self.configured_wake_ready_markers())
        if not ready_ok and mode_active_ok and ready_required:
            ready_ok = True
            if "modeState=1" not in ready_evidence:
                ready_evidence.append("modeState=1")
        signal_evidence: list[str] = []
        if mode_snapshot:
            signal_evidence.append(f"workMode={mode_snapshot}")
        if mode_states:
            signal_evidence.append(f"modeState={'/'.join(mode_states[-3:])}")
        signal_evidence.extend(ready_evidence)
        if protocols:
            signal_evidence.append(f"protocols={'/'.join(protocols[-3:])}")
        wake_word_ok = word in recognized_values
        wake_ok = tone_ok and protocol_ok and ((not ready_required) or ready_ok)
        return {
            "wake_values": wake_values,
            "asr_values": asr_values,
            "recognized_values": recognized_values,
            "tones": tones,
            "protocols": protocols,
            "recognition_clock": recognition_clock,
            "ready_ok": ready_ok,
            "ready_clock": ready_clock,
            "ready_evidence": ready_evidence,
            "ready_required": ready_required,
            "wake_word_ok": wake_word_ok,
            "acceptable_tones": acceptable_tones,
            "tone_ok": tone_ok,
            "mode_active_ok": mode_active_ok,
            "protocol_required": protocol_required,
            "protocol_ok": protocol_ok,
            "signal_evidence": signal_evidence,
            "wake_ok": wake_ok,
        }

    def probe_voice_chain(self) -> bool:
        assert self.reader is not None
        wake_word = self.config["wakeupWord"]
        settle_s = max(float(self.policy.get("silenceObserveS", 2.5)), 2.0)
        idle_wait_s = max(float(self.policy.get("preflightWakeIdleWaitS", 35.0)), 20.0)
        self.clear_observation()
        time.sleep(settle_s)
        self.clear_observation()
        wake_result = self.run_wake(
            {
                "word": wake_word,
                "expect_tone_id": "TONE_ID_0",
                "retries": 2,
                "observe_s": max(float(self.policy.get("commandObserveS", 3.0)), 3.0),
            }
        )
        recent = self.reader.get_recent_lines()[-10:]
        evidence = list(wake_result.get("evidence", [])) + recent
        if wake_result["status"] == "PASS":
            idle_result = self.run_wait(
                {
                    "seconds": idle_wait_s,
                    "expect_tone_id": "TONE_ID_2",
                    "expect_markers": ["TIME_OUT"],
                    "measure_timeout": False,
                    "validate_duration": False,
                }
            )
            evidence.extend(idle_result.get("evidence", []))
            if idle_result["status"] != "PASS":
                self.set_capability("voice", False, f"语音链路预检通过但未回到空闲态：{idle_result['detail']}", evidence[-10:])
                return False
            self.clear_observation()
            self.set_capability("voice", True, f"语音链路预检通过：{wake_result['detail']}", evidence[-10:])
            return True
        self.set_capability("voice", False, f"当前工位未打通 PC播放->设备收音 链路：{wake_result['detail']}", evidence[-10:])
        return False

    def probe_uart_inject(self) -> bool:
        assert self.reader is not None
        assert self.proto is not None
        inject_action = None
        for case_item in self.selected_cases:
            for action in case_item["actions"]:
                if action["type"] == "inject_protocol":
                    inject_action = action
                    break
            if inject_action:
                break
        if not inject_action:
            self.set_capability("uart_inject", True, "本次未选择协议注入类用例")
            return True
        if not self.maybe_recover_logs():
            self.set_capability("uart_inject", False, "日志恢复失败，无法执行协议注入预检")
            return False
        self.clear_observation()
        before = len(self.reader.get_recent_lines())
        protocol = normalize_hex(inject_action["protocol"])
        self.proto.write_hex(protocol)
        time.sleep(float(inject_action.get("observe_s", self.policy["commandObserveS"])))
        recent = self.reader.get_recent_lines()
        tones = self.observed_tones()
        recv_values = self.reader.get_all("recvMsg")
        if tones or recv_values or len(recent) > before:
            self.set_capability("uart_inject", True, "协议注入预检通过", recent[-10:])
            return True
        self.set_capability("uart_inject", False, "协议串口注入后无日志、无播报、无接收协议响应", recent[-10:])
        return False

    def record_serial_check(self, name: str, supported: bool, reason: str, evidence: list[str] | None = None) -> bool:
        payload = {
            "supported": supported,
            "reason": reason,
            "evidence": evidence or [],
        }
        self.serial_checks[name] = payload
        level = self.log.info if supported else self.log.error
        level(f"串口校验 {name}: {'READY' if supported else 'BLOCKED'} | {reason}")
        return supported

    def validate_control_serial(self) -> bool:
        if not self.power_control_supported():
            return self.record_serial_check("ctrl_serial", False, "未配置上下电控制串口")
        cfg = self.config.get("powerControl", {})
        commands = [str(item) for item in (cfg.get("powerOnCmds") or ["uut-switch1.on"]) if str(item).strip()]
        try:
            issued = self._control_commands(commands)
            boot_wait_s = self.configured_power_boot_wait_s()
            if not self.reopen_observers_after_power_event("控制串口上电后刷新观察串口", wait_s=boot_wait_s):
                reason = "控制串口命令执行成功，但上电后日志/协议串口重连失败"
                return self.record_serial_check("ctrl_serial", False, reason, issued)
            reason = f"控制串口命令执行成功，并已等待设备重新就绪 {boot_wait_s:.1f}s"
            return self.record_serial_check("ctrl_serial", True, reason, issued)
        except Exception as exc:
            reason = f"控制串口命令执行失败: {exc}"
            return self.record_serial_check("ctrl_serial", False, reason)

    def validate_log_serial(self) -> bool:
        assert self.reader is not None
        if not self.ensure_logs_available("串口基础校验/日志口"):
            recent = self.reader.get_recent_lines()[-10:]
            return self.record_serial_check("log_serial", False, "日志串口未观察到有效日志回显", recent)
        self.log.info("串口基础校验/日志口：主动发送 loglevel 4，确保详细日志开启")
        if not self.set_log_level():
            recent = self.reader.get_recent_lines()[-10:]
            return self.record_serial_check("log_serial", False, "日志串口已连接，但 loglevel 4 设置失败", recent)
        recent = self.reader.get_recent_lines()[-10:]
        return self.record_serial_check("log_serial", True, "日志串口可正常收发并已开启详细日志", recent)

    def find_protocol_validation_action(self) -> tuple[dict[str, Any], str] | None:
        preferred_words = ("恢复出厂模式", "打开窗帘", "关闭窗帘", "小声点")
        preferred_normal: tuple[dict[str, Any], str] | None = None
        preferred_setup: tuple[dict[str, Any], str] | None = None
        fallback_normal: tuple[dict[str, Any], str] | None = None
        fallback_setup: tuple[dict[str, Any], str] | None = None
        for case_item in self.selected_cases:
            for action in case_item["actions"]:
                if action.get("type") != "say":
                    continue
                action_expected = normalize_hex(
                    action.get("expect_send_protocol") or action.get("expected_protocol") or ""
                )
                if action_expected and action_expected != "-":
                    is_setup = bool(action.get("setup_action")) or bool(action.get("always_run"))
                    word = str(action.get("word") or action.get("text") or "").strip()
                    is_preferred = any(word == item for item in preferred_words)
                    candidate = (action, action_expected)
                    if is_preferred and not is_setup and preferred_normal is None:
                        preferred_normal = candidate
                    elif is_preferred and is_setup and preferred_setup is None:
                        preferred_setup = candidate
                    elif not is_setup and fallback_normal is None:
                        fallback_normal = candidate
                    elif is_setup and fallback_setup is None:
                        fallback_setup = candidate
        return preferred_normal or preferred_setup or fallback_normal or fallback_setup

    def validate_protocol_serial(self) -> bool:
        assert self.reader is not None
        if not self.proto:
            return self.record_serial_check("protocol_serial", False, "未配置协议串口监控")
        match = self.find_protocol_validation_action()
        if not match:
            return self.record_serial_check("protocol_serial", True, "本次未选中需要协议口观测的发送协议用例")
        action, expected = match
        if not self.maybe_recover_logs():
            return self.record_serial_check("protocol_serial", False, "日志恢复失败，无法执行协议串口预检")
        word = str(action.get("word") or action.get("text") or "").strip()
        if not word:
            return self.record_serial_check("protocol_serial", False, "未找到可用于协议串口预检的命令词")
        wake_word = str(self.config.get("wakeupWord") or "").strip()
        observe_s = float(action.get("observe_s", self.policy["commandObserveS"]))
        post_wake_gap_s = max(float(self.policy.get("postWakeGapS", 2.0)), 2.0)
        attempts = max(1, int(self.policy.get("commandRetries", 3)))
        result: dict[str, Any] = {}
        protocols: list[str] = []
        for attempt in range(attempts):
            if wake_word:
                if not self.maybe_recover_logs():
                    return self.record_serial_check("protocol_serial", False, "日志恢复失败，无法执行协议串口预检")
                self.clear_observation()
                if not self.play_word(wake_word):
                    result = self._make_action_result("FAIL", "say", f"协议预检前唤醒词播放失败：{wake_word}", time.perf_counter())
                    break
                time.sleep(post_wake_gap_s)
            result = self.run_say(
                {
                    "word": word,
                    "expect_send_protocol": expected,
                    "retries": 1,
                    "observe_s": observe_s,
                    "require_protocol_and_tone": False,
                    "auto_wake": False,
                }
            )
            protocols = result.get("actual_protocols", [])
            if protocols:
                break
            if attempt + 1 < attempts:
                self.log.warn(f"协议串口预检未命中期望协议，准备重试 {attempt + 1}/{attempts - 1}")
                time.sleep(0.5)
        recent = self.reader.get_recent_lines()[-10:]
        matched = bool(protocols)
        expected_hit = any(expected in item for item in protocols) if expected else matched
        detail = str(result.get("detail", "")).strip()
        if matched:
            reason = f"播放 {word} 后已观察到协议串口实收数据"
            if result.get("status") != "PASS" and detail:
                reason += f"（动作摘要：{detail}）"
        elif result.get("status") != "PASS" and detail:
            reason = f"播放 {word} 后未稳定观察到协议串口实收数据：{detail}"
        else:
            reason = f"播放 {word} 后未观察到协议串口实收数据"
        if matched and expected and not expected_hit:
            reason += f"；当前仅确认协议串口链路可观测，最近实收与期望 {expected} 不一致，实际为 {'/'.join(protocols[-3:]) or '-'}"
        evidence = (protocols[-5:] + recent)[-10:]
        return self.record_serial_check("protocol_serial", matched, reason, evidence)

    def validate_serial_links(self) -> bool:
        checks = [
            self.validate_control_serial(),
            self.validate_log_serial(),
        ]
        return all(checks)

    def run_preflights(self) -> None:
        if not self.validate_serial_links():
            self.stop_reason = "串口基础校验失败，停止测试"
            return
        if not self.validate_protocol_serial():
            self.stop_reason = "串口基础校验失败，停止测试"
            return
        if any("voice" in self.case_dependencies(case_item) for case_item in self.selected_cases):
            self.log.info("开始语音链路预检")
            if not self.probe_voice_chain():
                self.stop_reason = "语音链路预检失败，停止测试"
                return
        if any("uart_inject" in self.case_dependencies(case_item) for case_item in self.selected_cases):
            self.log.info("开始协议注入链路预检")
            self.probe_uart_inject()
        if any("manual_power" in self.case_dependencies(case_item) for case_item in self.selected_cases):
            if self.power_control_supported():
                self.set_capability("manual_power", True, "已接入控制口上下电/boot 控制接口")
            else:
                self.set_capability("manual_power", False, "未接入上下电控制接口")

    def set_log_level(self) -> bool:
        assert self.reader is not None
        retries = int(self.policy.get("setLogLevelRetries", 5))
        command = self.policy["loglevelCommand"]
        for index in range(retries):
            before_lines = self.reader.get_recent_lines()
            self.log.info(f"设置日志等级 {command}，第 {index + 1}/{retries} 次")
            self.reader.write(command)
            deadline = time.time() + 2.5
            while time.time() < deadline:
                time.sleep(0.2)
                after_lines = self.reader.get_recent_lines()
                if len(after_lines) > len(before_lines):
                    self.log.info("日志等级设置后已观察到新日志")
                    return True
                recent = after_lines[-5:]
                if any(command in line or "root:/$" in line for line in recent):
                    self.log.info("日志等级设置后已观察到控制台回显")
                    return True
            if index + 1 < retries:
                self.log.warn("未观察到新日志，继续重试 loglevel 4")
        self.log.error("loglevel 4 连续 5 次均失败")
        return False

    def logs_available(self) -> bool:
        assert self.reader is not None
        recent = self.reader.get_recent_lines()[-30:]
        return any(line.strip() for line in recent)

    def ensure_logs_available(self, reason: str) -> bool:
        time.sleep(1.0)
        if self.logs_available():
            self.log.info(f"{reason}：当前日志已可用，不额外发送 loglevel 4")
            return True
        self.log.warn(f"{reason}：当前未观察到日志，开始 loglevel 4 重试")
        return self.set_log_level()

    def recover_logs_after_reboot(self, reason: str, fatal: bool = False) -> bool:
        assert self.reader is not None
        time.sleep(1.0)
        self.log.info(f"{reason}：检测到重启后强制恢复 loglevel 4")
        if not self.set_log_level():
            detail = f"{reason}：重启后 loglevel 4 恢复失败"
            self.log.error(detail)
            if fatal:
                self.stop_reason = detail
            return False
        time.sleep(1.0)
        if not self.logs_available():
            detail = f"{reason}：loglevel 4 设置后仍未观察到日志输出"
            self.log.error(detail)
            if fatal:
                self.stop_reason = detail
            return False
        self.log.info(f"{reason}：已验证日志输出恢复，允许继续测试")
        return True

    def sync_reboot_baseline(self, reason: str = "") -> int:
        if not self.reader:
            return 0
        current = self.reader.get_reboot_count()
        self.reboot_count_baseline = current
        self.reader.clear_reboot_flag()
        if reason:
            self.log.info(f"{reason}：重启计数基线同步为 {current}")
        return current

    def check_unexpected_reboot(self, reason: str) -> bool:
        if not self.reader:
            return True
        current_reboot = self.reader.get_reboot_count()
        reboot_flag = self.reader.is_rebooted()
        if current_reboot <= self.reboot_count_baseline and not reboot_flag:
            return True
        if current_reboot > self.reboot_count_baseline:
            delta = current_reboot - self.reboot_count_baseline
            self.reboot_count_baseline = current_reboot
            if not self.note_unexpected_reboot(f"{reason}检测到 {delta} 次异常重启"):
                self.reader.clear_reboot_flag()
                return False
            if not self.recover_logs_after_reboot(f"{reason}异常重启后恢复日志", fatal=True):
                self.reader.clear_reboot_flag()
                return False
            self.sync_reboot_baseline(f"{reason}异常重启已处理")
            return True
        self.log.warn(f"{reason}检测到重启标记，按重启恢复流程强制设置 loglevel 4")
        if not self.recover_logs_after_reboot(f"{reason}检测到重启标记后恢复日志", fatal=True):
            self.reader.clear_reboot_flag()
            return False
        self.sync_reboot_baseline(f"{reason}重启标记已处理")
        return True

    def do_reboot(self) -> bool:
        assert self.reader is not None
        retries = int(self.policy.get("rebootRetries", 5))
        command = self.policy["rebootCommand"]
        for index in range(retries):
            before_count = self.reader.get_reboot_count()
            self.log.info(f"执行重启 {command}，第 {index + 1}/{retries} 次")
            self.reader.write(command)
            deadline = time.time() + 5.0
            while time.time() < deadline:
                time.sleep(0.2)
                if self.reader.get_reboot_count() > before_count or self.reader.is_rebooted():
                    self.sync_reboot_baseline("主动 reboot 已观察到重启，忽略本次预期重启")
                    self.log.info("已观察到有效重启日志")
                    return True
            recent = self.reader.get_recent_lines()[-20:]
            if any("reboot" in line.lower() or "RESET=" in line for line in recent):
                self.sync_reboot_baseline("主动 reboot 已观察到重启证据，忽略本次预期重启")
                self.log.info("已从最近日志中观察到重启证据")
                return True
        self.log.error("reboot 连续重试后仍未观察到有效重启日志")
        return False

    def maybe_recover_logs(self) -> bool:
        assert self.reader is not None
        if self.logs_available():
            return True
        self.log.warn("日志不活跃，按需补发 loglevel 4")
        return self.set_log_level()

    def note_unexpected_reboot(self, reason: str) -> bool:
        self.unexpected_reboot_count += 1
        self.log.warn(f"检测到设备异常重启，第 {self.unexpected_reboot_count}/3 次：{reason}")
        if self.unexpected_reboot_count > 3:
            self.stop_reason = f"设备异常重启超过 3 次：{reason}"
            self.log.error(self.stop_reason)
            return False
        return True

    def current_case_unexpected_reboot_detail(self, reason: str) -> str:
        if not self.reader:
            return ""
        current_reboot = self.reader.get_reboot_count()
        reboot_flag = self.reader.is_rebooted()
        if current_reboot <= self.reboot_count_baseline and not reboot_flag:
            return ""
        reboot_reason = str(self.reader.get("rebootReason") or self.reader.get("reboot") or "unknown").strip()
        delta = current_reboot - self.reboot_count_baseline if current_reboot > self.reboot_count_baseline else 1
        detail = f"{reason}检测到设备异常重启 {delta} 次（Boot Reason: {reboot_reason}），按规则判当前用例失败"
        allowed = self.note_unexpected_reboot(detail)
        if allowed:
            if not self.recover_logs_after_reboot(f"{reason}异常重启后恢复日志", fatal=True):
                self.reader.clear_reboot_flag()
                if self.stop_reason and self.stop_reason not in detail:
                    return f"{detail}；{self.stop_reason}"
                return detail
            self.sync_reboot_baseline(f"{reason}异常重启已处理")
        else:
            self.reader.clear_reboot_flag()
        if self.stop_reason and self.stop_reason not in detail and self.unexpected_reboot_count > 3:
            return f"{detail}；{self.stop_reason}"
        return detail

    def note_protocol_truncation(self, partial_frames: list[str]) -> bool:
        self.protocol_truncation_count += 1
        preview = partial_frames[-1] if partial_frames else ""
        self.log.warn(f"检测到协议截断，第 {self.protocol_truncation_count}/2 次：{preview}")
        if self.protocol_truncation_count > 2:
            self.stop_reason = f"串口日志协议截断超过 2 次：{preview}"
            self.log.error(self.stop_reason)
            return False
        return True

    def consecutive_wake_failure_limit(self) -> int:
        return max(0, int(self.policy.get("consecutiveWakeFailureLimit", 3) or 3))

    def is_wake_failure_result(self, result: dict[str, Any]) -> bool:
        if result.get("status") != "FAIL":
            return False
        parts = [str(result.get("detail", ""))]
        for action in result.get("actions", []):
            parts.append(str(action.get("detail", "")))
        combined = " | ".join(part for part in parts if part)
        return "唤醒失败" in combined

    def note_consecutive_wake_failures(self, result: dict[str, Any]) -> bool:
        limit = self.consecutive_wake_failure_limit()
        if limit <= 0:
            return True
        if self.is_wake_failure_result(result):
            self.consecutive_wake_failure_count += 1
            self.log.warn(
                f"检测到连续唤醒失败，第 {self.consecutive_wake_failure_count}/{limit} 条："
                f"{result.get('case_id', '')} {result.get('detail', '')}"
            )
            if self.consecutive_wake_failure_count >= limit:
                self.stop_reason = (
                    f"连续 {limit} 条用例唤醒失败，停止测试："
                    f"{result.get('case_id', '')} {result.get('detail', '')}"
                )
                self.log.error(self.stop_reason)
                return False
            return True
        if self.consecutive_wake_failure_count:
            self.log.info(
                f"{result.get('case_id', '')} 未出现唤醒失败，连续唤醒失败计数清零（此前 {self.consecutive_wake_failure_count} 条）"
            )
        self.consecutive_wake_failure_count = 0
        return True

    def _read_serial_tail(self, last_n: int = 200) -> list[str]:
        path = self.result_dir / self.policy["serialLogName"]
        if not path.is_file():
            return []
        with path.open("r", encoding="utf-8", errors="ignore") as fp:
            return [line.rstrip("\n") for line in deque(fp, maxlen=last_n)]

    def _clock_from_serial_line(self, line: str) -> float | None:
        match = re.match(r"\[(\d{2}):(\d{2}):(\d{2})\.(\d{3})\]", line)
        if not match:
            return None
        hour, minute, second, millis = (int(part) for part in match.groups())
        return hour * 3600 + minute * 60 + second + millis / 1000.0

    def _now_clock(self) -> float:
        now = dt.datetime.now()
        return now.hour * 3600 + now.minute * 60 + now.second + now.microsecond / 1_000_000.0

    def _clock_is_after(self, clock: float | None, after_clock: float | None) -> bool:
        if clock is None or after_clock is None:
            return clock is not None
        if clock + 0.05 >= after_clock:
            return True
        return after_clock - clock > 12 * 3600

    def _find_latest_serial_clock(self, markers: list[str], last_n: int = 200, after_clock: float | None = None) -> float | None:
        for line in reversed(self._read_serial_tail(last_n)):
            if any(marker in line for marker in markers):
                clock = self._clock_from_serial_line(line)
                if self._clock_is_after(clock, after_clock):
                    return clock
        return None

    def _find_first_serial_clock(self, markers: list[str], last_n: int = 200, after_clock: float | None = None) -> float | None:
        for line in self._read_serial_tail(last_n):
            if any(marker in line for marker in markers):
                clock = self._clock_from_serial_line(line)
                if self._clock_is_after(clock, after_clock):
                    return clock
        return None

    def _clock_diff(self, later_clock: float | None, earlier_clock: float | None) -> float | None:
        if later_clock is None or earlier_clock is None:
            return None
        diff = later_clock - earlier_clock
        if diff < -12 * 3600:
            diff += 24 * 3600
        if diff < 0:
            diff = 0.0
        return round(diff, 3)

    def _shift_clock(self, clock: float | None, delta_s: float) -> float | None:
        if clock is None:
            return None
        shifted = clock + delta_s
        while shifted < 0:
            shifted += 24 * 3600
        while shifted >= 24 * 3600:
            shifted -= 24 * 3600
        return shifted

    def describe_action(self, action: dict[str, Any]) -> str:
        custom = str(action.get("label") or "").strip()
        if custom:
            return custom
        action_type = action["type"]
        if action_type == "wake":
            return f"播放唤醒词【{action['word']}】"
        if action_type == "say":
            if action.get("expect_no_response"):
                return f"播放负例词【{action['word']}】并确认无播报/无协议"
            return f"播放命令词【{action['word']}】"
        if action_type == "wait":
            return f"等待 {action['seconds']}s"
        if action_type == "reboot":
            return "发送 reboot 并等待设备恢复"
        if action_type == "inject_protocol":
            return f"注入协议【{action['protocol']}】"
        if action_type == "volume_walk":
            return f"执行音量遍历【{action['command_word']}】"
        if action_type == "assert_wake_repeats":
            return f"连续唤醒校验【{action['word']}】x{action['repeats']}"
        if action_type == "assert_no_wake":
            return f"负例唤醒校验【{action['word']}】x{action.get('repeats', 1)}"
        if action_type == "phrase_check":
            return f"词条检查【{len(action.get('items', []))}条】"
        if action_type == "manual_power_cycle":
            return "人工上下电"
        if action_type == "log_probe":
            return f"校验日志字段【{', '.join(action.get('expected_keys', []))}】"
        return json.dumps(action, ensure_ascii=False)

    def _latest_play_id_clock(
        self,
        tone_id: str,
        fallback_markers: list[str] | None = None,
        after_clock: float | None = None,
    ) -> float | None:
        markers: list[str] = []
        tone_num = tone_id_number(tone_id)
        if tone_num:
            markers.extend([f"play id : {tone_num}", f"play id:{tone_num}", f"play id :{tone_num}"])
        if fallback_markers:
            markers.extend(fallback_markers)
        return self._find_latest_serial_clock(markers, after_clock=after_clock)

    def _make_action_result(self, status: str, action_type: str, detail: str, start: float, **extra: Any) -> dict[str, Any]:
        result = {
            "status": status,
            "action_type": action_type,
            "action_label": extra.get("action_label", ""),
            "detail": detail,
            "duration_s": round(time.perf_counter() - start, 3),
            "recognition_latency_s": extra.get("recognition_latency_s"),
            "actual_asr": extra.get("actual_asr", []),
            "actual_tones": extra.get("actual_tones", []),
            "actual_protocols": extra.get("actual_protocols", []),
            "actual_log_values": extra.get("actual_log_values", {}),
            "pending_advisory_logs": extra.get("pending_advisory_logs", []),
            "evidence": extra.get("evidence", []),
            "reference_clock": extra.get("reference_clock"),
        }
        self.last_action = result
        return result

    def run_wake(self, action: dict[str, Any]) -> dict[str, Any]:
        assert self.reader is not None
        start = time.perf_counter()
        word = action["word"]
        expect_tone = parse_tone_id(action.get("expect_tone_id", ""))
        require_any_tone = bool(action.get("require_any_tone"))
        expected_protocol = normalize_hex(action.get("expect_send_protocol", ""))
        require_protocol_when_configured = bool(
            action.get("require_protocol_when_configured", self.policy.get("requireWakeProtocolWhenConfigured", True))
        )
        require_ready_signal = bool(
            action.get("require_ready_signal", self.policy.get("requireWakeReadySignal", True))
        )
        retries = int(action.get("retries", self.policy["wakeRetries"]))
        observe_s = float(action.get("observe_s", self.policy["commandObserveS"]))
        observe_tail_s = float(action.get("observe_tail_s", self.policy.get("wakeObserveTailS", 15.0)))
        observe_hard_limit_s = float(action.get("observe_hard_limit_s", self.policy.get("wakeObserveHardLimitS", 20.0)))
        last_recognition_latency: float | None = None
        wake_values: list[str] = []
        tones: list[str] = []
        protocols: list[str] = []
        last_reference_clock: float | None = None
        last_evidence: list[str] = []
        last_detail = f"唤醒失败：{word}"
        retry_gate_clock = self.last_action.get("reference_clock") if self.last_action else None
        for index in range(retries):
            if retry_gate_clock is not None:
                gate_ok, gate_evidence = self.wait_for_response_playback_complete(
                    retry_gate_clock,
                    context=f"唤醒词【{word}】第 {index + 1} 次前",
                )
                if not gate_ok:
                    self.log.warn(f"唤醒词【{word}】第 {index + 1} 次前仍未确认播报结束：{'/'.join(gate_evidence) or '-'}")
                retry_gate_clock = None
            if not self.maybe_recover_logs():
                return self._make_action_result("FAIL", "wake", "日志恢复失败", start)
            mode_snapshot = self.current_work_mode()
            self.clear_observation()
            if not self.play_word(word):
                return self._make_action_result("FAIL", "wake", f"唤醒音频播放失败：{word}", start)
            play_end_clock = self._now_clock()
            match_after_clock = self._shift_clock(play_end_clock, -2.0)
            wake_eval: dict[str, Any] | None = None
            observe_start = time.time()
            base_deadline = observe_start + max(observe_s, 1.8)
            hard_deadline = observe_start + max(max(observe_s, 1.8), observe_hard_limit_s)
            tail_deadline: float | None = None
            while time.time() < min(tail_deadline or base_deadline, hard_deadline):
                time.sleep(0.2)
                wake_eval = self._evaluate_wake_signal(
                    word=word,
                    expect_tone=expect_tone,
                    require_any_tone=require_any_tone,
                    expected_protocol=expected_protocol,
                    mode_snapshot=mode_snapshot,
                    match_after_clock=match_after_clock,
                    require_protocol_when_configured=require_protocol_when_configured,
                    require_ready_signal=require_ready_signal,
                )
                wake_values = wake_eval["wake_values"]
                recognized_values = wake_eval["recognized_values"]
                tones = wake_eval["tones"]
                protocols = wake_eval["protocols"]
                recognition_clock = wake_eval["recognition_clock"]
                last_recognition_latency = self._clock_diff(recognition_clock, play_end_clock)
                last_evidence = wake_eval["signal_evidence"]
                partial_match = bool(
                    wake_eval["wake_word_ok"]
                    or wake_eval["ready_ok"]
                    or wake_eval["mode_active_ok"]
                    or wake_eval["protocol_ok"] and wake_eval["protocol_required"]
                    or (
                        wake_eval["acceptable_tones"]
                        and any(tone in tones for tone in wake_eval["acceptable_tones"])
                    )
                )
                if partial_match:
                    tail_deadline = min(
                        hard_deadline,
                        max(tail_deadline or base_deadline, time.time() + max(observe_tail_s, 0.5)),
                    )
                if wake_eval["wake_ok"]:
                    reference_clock = (
                        wake_eval["ready_clock"]
                        or self._latest_play_id_clock(expect_tone or (tones[-1] if tones else ""), ["Wakeup:"], after_clock=match_after_clock)
                        or recognition_clock
                    )
                    last_reference_clock = reference_clock
                    return self._make_action_result(
                        "PASS",
                        "wake",
                        f"唤醒成功，第 {index + 1} 次",
                        start,
                        recognition_latency_s=last_recognition_latency,
                        actual_asr=recognized_values,
                        actual_tones=tones,
                        actual_protocols=protocols,
                        evidence=wake_eval["signal_evidence"],
                        reference_clock=reference_clock,
                    )
            if wake_eval is None:
                wake_eval = self._evaluate_wake_signal(
                    word=word,
                    expect_tone=expect_tone,
                    require_any_tone=require_any_tone,
                    expected_protocol=expected_protocol,
                    mode_snapshot=mode_snapshot,
                    match_after_clock=match_after_clock,
                    require_protocol_when_configured=require_protocol_when_configured,
                    require_ready_signal=require_ready_signal,
                )
                wake_values = wake_eval["wake_values"]
                recognized_values = wake_eval["recognized_values"]
                tones = wake_eval["tones"]
                protocols = wake_eval["protocols"]
                recognition_clock = wake_eval["recognition_clock"]
                last_recognition_latency = self._clock_diff(recognition_clock, play_end_clock)
                last_evidence = wake_eval["signal_evidence"]
            mismatch_parts = []
            if not wake_eval["wake_word_ok"]:
                mismatch_parts.append(f"唤醒词不符(期望 {word} / 实际 {'/'.join(recognized_values) or '-'})")
            if wake_eval["acceptable_tones"] and not wake_eval["tone_ok"]:
                mismatch_parts.append(
                    f"唤醒播报不符(当前workMode={mode_snapshot or '-'} / 期望 {'/'.join(wake_eval['acceptable_tones'])} / 实际 {'/'.join(tones) or '-'})"
                )
            if wake_eval["protocol_required"] and not wake_eval["protocol_ok"]:
                mismatch_parts.append(f"唤醒协议不符(期望 {expected_protocol} / 实际 {'/'.join(protocols) or '-'})")
            if wake_eval["ready_required"] and not wake_eval["ready_ok"]:
                mismatch_parts.append("未观察到唤醒就绪日志")
            if mismatch_parts:
                last_detail = f"唤醒失败：{word}；" + "；".join(mismatch_parts)
            retry_gate_clock = play_end_clock
        return self._make_action_result(
            "FAIL",
            "wake",
            last_detail,
            start,
            recognition_latency_s=last_recognition_latency,
            actual_asr=wake_eval["recognized_values"] if wake_eval else self.reader.get_all("wakeKw"),
            actual_tones=self.observed_tones(),
            actual_protocols=self.observed_protocols(),
            evidence=last_evidence,
            reference_clock=last_reference_clock,
        )

    def run_say(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        word = action["word"]
        observe_s = float(action.get("observe_s", self.policy["commandObserveS"]))
        observe_floor_s = float(action.get("observe_floor_s", self.policy.get("commandObserveFloorS", 6.0)))
        expect_no_response = bool(action.get("expect_no_response"))
        auto_wake = bool(action.get("auto_wake", True))
        wake_word = str(action.get("wake_word") or self.config["wakeupWord"]).strip() or self.config["wakeupWord"]
        expected_asr = set(action.get("expect_asr", []))
        expected_tone = parse_tone_id(action.get("expect_tone_id", ""))
        advisory_tone = parse_tone_id(action.get("advisory_tone_id", ""))
        expect_any_tone = bool(action.get("expect_any_tone", False))
        expected_protocol = normalize_hex(action.get("expect_send_protocol", ""))

        def normalize_log_expectations(raw_values: dict[str, Any]) -> dict[str, list[str]]:
            return {
                key: [str(item).strip() for item in (values if isinstance(values, list) else [values]) if str(item).strip()]
                for key, values in raw_values.items()
            }

        expected_logs = normalize_log_expectations(action.get("expect_log_values", {}))
        advisory_logs = normalize_log_expectations(action.get("advisory_log_values", {}))
        forbidden_logs = normalize_log_expectations(action.get("forbid_log_values", {}))
        require_protocol_and_tone = bool(action.get("require_protocol_and_tone", True))
        config_saved_values = expected_logs.pop("configSaved", [])
        if config_saved_values:
            advisory_logs.setdefault("configSaved", [])
            for item in config_saved_values:
                if item not in advisory_logs["configSaved"]:
                    advisory_logs["configSaved"].append(item)

        retries = int(action.get("retries", 1 if (expected_logs or expected_protocol) else self.policy["commandRetries"]))
        default_hard_limit_s = 80.0 if expected_logs else 35.0
        observe_tail_s = float(action.get("observe_tail_s", self.policy.get("commandObserveTailS", 20.0)))
        observe_hard_limit_s = float(
            action.get("observe_hard_limit_s", self.policy.get("commandObserveHardLimitS", default_hard_limit_s))
        )
        advisory_log_wait_s = float(action.get("advisory_log_wait_s", self.policy.get("advisoryLogWaitS", 8.0)))
        advisory_tone_wait_s = float(
            action.get(
                "advisory_tone_wait_s",
                self.policy.get("advisoryToneWaitS", 0.0),
            )
        )
        secondary_signal_wait_s = float(
            action.get(
                "secondary_signal_wait_s",
                self.policy.get("secondarySignalWaitS", 0.0),
            )
        )
        default_settle_s = 0.0 if expect_no_response else 2.0
        success_settle_s = float(action.get("success_settle_s", self.policy.get("successSettleS", default_settle_s)))
        all_asr: list[str] = []
        all_tones: list[str] = []
        all_protocols: list[str] = []
        combined_log_keys = list(expected_logs.keys())
        combined_log_keys.extend(key for key in advisory_logs.keys() if key not in combined_log_keys)
        combined_log_keys.extend(key for key in forbidden_logs.keys() if key not in combined_log_keys)
        all_log_values: dict[str, list[str]] = {key: [] for key in combined_log_keys}
        last_recognition_latency: float | None = None
        last_reference_clock: float | None = None
        retry_gate_clock = self.last_action.get("reference_clock") if self.last_action else None

        for index in range(retries):
            pending_pass_reference_clock: float | None = None
            pending_pass_deadline: float | None = None
            pending_pass_reason = ""
            if retry_gate_clock is not None:
                gate_ok, gate_evidence = self.wait_for_response_playback_complete(
                    retry_gate_clock,
                    context=f"命令【{word}】第 {index + 1} 次前",
                )
                if not gate_ok:
                    self.log.warn(f"命令【{word}】第 {index + 1} 次前仍未确认播报结束：{'/'.join(gate_evidence) or '-'}")
                retry_gate_clock = None
            if auto_wake:
                wake_result = self.run_wake(
                    {
                        "word": wake_word,
                        "expect_tone_id": action.get("wake_expect_tone_id", "TONE_ID_0"),
                        "require_any_tone": bool(action.get("wake_require_any_tone", False)),
                        "retries": int(action.get("wake_retries", self.policy["wakeRetries"])),
                    }
                )
                if wake_result["status"] != "PASS":
                    all_asr.extend(item for item in wake_result.get("actual_asr", []) if item not in all_asr)
                    all_tones.extend(item for item in wake_result.get("actual_tones", []) if item not in all_tones)
                    all_protocols.extend(item for item in wake_result.get("actual_protocols", []) if item not in all_protocols)
                    return self._make_action_result(
                        "FAIL",
                        "say",
                        f"命令词执行前唤醒失败：{word}",
                        start,
                        actual_asr=all_asr,
                        actual_tones=all_tones,
                        actual_protocols=all_protocols,
                        actual_log_values=all_log_values,
                        evidence=[wake_result.get("detail", "")],
                        reference_clock=wake_result.get("reference_clock"),
                    )
                wake_gate_ok, wake_gate_evidence = self.wait_for_response_playback_complete(
                    wake_result.get("reference_clock"),
                    context=f"命令【{word}】前等待唤醒播报结束",
                )
                if not wake_gate_ok:
                    self.log.warn(f"命令【{word}】前仍未确认唤醒播报结束：{'/'.join(wake_gate_evidence) or '-'}")
                post_wake_gap_s = max(float(action.get("post_wake_gap_s", self.policy.get("postPlaybackGapS", 0.2))), 0.0)
                if post_wake_gap_s > 0:
                    time.sleep(post_wake_gap_s)
            elif not self.maybe_recover_logs():
                return self._make_action_result("FAIL", "say", "日志恢复失败", start, actual_log_values=all_log_values)

            self.clear_observation()
            if not self.play_word(word):
                return self._make_action_result("FAIL", "say", f"播放命令词失败：{word}", start, actual_log_values=all_log_values)
            play_end_clock = self._now_clock()
            match_after_clock = self._shift_clock(play_end_clock, -2.0)
            recognition_clock: float | None = None
            observe_window_s = observe_s if expect_no_response else max(observe_s, observe_floor_s)
            observe_start = time.time()
            base_deadline = observe_start + observe_window_s
            hard_deadline = observe_start + max(observe_window_s, observe_hard_limit_s)
            tail_deadline: float | None = None
            while time.time() < min(tail_deadline or base_deadline, hard_deadline):
                time.sleep(0.2)
                asr_values = self.observed_recognized_values()
                tones = self.observed_tones()
                protocols = self.observed_protocols()
                recognition_clock = self._find_latest_serial_clock(["Wakeup:", "keyword:"], after_clock=match_after_clock)
                last_recognition_latency = self._clock_diff(recognition_clock, play_end_clock)
                all_asr.extend(item for item in asr_values if item not in all_asr)
                all_tones.extend(item for item in tones if item not in all_tones)
                all_protocols.extend(item for item in protocols if item not in all_protocols)
                matched_logs: dict[str, list[str]] = {}
                for key in expected_logs:
                    current_values = self.observed_log_values(key)
                    all_log_values[key].extend(item for item in current_values if item not in all_log_values[key])
                    matched_logs[key] = [item for item in current_values if item in expected_logs[key]]
                matched_advisory_logs: dict[str, list[str]] = {}
                for key in advisory_logs:
                    current_values = self.observed_log_values(key)
                    all_log_values[key].extend(item for item in current_values if item not in all_log_values[key])
                    matched_advisory_logs[key] = [item for item in current_values if item in advisory_logs[key]]
                forbidden_hits: dict[str, list[str]] = {}
                for key in forbidden_logs:
                    current_values = self.observed_log_values(key)
                    all_log_values[key].extend(item for item in current_values if item not in all_log_values[key])
                    forbidden_hits[key] = [item for item in current_values if item in forbidden_logs[key]]
                if any(forbidden_hits.values()):
                    hit_parts = [
                        f"{key}={'/'.join(values)}"
                        for key, values in forbidden_hits.items()
                        if values
                    ]
                    return self._make_action_result(
                        "FAIL",
                        "say",
                        f"{'静默断言失败' if expect_no_response else '命令失败'}：{word}；出现禁止日志 {'；'.join(hit_parts)}",
                        start,
                        recognition_latency_s=last_recognition_latency,
                        actual_asr=all_asr,
                        actual_tones=all_tones,
                        actual_protocols=all_protocols,
                        actual_log_values=all_log_values,
                        reference_clock=recognition_clock,
                        evidence=[json.dumps(all_log_values, ensure_ascii=False)] if all_log_values else [],
                    )
                if expect_no_response:
                    if tones or protocols:
                        break
                    continue
                asr_ok = not expected_asr or bool(expected_asr.intersection(asr_values))
                tone_requested = bool(expected_tone) or expect_any_tone
                tone_ok = expected_tone in tones if expected_tone else (bool(tones) if expect_any_tone else True)
                proto_ok = not expected_protocol or expected_protocol in protocols
                protocol_requested = bool(expected_protocol)
                log_ok = all(bool(set(values).intersection(all_log_values.get(key, []))) for key, values in expected_logs.items())
                advisory_log_ok = all(bool(set(values).intersection(all_log_values.get(key, []))) for key, values in advisory_logs.items())
                if require_protocol_and_tone:
                    signal_ok = tone_ok and proto_ok
                elif tone_requested and protocol_requested:
                    # When a command declares an expected send-protocol, the
                    # protocol UART is the primary assertion source. Tone stays
                    # as a supplementary signal unless the case explicitly
                    # requires both protocol and tone together.
                    signal_ok = proto_ok
                elif tone_requested:
                    signal_ok = tone_ok
                elif protocol_requested:
                    signal_ok = proto_ok
                else:
                    signal_ok = True
                partial_match = bool(
                    (expected_asr and expected_asr.intersection(asr_values))
                    or (expected_tone and expected_tone in tones)
                    or (advisory_tone and advisory_tone in tones)
                    or (expect_any_tone and tones)
                    or (expected_protocol and expected_protocol in protocols)
                    or any(matched_logs.values())
                    or any(matched_advisory_logs.values())
                    or (
                        not expected_asr
                        and not expected_tone
                        and not expect_any_tone
                        and not expected_protocol
                        and not expected_logs
                        and not advisory_logs
                        and (asr_values or tones or protocols)
                    )
                )
                if partial_match:
                    tail_deadline = min(
                        hard_deadline,
                        max(tail_deadline or base_deadline, time.time() + max(observe_tail_s, 0.5)),
                    )
                verdict_ok = asr_ok and log_ok and signal_ok
                if verdict_ok:
                    reference_clock = self._latest_play_id_clock(
                        expected_tone or (all_tones[-1] if all_tones else ""),
                        ["Wakeup:"],
                        after_clock=match_after_clock,
                    )
                    if reference_clock is None and protocols:
                        reference_clock = self._find_latest_serial_clock(
                            ["send msg::", "receive msg::"],
                            after_clock=match_after_clock,
                        )
                    if reference_clock is None:
                        reference_clock = recognition_clock
                    last_reference_clock = reference_clock
                    pending_advisory_logs = [
                        key for key, values in advisory_logs.items() if not set(values).intersection(all_log_values.get(key, []))
                    ]
                    pending_advisory_tones = []
                    if advisory_tone and advisory_tone not in all_tones:
                        pending_advisory_tones.append(advisory_tone)
                    pending_secondary_signals: list[str] = []
                    if tone_requested and protocol_requested and not require_protocol_and_tone:
                        if not tone_ok:
                            pending_secondary_signals.append(f"播报:{expected_tone or 'ANY'}")
                        if not proto_ok:
                            pending_secondary_signals.append(f"协议:{expected_protocol}")
                    if advisory_logs and not advisory_log_ok and advisory_log_wait_s > 0:
                        pending_pass_reference_clock = reference_clock
                        pending_pass_reason = "advisory_logs"
                        pending_pass_deadline = min(
                            hard_deadline,
                            max(pending_pass_deadline or base_deadline, time.time() + advisory_log_wait_s),
                        )
                        tail_deadline = min(
                            hard_deadline,
                            max(tail_deadline or base_deadline, pending_pass_deadline),
                        )
                        continue
                    if pending_advisory_tones and advisory_tone_wait_s > 0:
                        pending_pass_reference_clock = reference_clock
                        pending_pass_reason = "advisory_tones"
                        pending_pass_deadline = min(
                            hard_deadline,
                            max(pending_pass_deadline or base_deadline, time.time() + advisory_tone_wait_s),
                        )
                        tail_deadline = min(
                            hard_deadline,
                            max(tail_deadline or base_deadline, pending_pass_deadline),
                        )
                        continue
                    if pending_secondary_signals and secondary_signal_wait_s > 0:
                        pending_pass_reference_clock = reference_clock
                        pending_pass_reason = "secondary_signals"
                        pending_pass_deadline = min(
                            hard_deadline,
                            max(pending_pass_deadline or base_deadline, time.time() + secondary_signal_wait_s),
                        )
                        tail_deadline = min(
                            hard_deadline,
                            max(tail_deadline or base_deadline, pending_pass_deadline),
                        )
                        continue
                    settle_needed = success_settle_s > 0
                    if settle_needed:
                        if pending_pass_reason != "settle":
                            pending_pass_reference_clock = reference_clock
                            pending_pass_reason = "settle"
                            pending_pass_deadline = min(
                                hard_deadline,
                                max(pending_pass_deadline or base_deadline, time.time() + success_settle_s),
                            )
                            tail_deadline = min(
                                hard_deadline,
                                max(tail_deadline or base_deadline, pending_pass_deadline),
                            )
                            continue
                        if pending_pass_deadline is not None and time.time() < pending_pass_deadline:
                            continue
                    detail = f"命令通过，第 {index + 1} 次：{word}"
                    if pending_advisory_logs:
                        detail = f"{detail}（辅助日志未到：{'/'.join(pending_advisory_logs)}）"
                    if pending_advisory_tones:
                        detail = f"{detail}（辅助播报未到：{'/'.join(pending_advisory_tones)}）"
                    if pending_secondary_signals:
                        detail = f"{detail}（补充信号未到：{'/'.join(pending_secondary_signals)}）"
                    return self._make_action_result(
                        "PASS",
                        "say",
                        detail,
                        start,
                        recognition_latency_s=last_recognition_latency,
                        actual_asr=all_asr,
                        actual_tones=all_tones,
                        actual_protocols=all_protocols,
                        actual_log_values=all_log_values,
                        pending_advisory_logs=pending_advisory_logs,
                        reference_clock=reference_clock,
                        evidence=[json.dumps(all_log_values, ensure_ascii=False)] if all_log_values else [],
                    )
            if pending_pass_deadline is not None:
                pending_advisory_logs = [
                    key for key, values in advisory_logs.items() if not set(values).intersection(all_log_values.get(key, []))
                ]
                pending_advisory_tones = []
                if advisory_tone and advisory_tone not in all_tones:
                    pending_advisory_tones.append(advisory_tone)
                pending_secondary_signals: list[str] = []
                if expected_tone and expected_protocol and not require_protocol_and_tone:
                    if expected_tone not in all_tones:
                        pending_secondary_signals.append(f"播报:{expected_tone}")
                    if expected_protocol not in all_protocols:
                        pending_secondary_signals.append(f"协议:{expected_protocol}")
                detail = f"命令通过，第 {index + 1} 次：{word}"
                if pending_advisory_logs:
                    detail = f"{detail}（辅助日志未到：{'/'.join(pending_advisory_logs)}）"
                if pending_advisory_tones:
                    detail = f"{detail}（辅助播报未到：{'/'.join(pending_advisory_tones)}）"
                if pending_secondary_signals:
                    detail = f"{detail}（补充信号未到：{'/'.join(pending_secondary_signals)}）"
                return self._make_action_result(
                    "PASS",
                    "say",
                    detail,
                    start,
                    recognition_latency_s=last_recognition_latency,
                    actual_asr=all_asr,
                    actual_tones=all_tones,
                    actual_protocols=all_protocols,
                    actual_log_values=all_log_values,
                    pending_advisory_logs=pending_advisory_logs,
                    evidence=[json.dumps(all_log_values, ensure_ascii=False)] if all_log_values else [],
                    reference_clock=pending_pass_reference_clock or last_reference_clock or recognition_clock,
                    )
            if expect_no_response and not all_tones and not all_protocols:
                detail = f"静默断言通过：{word}"
                if all_asr:
                    detail = f"静默断言通过（有识别无播报/协议）：{word}"
                return self._make_action_result(
                    "PASS",
                    "say",
                    detail,
                    start,
                    recognition_latency_s=last_recognition_latency,
                    actual_asr=all_asr,
                    actual_tones=all_tones,
                    actual_protocols=all_protocols,
                    actual_log_values=all_log_values,
                    evidence=[json.dumps(all_log_values, ensure_ascii=False)] if all_log_values else [],
                    reference_clock=recognition_clock,
                )
            retry_gate_clock = play_end_clock

        detail = f"静默断言失败：{word}" if expect_no_response else f"命令失败：{word}"
        if not expect_no_response:
            mismatch_parts = []
            if expected_asr and not expected_asr.intersection(all_asr):
                mismatch_parts.append(f"识别不符(期望 {'/'.join(expected_asr)} / 实际 {'/'.join(all_asr) or '-'})")
            if expected_tone and expected_tone not in all_tones:
                mismatch_parts.append(f"播报ID不符(期望 {expected_tone} / 实际 {'/'.join(all_tones) or '-'})")
            elif expect_any_tone and not all_tones:
                mismatch_parts.append("未观察到任何播报")
            if expected_protocol and expected_protocol not in all_protocols:
                mismatch_parts.append(f"协议不符(期望 {expected_protocol} / 实际 {'/'.join(all_protocols) or '-'})")
            for key, values in expected_logs.items():
                actual_values = all_log_values.get(key, [])
                if not set(values).intersection(actual_values):
                    mismatch_parts.append(f"日志值不符({key}: 期望 {'/'.join(values)} / 实际 {'/'.join(actual_values) or '-'})")
            if mismatch_parts:
                detail = f"{detail}；" + "；".join(mismatch_parts)
        return self._make_action_result(
            "FAIL",
            "say",
            detail,
            start,
            recognition_latency_s=last_recognition_latency,
            actual_asr=all_asr,
            actual_tones=all_tones,
            actual_protocols=all_protocols,
            actual_log_values=all_log_values,
            reference_clock=last_reference_clock,
            evidence=[json.dumps(all_log_values, ensure_ascii=False)] if all_log_values else [],
        )

    def run_wait(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        seconds = float(action["seconds"])
        tolerance = float(action.get("tolerance_s", 1.0))
        expect_tone = parse_tone_id(action.get("expect_tone_id", ""))
        expect_markers = [str(item) for item in action.get("expect_markers", []) if str(item)]
        validate_duration = bool(action.get("validate_duration", True))
        measure_timeout = bool(action.get("measure_timeout"))
        reference_clock = self.last_action.get("reference_clock") if self.last_action else None
        expected_desc = expect_tone or " / ".join(expect_markers) or "仅等待"
        self.log.info(
            f"等待观测：目标 {seconds:.1f}s，证据={expected_desc}，"
            f"measure_timeout={'Y' if measure_timeout else 'N'}，validate_duration={'Y' if validate_duration else 'N'}"
        )
        self.clear_observation()
        wait_start_clock = self._now_clock()
        deadline = time.perf_counter() + seconds + tolerance + 1.0
        while time.perf_counter() < deadline:
            time.sleep(0.2)
            tones = self.observed_tones()
            marker_clock = self._find_latest_serial_clock(expect_markers, after_clock=wait_start_clock) if expect_markers else None
            tone_clock = self._latest_play_id_clock(expect_tone, after_clock=wait_start_clock) if expect_tone else None
            tone_hit = bool(expect_tone and expect_tone in tones and tone_clock is not None)
            marker_hit = marker_clock is not None
            if tone_hit or marker_hit:
                elapsed = time.perf_counter() - start
                if measure_timeout and reference_clock is not None:
                    event_clock = marker_clock or tone_clock
                    if event_clock is not None:
                        elapsed = event_clock - reference_clock
                        if elapsed < 0:
                            elapsed += 24 * 3600
                if validate_duration and abs(elapsed - seconds) > tolerance:
                    return self._make_action_result("FAIL", "wait", f"等待耗时异常：{elapsed:.3f}s", start, actual_tones=tones)
                detail = f"等待超时通过：{expect_tone}" if tone_hit else f"等待超时通过：{' / '.join(expect_markers)}"
                if measure_timeout and reference_clock is not None:
                    detail += f"（{elapsed:.3f}s）"
                return self._make_action_result(
                    "PASS",
                    "wait",
                    detail,
                    start,
                    actual_tones=tones,
                    evidence=expect_markers,
                    reference_clock=reference_clock,
                )
        if not expect_tone and not expect_markers:
            return self._make_action_result("PASS", "wait", f"等待完成：{seconds}s", start)
        return self._make_action_result("FAIL", "wait", f"未等到超时证据：{expected_desc}", start, actual_tones=self.observed_tones())

    def run_reboot(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        if not self.do_reboot():
            return self._make_action_result("FAIL", "reboot", "reboot 重试耗尽仍失败", start)
        ready_delay = float(self.policy.get("postBootReadyDelayS", 4.0))
        self.log.info(f"重启完成后等待 {ready_delay:.1f}s，给算法资源加载")
        time.sleep(ready_delay)
        if not self.recover_logs_after_reboot("主动 reboot 后恢复日志", fatal=True):
            return self._make_action_result("FAIL", "reboot", self.stop_reason or "重启后 loglevel 4 恢复失败", start)
        self.sync_reboot_baseline("主动 reboot 动作完成，忽略本次预期重启")
        expect_tone = parse_tone_id(action.get("expect_tone_id", ""))
        deadline = time.perf_counter() + float(action.get("wait_log_recovery_s", self.policy["rebootRecoveryS"]))
        self.clear_observation()
        while time.perf_counter() < deadline:
            time.sleep(1.0)
            tones = self.observed_tones()
            recent = self.reader.get_recent_lines()[-50:]
            if recent and not expect_tone:
                return self._make_action_result("PASS", "reboot", "重启后日志恢复", start, actual_tones=tones, evidence=recent[-10:])
            if recent and expect_tone in tones:
                return self._make_action_result("PASS", "reboot", "重启后日志与播报恢复", start, actual_tones=tones, evidence=recent[-10:])
        if self.reader.get_recent_lines():
            return self._make_action_result("FAIL", "reboot", f"重启成功但未观察到期望播报：{expect_tone}", start, actual_tones=self.observed_tones(), evidence=self.reader.get_recent_lines()[-10:])
        return self._make_action_result("FAIL", "reboot", "重启后未恢复到可观测状态", start, actual_tones=self.observed_tones())

    def run_inject_protocol(self, action: dict[str, Any]) -> dict[str, Any]:
        assert self.proto is not None
        start = time.perf_counter()
        protocol = normalize_hex(action["protocol"])
        if not protocol:
            return self._make_action_result("FAIL", "inject_protocol", "协议注入动作缺少协议内容", start)
        expect_tone = parse_tone_id(action.get("expect_tone_id", ""))
        expect_recv_protocol = normalize_hex(action.get("expect_recv_protocol", protocol)) if action.get("expect_recv_protocol", protocol) else ""
        require_receive_msg = bool(action.get("require_receive_msg", False))
        require_active_window = bool(action.get("requires_active_window", False))
        expect_no_response = bool(action.get("expect_no_response", False))
        retries = int(action.get("truncation_retries", 2)) + 1
        last_partial: list[str] = []
        for attempt in range(1, retries + 1):
            self.clear_observation()
            if require_active_window:
                self.log.info(f"已进入遥控器配置窗口，等待协议串口发送：{protocol}")
            self.proto.write_hex(protocol)
            time.sleep(float(action.get("observe_s", self.policy["commandObserveS"])))
            tones = self.observed_tones()
            frames = self.proto.get_frames()
            partial = self.proto.get_partial_frames()
            recv_values = self.observed_log_protocols("recv")
            recent = self.reader.get_recent_lines()[-10:] if self.reader else []
            if partial and not frames:
                last_partial = partial
                if attempt < retries:
                    self.note_protocol_truncation(partial)
                    self.log.warn(f"协议截断，准备重试 {attempt}/{retries - 1}：{protocol}")
                    continue
                self.note_protocol_truncation(partial)
                return self._make_action_result(
                    "FAIL",
                    "inject_protocol",
                    "协议注入出现截断，重试耗尽",
                    start,
                    actual_tones=tones,
                    actual_protocols=frames,
                    actual_log_values={"recvMsg": recv_values} if recv_values else {},
                    evidence=recent + partial[-3:],
                )
            if expect_no_response:
                if not tones and not recv_values and not frames:
                    return self._make_action_result(
                        "PASS",
                        "inject_protocol",
                        f"窗口外注入无效：{protocol or '-'}",
                        start,
                        actual_tones=tones,
                        actual_protocols=frames,
                        actual_log_values={"recvMsg": recv_values} if recv_values else {},
                        evidence=recent + last_partial[-3:],
                    )
                detail = []
                if tones:
                    detail.append(f"出现播报 {'/'.join(tones)}")
                if recv_values:
                    detail.append(f"出现 receive msg {'/'.join(recv_values)}")
                if frames:
                    detail.append(f"协议口收到 {'/'.join(frames)}")
                return self._make_action_result(
                    "FAIL",
                    "inject_protocol",
                    "窗口外注入仍有响应：" + "；".join(detail),
                    start,
                    actual_tones=tones,
                    actual_protocols=frames,
                    actual_log_values={"recvMsg": recv_values} if recv_values else {},
                    evidence=recent + last_partial[-3:],
                )
            recv_ok = (not require_receive_msg) or (expect_recv_protocol in recv_values if expect_recv_protocol else bool(recv_values))
            tone_ok = (not expect_tone) or (expect_tone in tones)
            if recv_ok and tone_ok:
                detail = f"协议注入通过：{protocol}"
                if require_active_window:
                    detail = f"配置窗口内协议注入通过：{protocol}"
                return self._make_action_result(
                    "PASS",
                    "inject_protocol",
                    detail,
                    start,
                    actual_tones=tones,
                    actual_protocols=frames,
                    actual_log_values={"recvMsg": recv_values} if recv_values else {},
                    evidence=recent,
                )
            missing = []
            if require_receive_msg and not recv_ok:
                missing.append(f"未观察到 receive msg(期望 {expect_recv_protocol or '-'} / 实际 {'/'.join(recv_values) or '-'})")
            if expect_tone and not tone_ok:
                missing.append(f"播报ID不符(期望 {expect_tone} / 实际 {'/'.join(tones) or '-'})")
                return self._make_action_result(
                    "FAIL",
                    "inject_protocol",
                    "；".join(missing) if missing else "协议注入后未观察到期望响应",
                    start,
                    actual_tones=tones,
                    actual_protocols=frames,
                    actual_log_values={"recvMsg": recv_values} if recv_values else {},
                    evidence=recent + last_partial[-3:],
                )
        return self._make_action_result(
            "FAIL",
            "inject_protocol",
            "协议注入异常退出",
            start,
            actual_protocols=[],
            evidence=last_partial[-3:],
        )

    def run_volume_walk(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        expected_step_count = int(action["expected_step_count"])
        extra_boundary = int(action.get("extra_boundary_repeats", 1))
        step_tone = parse_tone_id(action["step_tone_id"])
        boundary_tone = parse_tone_id(action["boundary_tone_id"])
        step_hits = 0
        boundary_hits = 0
        while boundary_hits < 1 + extra_boundary and step_hits <= expected_step_count + 2:
            wake_result = self.run_wake({"word": self.config["wakeupWord"], "expect_tone_id": "TONE_ID_0", "retries": self.policy["wakeRetries"]})
            if wake_result["status"] != "PASS":
                return self._make_action_result("FAIL", "volume_walk", "音量遍历前唤醒失败", start)
            say_result = self.run_say({"word": action["command_word"], "expect_asr": [action["command_word"]], "retries": 1, "observe_s": self.policy["commandObserveS"]})
            tones = say_result["actual_tones"]
            if boundary_tone in tones:
                boundary_hits += 1
            elif step_tone in tones:
                step_hits += 1
            else:
                return self._make_action_result("FAIL", "volume_walk", f"音量遍历出现异常播报：{tones}", start, actual_tones=tones)
            time.sleep(0.5)
        if step_hits == expected_step_count and boundary_hits >= 1 + extra_boundary:
            return self._make_action_result("PASS", "volume_walk", f"音量遍历通过，step={step_hits}, boundary={boundary_hits}", start, actual_tones=[step_tone, boundary_tone])
        return self._make_action_result("FAIL", "volume_walk", f"音量遍历不符，step={step_hits}, boundary={boundary_hits}", start, actual_tones=[step_tone, boundary_tone])

    def run_assert_wake_repeats(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        repeats = int(action["repeats"])
        wait_for_idle = bool(action.get("wait_for_idle_between_repeats", False))
        idle_timeout_s = float(action.get("idle_timeout_s", 0.0) or 0.0)
        timeout_tone_id = parse_tone_id(action.get("timeout_tone_id", "TONE_ID_2"))
        timeout_markers = [str(item).strip() for item in action.get("timeout_markers", ["TIME_OUT"]) if str(item).strip()]
        for index in range(repeats):
            result = self.run_wake(
                {
                    "word": action["word"],
                    "expect_tone_id": action.get("expect_tone_id", "TONE_ID_0"),
                    "require_any_tone": action.get("require_any_tone", False),
                    "retries": self.policy["wakeRetries"],
                }
            )
            if result["status"] != "PASS":
                return self._make_action_result("FAIL", "assert_wake_repeats", f"连续唤醒失败，第 {index + 1} 次", start)
            if index + 1 < repeats and wait_for_idle and idle_timeout_s > 0:
                wait_result = self.run_wait(
                    {
                        "seconds": idle_timeout_s,
                        "expect_tone_id": timeout_tone_id,
                        "expect_markers": timeout_markers,
                        "measure_timeout": False,
                        "validate_duration": False,
                    }
                )
                if wait_result["status"] != "PASS":
                    return self._make_action_result(
                        "FAIL",
                        "assert_wake_repeats",
                        f"连续唤醒后未回到空闲态，第 {index + 1} 次；{wait_result['detail']}",
                        start,
                    )
            if index + 1 < repeats:
                repeat_gap_s = max(float(action.get("repeat_gap_s", self.policy.get("postPlaybackGapS", 0.2))), 0.0)
                if repeat_gap_s > 0:
                    time.sleep(repeat_gap_s)
        return self._make_action_result("PASS", "assert_wake_repeats", f"连续唤醒通过：{repeats} 次", start)

    def run_assert_no_wake(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        repeats = int(action.get("repeats", 1))
        observe_s = float(action.get("observe_s", self.policy["silenceObserveS"]))
        expect_tone = parse_tone_id(action.get("expect_tone_id", "TONE_ID_0"))
        last_recognition_latency: float | None = None
        words = action.get("words")
        if not isinstance(words, list) or not words:
            words = [action.get("word", "")]
        words = [str(item).strip() for item in words if str(item).strip()]
        wake_values: list[str] = []
        tones: list[str] = []
        protocols: list[str] = []
        expected_protocol = normalize_hex(action.get("expect_send_protocol", ""))
        require_protocol_when_configured = bool(
            action.get("require_protocol_when_configured", self.policy.get("requireWakeProtocolWhenConfigured", True))
        )
        require_ready_signal = bool(
            action.get("require_ready_signal", self.policy.get("requireWakeReadySignal", True))
        )
        for word in words:
            for index in range(repeats):
                self.log.info(f"负例校验：词【{word}】第 {index + 1}/{repeats} 次，观测 {observe_s:.1f}s")
                if not self.maybe_recover_logs():
                    return self._make_action_result("FAIL", "assert_no_wake", "日志恢复失败", start)
                mode_snapshot = self.current_work_mode()
                self.clear_observation()
                if not self.play_word(word):
                    return self._make_action_result("FAIL", "assert_no_wake", f"误唤醒音频播放失败：{word}", start)
                play_end_clock = self._now_clock()
                match_after_clock = self._shift_clock(play_end_clock, -2.0)
                time.sleep(observe_s)
                wake_eval = self._evaluate_wake_signal(
                    word=word,
                    expect_tone=expect_tone,
                    require_any_tone=False,
                    expected_protocol=expected_protocol,
                    mode_snapshot=mode_snapshot,
                    match_after_clock=match_after_clock,
                    require_protocol_when_configured=require_protocol_when_configured,
                    require_ready_signal=require_ready_signal,
                )
                wake_values = wake_eval["wake_values"]
                asr_values = wake_eval["asr_values"]
                recognized_values = wake_eval["recognized_values"]
                tones = wake_eval["tones"]
                protocols = wake_eval["protocols"]
                recognition_clock = wake_eval["recognition_clock"]
                last_recognition_latency = self._clock_diff(recognition_clock, play_end_clock)
                if wake_eval["wake_ok"]:
                    detail_parts = []
                    if wake_values:
                        detail_parts.append(f"wakeKw={'/'.join(wake_values)}")
                    if asr_values:
                        detail_parts.append(f"asrKw={'/'.join(asr_values)}")
                    if tones:
                        detail_parts.append(f"playId={'/'.join(tones)}")
                    if protocols:
                        detail_parts.append(f"protocols={'/'.join(protocols)}")
                    if wake_eval["ready_ok"]:
                        detail_parts.append("/".join(wake_eval["ready_evidence"]) or "wakeReady=marker")
                    return self._make_action_result(
                        "FAIL",
                        "assert_no_wake",
                        f"发生误唤醒：{word}，第 {index + 1} 次；" + "；".join(detail_parts),
                        start,
                        recognition_latency_s=last_recognition_latency,
                        actual_asr=recognized_values,
                        actual_tones=tones,
                        actual_protocols=protocols,
                    )
        return self._make_action_result(
            "PASS",
            "assert_no_wake",
            f"未误唤醒：{len(words)} 个词，每词 {repeats} 次",
            start,
            recognition_latency_s=last_recognition_latency,
            actual_asr=wake_values if repeats else [],
            actual_tones=tones,
            actual_protocols=protocols,
        )

    def run_phrase_check(self, action: dict[str, Any]) -> dict[str, Any]:
        assert self.reader is not None
        start = time.perf_counter()
        observe_s = float(action.get("observe_s", 1.8))
        inter_word_gap_s = float(action.get("inter_word_gap_s", 0.2))
        wake_word = str(action.get("wake_word") or self.config["wakeupWord"]).strip() or self.config["wakeupWord"]
        timeout_markers = [str(item).strip() for item in action.get("timeout_markers", ["TIME_OUT"]) if str(item).strip()]
        timeout_tone = parse_tone_id(action.get("timeout_tone_id", "TONE_ID_2"))
        items = list(action.get("items", []))
        all_asr: list[str] = []
        all_tones: list[str] = []
        all_protocols: list[str] = []
        evidence: list[str] = []
        failures: list[str] = []
        last_recognition_latency: float | None = None
        need_wake = True
        wake_count = 0

        for index, item in enumerate(items, start=1):
            word = str(item.get("word", "")).strip()
            if not word:
                continue
            accepted_asr = {
                str(value).strip()
                for value in item.get("accept_asr", [word])
                if str(value).strip()
            } or {word}

            if need_wake:
                wake_result = self.run_wake(
                    {
                        "word": wake_word,
                        "retries": self.policy["wakeRetries"],
                    }
                )
                wake_count += 1
                evidence.append(f"唤醒#{wake_count}:{wake_result['status']}:{wake_result['detail']}")
                all_tones.extend(item for item in wake_result.get("actual_tones", []) if item not in all_tones)
                if wake_result["status"] != "PASS":
                    return self._make_action_result(
                        "FAIL",
                        "phrase_check",
                        f"词条检查前唤醒失败：{word}",
                        start,
                        recognition_latency_s=last_recognition_latency,
                        actual_asr=all_asr,
                        actual_tones=all_tones,
                        actual_protocols=all_protocols,
                        evidence=evidence,
                    )
                need_wake = False
                time.sleep(0.2)

            if not self.maybe_recover_logs():
                return self._make_action_result(
                    "FAIL",
                    "phrase_check",
                    "日志恢复失败",
                    start,
                    recognition_latency_s=last_recognition_latency,
                    actual_asr=all_asr,
                    actual_tones=all_tones,
                    actual_protocols=all_protocols,
                    evidence=evidence,
                )

            self.clear_observation()
            if not self.play_word(word):
                return self._make_action_result(
                    "FAIL",
                    "phrase_check",
                    f"词条音频播放失败：{word}",
                    start,
                    recognition_latency_s=last_recognition_latency,
                    actual_asr=all_asr,
                    actual_tones=all_tones,
                    actual_protocols=all_protocols,
                    evidence=evidence,
                )

            play_end_clock = self._now_clock()
            match_after_clock = self._shift_clock(play_end_clock, -2.0)
            recognition_clock: float | None = None
            timeout_clock: float | None = None
            item_asr: list[str] = []
            item_tones: list[str] = []
            item_protocols: list[str] = []
            recognized = False
            timeout_hit = False
            deadline = time.perf_counter() + observe_s

            while time.perf_counter() < deadline:
                time.sleep(0.2)
                item_asr = self.observed_asr()
                item_tones = self.observed_tones()
                item_protocols = self.observed_protocols()
                if any(value in accepted_asr for value in item_asr):
                    recognized = True
                    recognition_clock = self._find_latest_serial_clock(["Wakeup:", "keyword:"], after_clock=match_after_clock)
                    break
                marker_clock = self._find_latest_serial_clock(timeout_markers, after_clock=match_after_clock) if timeout_markers else None
                tone_clock = self._latest_play_id_clock(timeout_tone, after_clock=match_after_clock) if timeout_tone else None
                if marker_clock is not None or tone_clock is not None:
                    timeout_hit = True
                    timeout_clock = marker_clock or tone_clock
                    break

            if recognized and recognition_clock is not None:
                last_recognition_latency = self._clock_diff(recognition_clock, play_end_clock)
            all_asr.extend(value for value in item_asr if value not in all_asr)
            all_tones.extend(value for value in item_tones if value not in all_tones)
            all_protocols.extend(value for value in item_protocols if value not in all_protocols)

            if recognized:
                actual_text = " / ".join(item_asr) if item_asr else "-"
                evidence.append(f"{index:02d}.{word}:PASS:ASR={actual_text}")
            else:
                reason = "timeout" if timeout_hit else "no_asr"
                actual_text = " / ".join(item_asr) if item_asr else "-"
                failures.append(word)
                evidence.append(f"{index:02d}.{word}:FAIL:{reason}:ASR={actual_text}")

            if timeout_hit or timeout_clock is not None:
                need_wake = True

            if inter_word_gap_s > 0 and index < len(items):
                time.sleep(inter_word_gap_s)

        if failures:
            detail = f"词条检查失败：{', '.join(failures)}"
            return self._make_action_result(
                "FAIL",
                "phrase_check",
                detail,
                start,
                recognition_latency_s=last_recognition_latency,
                actual_asr=all_asr,
                actual_tones=all_tones,
                actual_protocols=all_protocols,
                evidence=evidence,
            )
        return self._make_action_result(
            "PASS",
            "phrase_check",
            f"词条检查通过：{len(items)} 条",
            start,
            recognition_latency_s=last_recognition_latency,
            actual_asr=all_asr,
            actual_tones=all_tones,
            actual_protocols=all_protocols,
            evidence=evidence,
        )

    def power_control_supported(self) -> bool:
        cfg = self.config.get("powerControl", {})
        return bool(cfg.get("port") and cfg.get("baudRate"))

    def _control_commands(self, commands: list[str]) -> list[str]:
        cfg = self.config.get("powerControl", {})
        port_name = str(cfg.get("port", "")).strip()
        baudrate = int(cfg.get("baudRate", 115200) or 115200)
        delay_ms = int(cfg.get("commandDelayMs", 300) or 300)
        if not port_name:
            raise RuntimeError("power control port is not configured")
        port = serial.Serial(port_name, baudrate=baudrate, timeout=0.5, write_timeout=1)
        issued: list[str] = []
        try:
            for command in commands:
                line = str(command).strip()
                if not line:
                    continue
                self.log.info(f"控制串口 {port_name} <- {line}")
                port.write((line + "\r\n").encode("ascii", errors="ignore"))
                port.flush()
                issued.append(line)
                time.sleep(delay_ms / 1000.0)
        finally:
            port.close()
        return issued

    def run_manual_power_cycle(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        note = action.get("note", "执行上下电循环")
        if not self.power_control_supported():
            if self.args.manual_block_as_fail:
                return self._make_action_result("FAIL", "manual_power_cycle", note, start)
            return self._make_action_result("BLOCKED", "manual_power_cycle", note, start)

        off_wait_s = float(action.get("off_wait_s", 1.0))
        boot_wait_s = float(action.get("boot_wait_s", 8.0))
        config_saved_guard_s = float(action.get("config_saved_guard_s", self.policy.get("configSavedGuardS", 5.0)))
        expect_tone = parse_tone_id(action.get("expect_tone_id", ""))
        power_off_cmds = action.get("power_off_cmds") or ["uut-switch1.off", "uut-switch2.off"]
        power_on_cmds = action.get("power_on_cmds") or ["uut-switch1.on"]
        try:
            evidence: list[str] = []
            pending_advisory_logs = set(self.last_action.get("pending_advisory_logs", [])) if self.last_action else set()
            if "configSaved" in pending_advisory_logs and self.reader and config_saved_guard_s > 0:
                guard_deadline = time.time() + config_saved_guard_s
                while time.time() < guard_deadline:
                    current_values = self.observed_log_values("configSaved")
                    if current_values:
                        evidence.append("configSaved observed before power cycle")
                        break
                    time.sleep(0.2)
                else:
                    evidence.append("configSaved still missing before power cycle")
            self.clear_observation()
            evidence.extend(self._control_commands([str(cmd) for cmd in power_off_cmds]))
            if off_wait_s > 0:
                time.sleep(off_wait_s)
            evidence.extend(self._control_commands([str(cmd) for cmd in power_on_cmds]))
            if self.reader or self.proto:
                if not self.reopen_observers_after_power_event("上下电后刷新观察串口", wait_s=boot_wait_s):
                    self.stop_reason = "上下电后日志/协议串口重连失败"
                    self.sync_reboot_baseline("主动上下电已执行但观察串口重连失败")
                    return self._make_action_result(
                        "FAIL",
                        "manual_power_cycle",
                        self.stop_reason,
                        start,
                        evidence=evidence,
                    )
            elif boot_wait_s > 0:
                time.sleep(boot_wait_s)
            tones = self.observed_tones()
            if self.reader and not self.recover_logs_after_reboot("上下电后恢复日志", fatal=True):
                self.sync_reboot_baseline("主动上下电已执行但日志恢复失败")
                return self._make_action_result(
                    "FAIL",
                    "manual_power_cycle",
                    self.stop_reason or "上下电后 loglevel 4 恢复失败",
                    start,
                    evidence=evidence,
                )
            tones = self.observed_tones()
            if expect_tone and expect_tone not in tones:
                self.sync_reboot_baseline("主动上下电已执行，忽略本次预期重启")
                return self._make_action_result(
                    "FAIL",
                    "manual_power_cycle",
                    f"上下电后未观察到期望播报：{expect_tone}",
                    start,
                    actual_tones=tones,
                    evidence=evidence + (self.reader.get_recent_lines()[-10:] if self.reader else []),
                )
            self.sync_reboot_baseline("主动上下电已执行，忽略本次预期重启")
            return self._make_action_result("PASS", "manual_power_cycle", note, start, actual_tones=tones, evidence=evidence)
        except Exception as exc:
            return self._make_action_result("FAIL", "manual_power_cycle", f"上下电执行失败: {exc}", start)

    def run_log_probe(self, action: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        if not self.last_action:
            return self._make_action_result("FAIL", "log_probe", "前置动作不存在", start)
        mapping = {
            "wakeKw": self.last_action.get("actual_asr", []),
            "asrKw": self.last_action.get("actual_asr", []),
            "playId": self.last_action.get("actual_tones", []),
            "sendMsg": self.last_action.get("actual_protocols", []),
            "recvMsg": self.proto.get_frames() if self.proto else [],
        }
        missing = [key for key in action.get("expected_keys", []) if not mapping.get(key)]
        if missing:
            return self._make_action_result("FAIL", "log_probe", f"日志字段缺失：{missing}", start, evidence=[str(mapping)])
        return self._make_action_result("PASS", "log_probe", "日志探针校验通过", start, evidence=[str(mapping)])

    def run_action(self, action: dict[str, Any]) -> dict[str, Any]:
        action_type = action["type"]
        action_label = self.describe_action(action)
        result: dict[str, Any]
        if action_type == "wake":
            result = self.run_wake(action)
        elif action_type == "say":
            result = self.run_say(action)
        elif action_type == "wait":
            result = self.run_wait(action)
        elif action_type == "reboot":
            result = self.run_reboot(action)
        elif action_type == "inject_protocol":
            result = self.run_inject_protocol(action)
        elif action_type == "volume_walk":
            result = self.run_volume_walk(action)
        elif action_type == "assert_wake_repeats":
            result = self.run_assert_wake_repeats(action)
        elif action_type == "assert_no_wake":
            result = self.run_assert_no_wake(action)
        elif action_type == "phrase_check":
            result = self.run_phrase_check(action)
        elif action_type == "manual_power_cycle":
            result = self.run_manual_power_cycle(action)
        elif action_type == "log_probe":
            result = self.run_log_probe(action)
        else:
            result = self._make_action_result("FAIL", action_type, f"未知动作类型：{action_type}", time.perf_counter())
        if action_type not in {"reboot", "manual_power_cycle"}:
            reboot_detail = self.current_case_unexpected_reboot_detail(f"动作【{action_label}】后")
            if reboot_detail:
                original_detail = str(result.get("detail", "")).strip()
                result["status"] = "FAIL"
                result["detail"] = reboot_detail if not original_detail else f"{reboot_detail}；原动作结果：{original_detail}"
        result["action_label"] = action_label
        self.last_action = result
        return result

    def run_case(self, case_item: dict[str, Any]) -> dict[str, Any]:
        start = time.perf_counter()
        if bool(case_item.get("manual_only")):
            return {
                "case_id": case_item["case_id"],
                "module": case_item["module"],
                "name": case_item["name"],
                "priority": case_item["priority"],
                "status": "DRY_RUN",
                "detail": str(case_item.get("manual_reason") or "人工验证项，已跳过自动执行"),
                "duration_s": round(time.perf_counter() - start, 3),
                "expected_protocol": case_item["expected_protocol"],
                "expected_tone_id": case_item["expected_tone_id"],
                "expected_tone_text": case_item["expected_tone_text"],
                "actual_protocols": [],
                "actual_tones": [],
                "actions": [],
            }
        if self.stop_reason:
            return {
                "case_id": case_item["case_id"],
                "module": case_item["module"],
                "name": case_item["name"],
                "priority": case_item["priority"],
                "status": "BLOCKED",
                "detail": self.stop_reason,
                "duration_s": round(time.perf_counter() - start, 3),
                "expected_protocol": case_item["expected_protocol"],
                "expected_tone_id": case_item["expected_tone_id"],
                "expected_tone_text": case_item["expected_tone_text"],
                "actual_protocols": [],
                "actual_tones": [],
                "actions": [],
            }
        blocked_by_case_ids = [str(item).strip() for item in case_item.get("blocked_by_case_ids", []) if str(item).strip()]
        dependency_reasons = []
        for dep_case_id in blocked_by_case_ids:
            dep_result = self.case_results_by_id.get(dep_case_id)
            if dep_result and dep_result.get("status") != "PASS":
                dependency_reasons.append(f"依赖用例 {dep_case_id}={dep_result.get('status')}：{dep_result.get('detail', '')}")
        if dependency_reasons:
            return {
                "case_id": case_item["case_id"],
                "module": case_item["module"],
                "name": case_item["name"],
                "priority": case_item["priority"],
                "status": "BLOCKED",
                "detail": "；".join(dependency_reasons),
                "duration_s": round(time.perf_counter() - start, 3),
                "expected_protocol": case_item["expected_protocol"],
                "expected_tone_id": case_item["expected_tone_id"],
                "expected_tone_text": case_item["expected_tone_text"],
                "actual_protocols": [],
                "actual_tones": [],
                "actions": [],
            }
        blocked_reasons = [
            self.capabilities[dep]["reason"]
            for dep in self.case_dependencies(case_item)
            if dep in self.capabilities and not self.capabilities[dep]["supported"]
        ]
        if blocked_reasons:
            return {
                "case_id": case_item["case_id"],
                "module": case_item["module"],
                "name": case_item["name"],
                "priority": case_item["priority"],
                "status": "BLOCKED",
                "detail": "；".join(blocked_reasons),
                "duration_s": round(time.perf_counter() - start, 3),
                "expected_protocol": case_item["expected_protocol"],
                "expected_tone_id": case_item["expected_tone_id"],
                "expected_tone_text": case_item["expected_tone_text"],
                "actual_protocols": [],
                "actual_tones": [],
                "actions": [],
            }
        case_status = "PASS"
        case_detail = "全部动作通过"
        action_results: list[dict[str, Any]] = []
        actual_protocols: list[str] = []
        actual_tones: list[str] = []
        advisory_failures: list[str] = []
        stop_main_actions = False
        for action in case_item["actions"]:
            if stop_main_actions and not bool(action.get("always_run", False)):
                continue
            result = self.run_action(action)
            action_results.append(result)
            if not bool(action.get("setup_action")) and not bool(action.get("always_run", False)):
                actual_protocols.extend(result.get("actual_protocols", []))
                actual_tones.extend(result.get("actual_tones", []))
            if result["status"] == "FAIL":
                if bool(action.get("advisory_failure", False)):
                    advisory_failures.append(result["detail"])
                elif case_status != "FAIL":
                    case_status = "FAIL"
                    case_detail = result["detail"]
                elif bool(action.get("always_run", False)):
                    case_detail = f"{case_detail}；清理失败：{result['detail']}"
                if not action.get("allow_fail_continue", False):
                    stop_main_actions = True
            if result["status"] == "BLOCKED" and case_status != "FAIL":
                case_status = "BLOCKED"
                case_detail = result["detail"]
        if case_status == "PASS" and advisory_failures:
            case_detail = f"主断言通过（前置差异：{advisory_failures[0]}）"
        self._record_runtime_gap(case_item["case_id"], actual_protocols)
        return {
            "case_id": case_item["case_id"],
            "module": case_item["module"],
            "name": case_item["name"],
            "priority": case_item["priority"],
            "status": case_status,
            "detail": case_detail,
            "duration_s": round(time.perf_counter() - start, 3),
            "expected_protocol": case_item["expected_protocol"],
            "expected_tone_id": case_item["expected_tone_id"],
            "expected_tone_text": case_item["expected_tone_text"],
            "actual_protocols": actual_protocols,
            "actual_tones": actual_tones,
            "actions": action_results,
        }

    def _make_blocked_case_result(self, case_item: dict[str, Any], detail: str, source: str) -> dict[str, Any]:
        result = {
            "case_id": case_item["case_id"],
            "module": case_item["module"],
            "name": case_item["name"],
            "priority": case_item["priority"],
            "status": "BLOCKED",
            "detail": detail,
            "duration_s": 0.0,
            "expected_protocol": case_item["expected_protocol"],
            "expected_tone_id": case_item["expected_tone_id"],
            "expected_tone_text": case_item["expected_tone_text"],
            "actual_protocols": [],
            "actual_tones": [],
            "actions": [],
        }
        self._mark_case_attempt(result, source, 0)
        result["final_source"] = source
        return result

    def _blocked_results_for_selected_cases(self, detail: str, source: str) -> list[dict[str, Any]]:
        return [self._make_blocked_case_result(case_item, detail, source) for case_item in self.selected_cases]

    def _append_unexecuted_blocked_cases(
        self,
        results: list[dict[str, Any]],
        detail: str,
        source: str,
    ) -> list[dict[str, Any]]:
        existing_case_ids = {item["case_id"] for item in results}
        pending_results = [
            self._make_blocked_case_result(case_item, detail, source)
            for case_item in self.selected_cases
            if case_item["case_id"] not in existing_case_ids
        ]
        return [*results, *pending_results]

    def _record_runtime_gap(self, case_id: str, actual_protocols: list[str]) -> None:
        rule = RUNTIME_GAP_RULES.get(case_id)
        if not rule:
            return
        expected = normalize_hex(rule["expected"])
        actual = next((normalize_hex(item) for item in actual_protocols if normalize_hex(item)), "")
        if not actual or actual == expected:
            return
        entry = {
            "type": "runtime_protocol_mismatch",
            "field": rule["field"],
            "case_id": case_id,
            "label": rule["label"],
            "source_value": expected,
            "runtime_value": actual,
            "note": "现网实测协议与词表/生成用例期望不一致，请同步回需求表、词表和结果摘要。",
        }
        if entry not in self.runtime_gaps:
            self.runtime_gaps.append(entry)

    def _normalized_gaps(self) -> list[dict[str, str]]:
        config_gaps = self.config.get("gaps", [])
        merged: list[dict[str, str]] = []
        seen: set[tuple[str, str, str, str]] = set()
        for gap in [*config_gaps, *self.runtime_gaps]:
            normalized = {
                "type": str(gap.get("type", "")),
                "field": str(gap.get("field", "")),
                "case_id": str(gap.get("case_id", "")),
                "label": str(gap.get("label", "")),
                "source_value": normalize_hex(gap.get("source_value", "")) or str(gap.get("source_value", "")),
                "runtime_value": normalize_hex(gap.get("runtime_value", "")) or str(gap.get("runtime_value", "")),
                "note": str(gap.get("note", "")),
            }
            key = (
                normalized["case_id"],
                normalized["field"],
                normalized["source_value"],
                normalized["runtime_value"],
            )
            if key in seen:
                continue
            seen.add(key)
            merged.append(normalized)
        return merged

    def classify_case_result(self, item: dict[str, Any]) -> str:
        if item["status"] == "PASS":
            return ""
        text_parts = [str(item.get("detail", "")), str(item.get("expected_protocol", "")), str(item.get("expected_tone_id", ""))]
        for action in item.get("actions", []):
            text_parts.extend(
                [
                    str(action.get("action_type", "")),
                    str(action.get("detail", "")),
                    str(action.get("action_label", "")),
                ]
            )
            if action.get("action_type") == "inject_protocol" and not normalize_hex(action.get("actual_protocols", [])):
                pass
        combined = " | ".join(part for part in text_parts if part)
        combined_lower = combined.lower()
        gap_case_ids = {gap.get("case_id", "") for gap in self._normalized_gaps()}

        if any(action.get("action_type") == "inject_protocol" and not action.get("detail") for action in item.get("actions", [])):
            return "generator issue"
        if any(
            action.get("action_type") == "inject_protocol"
            and "缺少协议内容" in str(action.get("detail", ""))
            for action in item.get("actions", [])
        ):
            return "generator issue"
        if item["status"] == "BLOCKED":
            if "依赖用例" in combined:
                return "dependency blocked"
            return "environment issue"
        if item.get("case_id", "") in gap_case_ids:
            return "requirement/expectation issue"
        if any(keyword in combined for keyword in ["异常重启", "Boot Reason"]):
            return "firmware issue"
        if any(keyword in combined for keyword in ["未知动作", "协议注入异常退出", "runner", "truncation"]):
            return "runner capability issue"
        if any(keyword in combined for keyword in ["词表/需求", "协议不符", "播报ID不符", "现网/运行"]):
            return "requirement/expectation issue"
        if any(
            keyword in combined
            for keyword in ["串口", "日志", "TTS", "音频", "未接入", "播放失败", "loglevel", "COM", "恢复失败"]
        ):
            return "environment issue"
        if any(keyword in combined_lower for keyword in ["save config", "work mode", "curtain", "volume", "reboot"]):
            return "firmware issue"
        return "firmware issue"

    def build_failure_category_summary(self, results: list[dict[str, Any]]) -> list[str]:
        categories = [
            "generator issue",
            "requirement/expectation issue",
            "runner capability issue",
            "dependency blocked",
            "environment issue",
            "firmware issue",
        ]
        grouped: dict[str, list[str]] = {category: [] for category in categories}
        for item in results:
            if item["status"] in {"PASS", "DRY_RUN"}:
                continue
            category = self.classify_case_result(item)
            if category:
                grouped.setdefault(category, []).append(item["case_id"])
        lines: list[str] = []
        for category in categories:
            case_ids = grouped.get(category, [])
            lines.append(f"- {category}: `{len(case_ids)}`" + (f" | {', '.join(case_ids)}" if case_ids else ""))
        return lines

    def _mark_case_attempt(self, result: dict[str, Any], source: str, retry_index: int = 0) -> dict[str, Any]:
        result["result_source"] = source
        result["retry_index"] = retry_index
        result.setdefault("initial_status", result["status"])
        result.setdefault("initial_detail", result["detail"])
        result.setdefault("had_rerun", False)
        result.setdefault("rerun_success", False)
        result.setdefault("rerun_attempt_count", retry_index if source == "失败重跑" else 0)
        result.setdefault("final_source", "正式执行")
        return result

    def _clone_case_result(self, result: dict[str, Any]) -> dict[str, Any]:
        return copy.deepcopy(result)

    def _merge_case_result(self, initial_result: dict[str, Any], retry_attempts: list[dict[str, Any]]) -> dict[str, Any]:
        if initial_result["status"] != "FAIL" or not retry_attempts:
            return initial_result

        success_attempt = next((item for item in retry_attempts if item["status"] == "PASS"), None)
        merged = self._clone_case_result(success_attempt or retry_attempts[-1])
        merged["result_source"] = "最终结果"
        merged["had_rerun"] = True
        merged["rerun_attempt_count"] = len(retry_attempts)
        merged["initial_status"] = initial_result["status"]
        merged["initial_detail"] = initial_result["detail"]

        if success_attempt:
            merged["status"] = "PASS"
            merged["rerun_success"] = True
            merged["final_source"] = f"失败重跑第{success_attempt['retry_index']}次"
            merged["detail"] = f"首轮失败（{initial_result['detail']}）；失败重跑第{success_attempt['retry_index']}次通过"
        else:
            merged["status"] = "FAIL"
            merged["rerun_success"] = False
            merged["final_source"] = f"失败重跑第{retry_attempts[-1]['retry_index']}次"
            merged["detail"] = (
                f"首轮失败（{initial_result['detail']}）；"
                f"失败重跑{len(retry_attempts)}次均失败；"
                f"最新失败：{retry_attempts[-1]['detail']}"
            )
        return merged

    def run_failed_case_reruns(self, initial_results: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        rerun_count = max(0, int(getattr(self.args, "failed_case_reruns", 2)))
        if rerun_count <= 0:
            return initial_results, []

        case_map = {case_item["case_id"]: case_item for case_item in self.selected_cases}
        final_results: list[dict[str, Any]] = []
        retry_results: list[dict[str, Any]] = []

        for item in initial_results:
            if item["status"] != "FAIL":
                final_results.append(item)
                continue

            case_item = case_map.get(item["case_id"])
            if not case_item:
                final_results.append(item)
                continue

            current_attempts: list[dict[str, Any]] = []
            for retry_index in range(1, rerun_count + 1):
                if not self.check_unexpected_reboot(f"失败重跑 {case_item['case_id']} 前"):
                    break
                self.log.info(f"[RERUN {retry_index}/{rerun_count}] 执行 {case_item['case_id']} {case_item['name']}")
                retry_result = self.run_case(case_item)
                self._mark_case_attempt(retry_result, "失败重跑", retry_index)
                retry_result["initial_status"] = item["status"]
                retry_result["initial_detail"] = item["detail"]
                retry_results.append(retry_result)
                current_attempts.append(retry_result)
                if retry_result["status"] == "PASS":
                    break

            final_results.append(self._merge_case_result(item, current_attempts))

        return final_results, retry_results

    def _action_plan_text(self, item: dict[str, Any]) -> str:
        return "\n".join(
            f"{index}. {action.get('action_label', action['action_type'])}"
            for index, action in enumerate(item["actions"], start=1)
        )

    def _action_actual_text(self, item: dict[str, Any]) -> str:
        return "\n".join(
            f"{index}. [{action['status']}] {action['detail']}"
            for index, action in enumerate(item["actions"], start=1)
        )

    def build_requirement_results(self, results: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if not self.requirement_catalog:
            return []
        result_map = {item["case_id"]: item for item in results}
        selected_case_ids = {item["case_id"] for item in self.selected_cases}
        requirement_results: list[dict[str, Any]] = []
        for requirement in self.requirement_catalog:
            requirement_id = str(requirement.get("requirement_id", "")).strip()
            mapped_case_ids = list(self.requirement_case_map.get(requirement_id, []))
            selected_mapped_case_ids = [case_id for case_id in mapped_case_ids if case_id in selected_case_ids]
            selected_results = [result_map[case_id] for case_id in selected_mapped_case_ids if case_id in result_map]
            validation_target = str(requirement.get("validation_target", "runtime_case"))
            status = "NO_CASE"
            detail = "当前未绑定自动化用例"
            if validation_target == "static_config":
                status = "NO_METHOD"
                detail = "静态配置项，本轮无自动运行期校验方法"
            elif not mapped_case_ids:
                status = "NO_CASE"
                detail = "当前未绑定自动化用例"
            elif not selected_mapped_case_ids:
                status = "NOT_SELECTED"
                detail = "本次执行未选中覆盖该需求的用例"
            elif not selected_results:
                status = "NOT_RUN"
                detail = "覆盖该需求的已选用例未产出执行结果"
            else:
                statuses = [item["status"] for item in selected_results]
                omitted_case_ids = [case_id for case_id in mapped_case_ids if case_id not in selected_mapped_case_ids]
                if any(item == "FAIL" for item in statuses):
                    status = "FAIL"
                    failures = [f"{item['case_id']}={item['detail']}" for item in selected_results if item["status"] == "FAIL"]
                    detail = "；".join(failures[:3])
                elif all(item == "PASS" for item in statuses) and not omitted_case_ids:
                    status = "PASS"
                    detail = "覆盖用例全部通过"
                elif all(item == "DRY_RUN" for item in statuses) and not omitted_case_ids:
                    status = "DRY_RUN"
                    detail = "仅完成 dry-run，未连接设备执行"
                elif any(item == "BLOCKED" for item in statuses) and not any(item == "PASS" for item in statuses):
                    status = "BLOCKED"
                    blocked = [f"{item['case_id']}={item['detail']}" for item in selected_results if item["status"] == "BLOCKED"]
                    detail = "；".join(blocked[:3])
                else:
                    status = "PARTIAL"
                    summary = [f"{item['case_id']}={item['status']}" for item in selected_results]
                    if omitted_case_ids:
                        summary.append(f"未执行覆盖用例: {', '.join(omitted_case_ids)}")
                    detail = "；".join(summary[:4])
            requirement_results.append(
                {
                    "requirement_id": requirement_id,
                    "module": str(requirement.get("module", "")),
                    "requirement": str(requirement.get("requirement", "")),
                    "acceptance_level": str(requirement.get("acceptance_level", "main")),
                    "source_kind": str(requirement.get("source_kind", "")),
                    "source_section": str(requirement.get("source_section", "")),
                    "validation_target": validation_target,
                    "mapped_case_ids": mapped_case_ids,
                    "selected_case_ids": selected_mapped_case_ids,
                    "case_statuses": [f"{item['case_id']}={item['status']}" for item in selected_results],
                    "status": status,
                    "detail": detail,
                }
            )
        return requirement_results

    def write_requirement_report(self, requirement_results: list[dict[str, Any]]) -> None:
        if not requirement_results:
            return
        report_path = self.result_dir / "requirement_status.md"
        counts = Counter(item["status"] for item in requirement_results)
        lines = [
            "# Requirement Status",
            "",
            f"- 时间: `{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`",
            f"- 需求总数: `{len(requirement_results)}`",
        ]
        for status in ["PASS", "FAIL", "BLOCKED", "PARTIAL", "DRY_RUN", "NO_METHOD", "NO_CASE", "NOT_SELECTED", "NOT_RUN"]:
            if counts.get(status, 0):
                lines.append(f"- {status}: `{counts[status]}`")
        lines.extend(
            [
                "",
                "| 需求ID | 模块 | 验收层级 | 结果 | 覆盖用例 | 需求描述 | 摘要 |",
                "| --- | --- | --- | --- | --- | --- | --- |",
            ]
        )
        for item in requirement_results:
            case_ids = ", ".join(item.get("selected_case_ids", []) or item.get("mapped_case_ids", []))
            requirement_text = str(item.get("requirement", "")).replace("|", "/")
            detail = str(item.get("detail", "")).replace("|", "/")
            lines.append(
                f"| {item['requirement_id']} | {item['module']} | {item['acceptance_level']} | {item['status']} | "
                f"{case_ids or '-'} | {requirement_text} | {detail or '-'} |"
            )
        report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        self.log.info(f"需求结果 Markdown 已保存：{report_path}")

    def write_results(self, results: list[dict[str, Any]], retry_results: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
        retry_results = retry_results or []
        requirement_results = self.build_requirement_results(results)
        workbook_path = self.result_dir / self.policy["resultWorkbookName"]
        wb = Workbook()
        ws = wb.active
        ws.title = "结果明细"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap = Alignment(vertical="top", wrap_text=True)

        headers = ["用例编号", "模块", "名称", "优先级", "结果", "用例耗时(s)", "摘要", "执行动作", "实际结果", "期望协议", "实际协议", "期望播报ID", "实际播报ID", "期望播报内容", "最终来源"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row_index, item in enumerate(results, start=2):
            action_plan = self._action_plan_text(item)
            actual_result = self._action_actual_text(item)
            values = [
                item["case_id"],
                item["module"],
                item["name"],
                item["priority"],
                item["status"],
                item["duration_s"],
                item["detail"],
                action_plan,
                actual_result,
                item["expected_protocol"],
                "\n".join(item["actual_protocols"]),
                item["expected_tone_id"],
                "\n".join(item["actual_tones"]),
                item["expected_tone_text"],
                item.get("final_source", "正式执行"),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col, value=value)
                cell.border = border
                cell.alignment = center if col in {1, 2, 4, 5, 6, 10, 12, 15} else wrap

        retry_ws = wb.create_sheet("失败重跑")
        retry_headers = ["用例编号", "模块", "名称", "重跑序号", "结果", "用例耗时(s)", "首轮摘要", "本次摘要", "执行动作", "实际结果", "期望协议", "实际协议", "期望播报ID", "实际播报ID", "期望播报内容"]
        for col, header in enumerate(retry_headers, start=1):
            cell = retry_ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row_index, item in enumerate(retry_results, start=2):
            values = [
                item["case_id"],
                item["module"],
                item["name"],
                item.get("retry_index", 0),
                item["status"],
                item["duration_s"],
                item.get("initial_detail", ""),
                item["detail"],
                self._action_plan_text(item),
                self._action_actual_text(item),
                item["expected_protocol"],
                "\n".join(item["actual_protocols"]),
                item["expected_tone_id"],
                "\n".join(item["actual_tones"]),
                item["expected_tone_text"],
            ]
            for col, value in enumerate(values, start=1):
                cell = retry_ws.cell(row=row_index, column=col, value=value)
                cell.border = border
                cell.alignment = center if col in {1, 2, 4, 5, 6, 11, 13} else wrap

        action_ws = wb.create_sheet("动作明细")
        action_headers = ["来源", "重跑序号", "用例编号", "动作序号", "动作类型", "执行动作", "动作结果", "总耗时(s)", "识别耗时(s)", "实际ASR", "实际播报ID", "实际协议", "实际结果"]
        for col, header in enumerate(action_headers, start=1):
            cell = action_ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row_index = 2
        for item in [*results, *retry_results]:
            for index, action in enumerate(item["actions"], start=1):
                values = [
                    item.get("result_source", "正式执行"),
                    item.get("retry_index", 0),
                    item["case_id"],
                    index,
                    action["action_type"],
                    action.get("action_label", action["action_type"]),
                    action["status"],
                    action["duration_s"],
                    action.get("recognition_latency_s"),
                    "\n".join(action["actual_asr"]),
                    "\n".join(action["actual_tones"]),
                    "\n".join(action["actual_protocols"]),
                    action["detail"],
                ]
                for col, value in enumerate(values, start=1):
                    cell = action_ws.cell(row=row_index, column=col, value=value)
                    cell.border = border
                    cell.alignment = center if col in {1, 2, 3, 4, 5, 7, 8, 9} else wrap
                row_index += 1

        stats_ws = wb.create_sheet("统计")
        stats_ws.cell(row=1, column=1, value="状态").fill = header_fill
        stats_ws.cell(row=1, column=2, value="数量").fill = header_fill
        stats_ws.cell(row=1, column=1).font = header_font
        stats_ws.cell(row=1, column=2).font = header_font
        for row_index, (status, count) in enumerate(sorted(Counter(item["status"] for item in results).items()), start=2):
            stats_ws.cell(row=row_index, column=1, value=status).border = border
            stats_ws.cell(row=row_index, column=2, value=count).border = border

        case_widths = {1: 16, 2: 16, 3: 24, 4: 10, 5: 10, 6: 12, 7: 40, 8: 42, 9: 42, 10: 28, 11: 28, 12: 18, 13: 24, 14: 24, 15: 16}
        for col, width in case_widths.items():
            ws.column_dimensions[chr(64 + col)].width = width
        retry_widths = {1: 16, 2: 16, 3: 24, 4: 10, 5: 10, 6: 12, 7: 34, 8: 38, 9: 42, 10: 42, 11: 28, 12: 28, 13: 18, 14: 24, 15: 24}
        for col, width in retry_widths.items():
            retry_ws.column_dimensions[chr(64 + col)].width = width
        action_widths = {"A": 12, "B": 10, "C": 16, "D": 10, "E": 14, "F": 40, "G": 10, "H": 12, "I": 12, "J": 24, "K": 20, "L": 28, "M": 40}
        for col, width in action_widths.items():
            action_ws.column_dimensions[col].width = width
        stats_ws.column_dimensions["A"].width = 12
        stats_ws.column_dimensions["B"].width = 10

        requirement_ws = wb.create_sheet("需求结果")
        requirement_headers = ["需求ID", "模块", "验收层级", "结果", "需求描述", "摘要", "覆盖用例", "本次执行用例", "来源类型", "来源章节", "验证目标"]
        for col, header in enumerate(requirement_headers, start=1):
            cell = requirement_ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row_index, item in enumerate(requirement_results, start=2):
            values = [
                item["requirement_id"],
                item["module"],
                item["acceptance_level"],
                item["status"],
                item["requirement"],
                item["detail"],
                "\n".join(item["mapped_case_ids"]),
                "\n".join(item["selected_case_ids"]),
                item["source_kind"],
                item["source_section"],
                item["validation_target"],
            ]
            for col, value in enumerate(values, start=1):
                cell = requirement_ws.cell(row=row_index, column=col, value=value)
                cell.border = border
                cell.alignment = center if col in {1, 2, 3, 4, 11} else wrap
        requirement_widths = {"A": 16, "B": 14, "C": 12, "D": 12, "E": 42, "F": 48, "G": 22, "H": 22, "I": 16, "J": 26, "K": 16}
        for col, width in requirement_widths.items():
            requirement_ws.column_dimensions[col].width = width

        wb.save(workbook_path)
        self.log.info(f"结果 Excel 已保存：{workbook_path}")
        self.write_requirement_report(requirement_results)
        return requirement_results

    def write_summary(
        self,
        results: list[dict[str, Any]],
        title: str,
        retry_results: list[dict[str, Any]] | None = None,
        requirement_results: list[dict[str, Any]] | None = None,
    ) -> None:
        retry_results = retry_results or []
        requirement_results = requirement_results or []
        summary_path = self.result_dir / self.policy["summaryName"]
        counts = Counter(item["status"] for item in results)
        lines = [
            f"# {title}",
            "",
            f"- 时间: `{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`",
            f"- 用例总数: `{len(results)}`",
            f"- PASS: `{counts.get('PASS', 0)}`",
            f"- FAIL: `{counts.get('FAIL', 0)}`",
            f"- BLOCKED: `{counts.get('BLOCKED', 0)}`",
            f"- DRY_RUN: `{counts.get('DRY_RUN', 0)}`",
            "",
        ]
        if self.serial_checks or self.capabilities:
            lines.append("## Preflight")
            lines.append("")
            for name, info in sorted(self.serial_checks.items()):
                status = "READY" if info["supported"] else "BLOCKED"
                lines.append(f"- {name}: `{status}` {info['reason']}")
            for name, info in sorted(self.capabilities.items()):
                status = "READY" if info["supported"] else "BLOCKED"
                lines.append(f"- {name}: `{status}` {info['reason']}")
            lines.append("")
        if retry_results:
            rerun_case_ids = sorted({item["case_id"] for item in retry_results})
            rerun_success = [item for item in results if item.get("had_rerun") and item.get("rerun_success")]
            rerun_fail = [item for item in results if item.get("had_rerun") and not item.get("rerun_success")]
            lines.append("## 失败重跑")
            lines.append("")
            lines.append(f"- 首轮 FAIL 用例数: `{len(rerun_case_ids)}`")
            lines.append(f"- 实际执行重跑次数: `{len(retry_results)}`")
            lines.append(f"- 重跑后转 PASS: `{len(rerun_success)}`")
            lines.append(f"- 重跑后仍 FAIL: `{len(rerun_fail)}`")
            lines.append("")
        if requirement_results:
            requirement_counts = Counter(item["status"] for item in requirement_results)
            lines.append("## 需求结果")
            lines.append("")
            for status in ["PASS", "FAIL", "BLOCKED", "PARTIAL", "DRY_RUN", "NO_METHOD", "NO_CASE", "NOT_SELECTED", "NOT_RUN"]:
                if requirement_counts.get(status, 0):
                    lines.append(f"- {status}: `{requirement_counts[status]}`")
            lines.append("- 详情见：`requirement_status.md` 与 `testResult.xlsx` / `需求结果`")
            lines.append("")
        failure_lines = self.build_failure_category_summary(results)
        if failure_lines:
            lines.append("## 失败分类")
            lines.append("")
            lines.extend(failure_lines)
            lines.append("")
        all_gaps = self._normalized_gaps()
        if all_gaps:
            lines.append("## 文档/词表差异")
            lines.append("")
            for gap in all_gaps:
                label = gap.get("label") or gap.get("field") or gap.get("type") or "unknown_gap"
                case_id = gap.get("case_id", "")
                prefix = f"{case_id} / " if case_id else ""
                source_value = gap.get("source_value", "")
                runtime_value = gap.get("runtime_value", "")
                note = gap.get("note", "")
                lines.append(f"- {prefix}{label}: `词表/需求={source_value}` | `现网/运行={runtime_value}` | {note}")
            lines.append("")
        lines.append("## 停测规则")
        lines.append("")
        lines.append("- 烧录最多重试 3 次，3 次失败则退出整个任务")
        lines.append("- 烧录后设备无法正常启动则停止测试")
        lines.append("- 启动阶段或执行后日志无法恢复则停止测试")
        lines.append("- 设备异常重启超过 3 次则立即停止测试")
        lines.append(f"- 连续 {self.consecutive_wake_failure_limit()} 条用例唤醒失败则立即停止测试")
        lines.append("- 串口日志协议截断最多重试 2 次，超过后判定失败")
        lines.append("")
        summary_path.write_text("\n".join(lines), encoding="utf-8")
        self.archive_workspace_artifacts()

    def run(self) -> int:
        self.log.info(f"项目：{self.config['projectInfo']}")
        self.log.info(f"结果目录：{self.result_dir}")
        self.log.info(f"待执行用例数：{len(self.selected_cases)}")
        if not self.selected_cases:
            requirement_results = self.write_results([], [])
            self.write_summary([], "没有匹配到待执行用例", [], requirement_results)
            return 0
        if not self.ensure_audio():
            blocked_results = self._blocked_results_for_selected_cases("音频准备失败", "预检阻塞")
            requirement_results = self.write_results(blocked_results, [])
            self.write_summary(blocked_results, "音频准备失败", [], requirement_results)
            return 1
        if self.args.dry_run:
            results = []
            for case_item in self.selected_cases:
                result = {
                    "case_id": case_item["case_id"],
                    "module": case_item["module"],
                    "name": case_item["name"],
                    "priority": case_item["priority"],
                    "status": "DRY_RUN",
                    "detail": "dry-run 未实际执行",
                    "duration_s": 0.0,
                    "expected_protocol": case_item["expected_protocol"],
                    "expected_tone_id": case_item["expected_tone_id"],
                    "expected_tone_text": case_item["expected_tone_text"],
                    "actual_protocols": [],
                    "actual_tones": [],
                    "actions": [],
                }
                self._mark_case_attempt(result, "正式执行", 0)
                results.append(result)
            requirement_results = self.write_results(results, [])
            self.write_summary(results, "dry-run 完成", [], requirement_results)
            return 0
        if not self.connect():
            blocked_results = self._blocked_results_for_selected_cases("串口连接失败", "预检阻塞")
            requirement_results = self.write_results(blocked_results, [])
            self.write_summary(blocked_results, "串口连接失败", [], requirement_results)
            return 2
        try:
            if not self.ensure_logs_available("启动日志检查"):
                self.stop_reason = "启动阶段日志不可用，loglevel 4 重试失败"
                blocked_results = self._blocked_results_for_selected_cases(self.stop_reason, "启动预检阻塞")
                requirement_results = self.write_results(blocked_results, [])
                self.write_summary(blocked_results, self.stop_reason, [], requirement_results)
                return 3
            self.run_preflights()
            if self.stop_reason:
                blocked_results = self._blocked_results_for_selected_cases(self.stop_reason, "预检阻塞")
                requirement_results = self.write_results(blocked_results, [])
                self.write_summary(blocked_results, self.stop_reason, [], requirement_results)
                return 4
            initial_results = []
            self.sync_reboot_baseline("启动预检完成")
            for index, case_item in enumerate(self.selected_cases, start=1):
                if not self.check_unexpected_reboot(f"执行 {case_item['case_id']} 前"):
                    break
                self.log.info(f"[{index}/{len(self.selected_cases)}] 执行 {case_item['case_id']} {case_item['name']}")
                result = self.run_case(case_item)
                self._mark_case_attempt(result, "正式执行", 0)
                initial_results.append(result)
                self.case_results_by_id[result["case_id"]] = result
                if not self.note_consecutive_wake_failures(result):
                    break
                if self.stop_reason:
                    self.log.error(f"停止测试：{self.stop_reason}")
                    break
                if self.reader and not self.ensure_logs_available(f"{case_item['case_id']} 执行后日志检查"):
                    self.stop_reason = f"{case_item['case_id']} 执行后日志不可用，停止测试"
                    self.log.error(self.stop_reason)
                    break
            results, retry_results = self.run_failed_case_reruns(initial_results) if not self.stop_reason else (initial_results, [])
            if self.stop_reason:
                results = self._append_unexecuted_blocked_cases(results, self.stop_reason, "停止后未执行")
            for result in results:
                self.case_results_by_id[result["case_id"]] = result
            requirement_results = self.write_results(results, retry_results)
            title = self.stop_reason or "执行完成"
            self.write_summary(results, title, retry_results, requirement_results)
            return 0 if (not self.stop_reason and all(item["status"] in {"PASS", "BLOCKED"} for item in results)) else 4
        finally:
            self.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Dooya curtain case-first runner")
    parser.add_argument("-f", "--config", default=str(DEFAULT_CONFIG), help="deviceInfo 配置路径")
    parser.add_argument("-c", "--cases", default=str(DEFAULT_CASES), help="cases.json 路径")
    parser.add_argument("--case-id", default="", help="用例编号筛选，逗号分隔")
    parser.add_argument("--module", default="", help="模块筛选，逗号分隔")
    parser.add_argument("--priority", default="", help="优先级筛选，逗号分隔")
    parser.add_argument("--limit", type=int, default=0, help="最多执行前 N 条")
    parser.add_argument("--dry-run", action="store_true", help="仅做静态校验，不连接设备")
    parser.add_argument("--manual-block-as-fail", action="store_true", help="将 manual_power_cycle 视为失败")
    parser.add_argument("--quiet", action="store_true", help="减少控制台输出")
    parser.add_argument("--result-dir", default="", help="指定本次执行结果目录")
    parser.add_argument("--ctrl-port", default="", help="覆盖上下电控制串口")
    parser.add_argument("--ctrl-baud", type=int, default=0, help="覆盖上下电控制串口波特率")
    parser.add_argument("--log-port", default="", help="覆盖日志串口")
    parser.add_argument("--log-baud", type=int, default=0, help="覆盖日志串口波特率")
    parser.add_argument("--uart1-port", default="", help="覆盖协议串口")
    parser.add_argument("--uart1-baud", type=int, default=0, help="覆盖协议串口波特率")
    parser.add_argument("--uart1-frame-header", default="", help="覆盖协议帧头，例如 55 AA")
    parser.add_argument("--uart1-frame-length", type=int, default=0, help="覆盖协议帧长度")
    parser.add_argument("--failed-case-reruns", type=int, default=2, help="首轮 FAIL 用例在收尾阶段最多重跑次数")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    runner = DooyaRunner(Path(args.config), Path(args.cases), args)
    raise SystemExit(runner.run())


if __name__ == "__main__":
    main()
