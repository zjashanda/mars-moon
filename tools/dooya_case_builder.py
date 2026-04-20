#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Build requirement-driven executable test cases for the Dooya curtain project."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
WORK_DIR = ROOT / "work"
GENERATED_DIR = ROOT / "generated"
MODULE_ORDER = [
    "BOOT",
    "BASE",
    "VOL",
    "UART",
    "POWER",
    "FACTORY",
    "WAKEWORD",
    "WORKMODE",
    "CURTAINMODE",
    "CURTAIN",
    "PHRASE",
    "SELECTOR",
    "CTRL",
]

CURTAIN_INTENT_GROUPS = {
    "default_open": ["打开窗帘", "窗帘打开"],
    "default_close": ["窗帘关闭", "关闭窗帘"],
    "default_stop": ["停止窗帘", "窗帘停止"],
    "cloth_open": ["打开布帘", "布帘打开"],
    "cloth_close": ["关闭布帘", "布帘关闭"],
    "cloth_stop": ["停止布帘", "布帘停止"],
    "sheer_open": ["打开纱帘", "纱帘打开"],
    "sheer_close": ["关闭纱帘", "纱帘关闭"],
    "sheer_stop": ["停止纱帘", "纱帘停止"],
    "window_sheer_open": ["打开窗纱", "窗纱打开"],
    "window_sheer_close": ["窗纱关闭", "关闭窗纱"],
    "window_sheer_stop": ["停止窗纱", "窗纱停止"],
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def norm_hex(value: str | None) -> str:
    return " ".join(str(value or "").upper().split())


def norm_text(value: str | None, fallback: str = "") -> str:
    text = str(value or "").strip()
    return text or fallback


def tone_lookup(spec: dict[str, Any]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for item in spec.get("tones", {}).get("items", []):
        lookup[str(item.get("name", "")).strip()] = str(item.get("text", "")).strip()
    return lookup


def case(
    case_id: str,
    module: str,
    name: str,
    precondition: str,
    steps: str,
    expected_result: str,
    expected_protocol: str,
    expected_tone_id: str,
    expected_tone_text: str,
    priority: str,
    actions: list[dict[str, Any]],
    requirement_ids: list[str],
    case_type: str,
    blocked_by_case_ids: list[str] | None = None,
    **extra: Any,
) -> dict[str, Any]:
    payload = {
        "case_id": case_id,
        "module": module,
        "name": name,
        "precondition": precondition,
        "steps": render_case_steps(actions),
        "design_steps": steps,
        "expected_result": expected_result,
        "expected_protocol": expected_protocol,
        "expected_tone_id": expected_tone_id,
        "expected_tone_text": expected_tone_text,
        "priority": priority,
        "requirement_ids": requirement_ids,
        "case_type": case_type,
        "actions": actions,
        "blocked_by_case_ids": blocked_by_case_ids or [],
    }
    payload.update(extra)
    return payload


def wake_action(
    word: str,
    expect_tone_id: str = "TONE_ID_0",
    retries: int = 10,
    expect_send_protocol: str | None = None,
) -> dict[str, Any]:
    return {
        "type": "wake",
        "word": word,
        "expect_tone_id": expect_tone_id,
        "retries": retries,
        "require_any_tone": False,
        "expect_send_protocol": norm_hex(expect_send_protocol),
    }


def say_action(
    word: str,
    expect_asr: list[str] | None = None,
    expect_tone_id: str | None = None,
    expect_tone_text: str | None = None,
    advisory_tone_id: str | None = None,
    advisory_tone_text: str | None = None,
    expect_send_protocol: str | None = None,
    expect_log_values: dict[str, list[str] | str] | None = None,
    advisory_log_values: dict[str, list[str] | str] | None = None,
    forbid_log_values: dict[str, list[str] | str] | None = None,
    expect_no_response: bool = False,
    expect_any_tone: bool = False,
    retries: int = 3,
    observe_s: float = 3.0,
    allow_fail_continue: bool = False,
    advisory_failure: bool = False,
    require_protocol_and_tone: bool = True,
    auto_wake: bool = True,
) -> dict[str, Any]:
    return {
        "type": "say",
        "word": word,
        "expect_asr": expect_asr or [word],
        "expect_tone_id": expect_tone_id or "",
        "expect_tone_text": expect_tone_text or "",
        "advisory_tone_id": advisory_tone_id or "",
        "advisory_tone_text": advisory_tone_text or "",
        "expect_send_protocol": norm_hex(expect_send_protocol),
        "expect_log_values": expect_log_values or {},
        "advisory_log_values": advisory_log_values or {},
        "forbid_log_values": forbid_log_values or {},
        "expect_no_response": expect_no_response,
        "expect_any_tone": expect_any_tone,
        "retries": retries,
        "observe_s": observe_s,
        "allow_fail_continue": allow_fail_continue,
        "advisory_failure": advisory_failure,
        "require_protocol_and_tone": require_protocol_and_tone,
        "auto_wake": auto_wake,
    }


def inject_protocol_action(
    protocol: str,
    expect_tone_id: str | None = None,
    expect_tone_text: str | None = None,
    expect_recv_protocol: str | None = None,
    require_receive_msg: bool = False,
    requires_active_window: bool = False,
    observe_s: float = 3.0,
    expect_no_response: bool = False,
) -> dict[str, Any]:
    return {
        "type": "inject_protocol",
        "protocol": norm_hex(protocol),
        "expect_tone_id": expect_tone_id or "",
        "expect_tone_text": expect_tone_text or "",
        "expect_recv_protocol": norm_hex(expect_recv_protocol),
        "require_receive_msg": require_receive_msg,
        "requires_active_window": requires_active_window,
        "observe_s": observe_s,
        "expect_no_response": expect_no_response,
    }


def wait_action(
    seconds: float,
    expect_tone_id: str | None = None,
    expect_markers: list[str] | None = None,
    tolerance_s: float = 1.5,
    measure_timeout: bool = False,
    validate_duration: bool = True,
) -> dict[str, Any]:
    return {
        "type": "wait",
        "seconds": seconds,
        "expect_tone_id": expect_tone_id or "",
        "expect_markers": expect_markers or [],
        "tolerance_s": tolerance_s,
        "measure_timeout": measure_timeout,
        "validate_duration": validate_duration,
    }


def manual_power_cycle_action(
    note: str = "执行上下电循环",
    off_wait_s: float = 1.0,
    boot_wait_s: float = 8.0,
    expect_tone_id: str | None = None,
) -> dict[str, Any]:
    return {
        "type": "manual_power_cycle",
        "note": note,
        "off_wait_s": off_wait_s,
        "boot_wait_s": boot_wait_s,
        "expect_tone_id": expect_tone_id or "",
    }


def phrase_check_action(items: list[dict[str, Any]], wake_word: str, observe_s: float = 1.8, inter_word_gap_s: float = 0.2) -> dict[str, Any]:
    return {
        "type": "phrase_check",
        "items": items,
        "wake_word": wake_word,
        "observe_s": observe_s,
        "inter_word_gap_s": inter_word_gap_s,
        "timeout_markers": ["TIME_OUT"],
        "timeout_tone_id": "TONE_ID_2",
    }


def wake_repeat_action(
    word: str,
    repeats: int = 3,
    expect_tone_id: str = "TONE_ID_0",
    *,
    idle_timeout_s: float | None = None,
    timeout_tone_id: str = "TONE_ID_2",
    timeout_markers: list[str] | None = None,
) -> dict[str, Any]:
    action = {
        "type": "assert_wake_repeats",
        "word": word,
        "repeats": repeats,
        "expect_tone_id": expect_tone_id,
    }
    if idle_timeout_s is not None and float(idle_timeout_s) > 0:
        action.update(
            {
                "wait_for_idle_between_repeats": True,
                "idle_timeout_s": float(idle_timeout_s),
                "timeout_tone_id": timeout_tone_id,
                "timeout_markers": timeout_markers or ["TIME_OUT"],
            }
        )
    return action


def no_wake_action(word: str | list[str], expect_silence_s: float = 2.0, repeats: int = 1) -> dict[str, Any]:
    words = [item for item in (word if isinstance(word, list) else [word]) if str(item).strip()]
    return {
        "type": "assert_no_wake",
        "word": words[0] if words else "",
        "words": words,
        "expect_silence_s": expect_silence_s,
        "observe_s": expect_silence_s,
        "repeats": repeats,
        # Runner currently reads expect_tone_id to decide which wake prompt is forbidden.
        "expect_tone_id": "TONE_ID_0",
        "forbidden_tone_id": "TONE_ID_0",
    }


def with_label(action: dict[str, Any], label: str) -> dict[str, Any]:
    cloned = dict(action)
    cloned["label"] = label
    return cloned


class RequirementCaseBuilder:
    def __init__(self, spec: dict[str, Any]) -> None:
        self.spec = spec
        self.behavior_rules = spec.get("behavior_rules", {})
        self.matrix = spec.get("coverage_matrix", [])
        self.requirement_catalog = spec.get("requirement_catalog", self.matrix)
        self.word_rows = spec.get("word_sheet", {}).get("rows", [])
        self.word_lookup = {
            norm_text(row.get("semantic")): row
            for row in self.word_rows
            if norm_text(row.get("semantic"))
        }
        self.semantic_groups = spec.get("semantic_groups", {})
        self.tones = tone_lookup(spec)
        base = self.behavior_rules.get("base", {})
        self.wake_word = norm_text(base.get("default_wake_word"), "你好杜亚")
        self.default_work_mode = norm_text(self.behavior_rules.get("work_mode_setting", {}).get("default_mode"), "语音模式")
        self.default_curtain_mode = norm_text(self.behavior_rules.get("curtain_mode_setting", {}).get("default_mode"), "窗帘模式")
        self.curtain_groups = self.behavior_rules.get("curtain_control", {}).get("groups", {})
        self.case_counters = {module: 1 for module in MODULE_ORDER}
        self.requirement_map: dict[str, list[str]] = {}
        self.requirement_index: dict[str, dict[str, Any]] = {}
        for item in self.requirement_catalog:
            self.requirement_index[item["requirement_id"]] = item
        for item in self.requirement_catalog:
            self.requirement_map.setdefault(item["module"], []).append(item["requirement_id"])

    def tone_text(self, tone_id: str | None) -> str:
        if not tone_id:
            return ""
        return self.tones.get(str(tone_id).strip(), "")

    def next_case_id(self, module: str) -> str:
        index = self.case_counters[module]
        self.case_counters[module] += 1
        return f"TC_{module}_{index:03d}"

    def requirement_ids(self, module: str, *indices: int) -> list[str]:
        items = self.requirement_map.get(module, [])
        if not indices:
            return items
        result: list[str] = []
        for index in indices:
            if index <= 0:
                continue
            if index <= len(items):
                result.append(items[index - 1])
        return result

    def requirement_ids_direct(self, *requirement_ids: str) -> list[str]:
        return [item for item in requirement_ids if item in self.requirement_index]

    def accepted_asr_variants(self, word: str) -> list[str]:
        groups = [
            {"滴答模式", "嘀嗒模式"},
            {"纱帘模式", "窗纱模式"},
        ]
        variants = [word]
        for group in groups:
            if word in group:
                for item in group:
                    if item not in variants:
                        variants.append(item)
        return variants

    def setting_invalid_word(self) -> str:
        return "测试无效词"

    def recognized_invalid_word(self, fallback: str = "打开窗帘") -> str:
        for candidate in [fallback, "打开窗帘", "关闭窗帘", "小声点", "恢复出厂模式"]:
            if candidate in self.word_lookup:
                return candidate
        return fallback

    def setting_entry_actions(
        self,
        entry_word: str,
        entry_tone_id: str,
        entry_tone_text: str,
        expect_send_protocol: str | None = None,
        expect_log_values: dict[str, list[str] | str] | None = None,
        advisory_log_values: dict[str, list[str] | str] | None = None,
        allow_fail_continue: bool = False,
        retries: int = 3,
        observe_s: float = 3.0,
    ) -> list[dict[str, Any]]:
        return [
            wake_action(self.wake_word),
            say_action(
                entry_word,
                expect_asr=[entry_word],
                advisory_tone_id=entry_tone_id,
                advisory_tone_text=entry_tone_text,
                expect_send_protocol=expect_send_protocol,
                expect_log_values=expect_log_values,
                advisory_log_values=advisory_log_values,
                allow_fail_continue=allow_fail_continue,
                require_protocol_and_tone=False,
                auto_wake=False,
                retries=retries,
                observe_s=observe_s,
            ),
        ]

    def setting_success_tone_id(self) -> str:
        if self.tone_text("TONE_ID_14"):
            return "TONE_ID_14"
        if self.tone_text("TONE_ID_10"):
            return "TONE_ID_10"
        return "TONE_ID_14"

    def setting_success_tone_text(self) -> str:
        tone_id = self.setting_success_tone_id()
        return self.tone_text(tone_id) or "设置成功"

    def word_row(self, semantic: str) -> dict[str, Any]:
        return self.word_lookup.get(semantic, {})

    def first_word(self, options: list[str], fallback: str = "") -> str:
        for item in options:
            if item in self.word_lookup:
                return item
        return fallback or (options[0] if options else "")

    def row_tone_id(self, row: dict[str, Any], fallback: str) -> str:
        return norm_text(row.get("tone_id"), fallback)

    def row_tone_text(self, row: dict[str, Any], fallback_tone_id: str, fallback_text: str = "") -> str:
        tone_id = self.row_tone_id(row, fallback_tone_id)
        return norm_text(row.get("tts_text"), self.tone_text(tone_id) or fallback_text)

    def curtain_words(self, key: str, fallback: list[str]) -> list[str]:
        rows = self.curtain_groups.get(key, [])
        words = [norm_text(row.get("semantic")) for row in rows if norm_text(row.get("semantic"))]
        return words or fallback

    def curtain_primary_row(self, key: str, fallback_words: list[str]) -> dict[str, Any]:
        words = self.curtain_words(key, fallback_words)
        return self.word_row(words[0]) if words else {}

    def row_protocol(self, row: dict[str, Any]) -> str:
        return norm_hex(row.get("send_protocol"))

    def positive_command_action(
        self,
        word: str,
        *,
        accept_asr: list[str] | None = None,
        auto_wake: bool = True,
        label: str | None = None,
    ) -> dict[str, Any]:
        row = self.word_row(word)
        action = say_action(
            word,
            expect_asr=accept_asr or [word],
            expect_tone_id=self.row_tone_id(row, "TONE_ID_1"),
            expect_tone_text=self.row_tone_text(row, "TONE_ID_1", "好的"),
            expect_send_protocol=self.row_protocol(row),
            require_protocol_and_tone=True,
            auto_wake=auto_wake,
        )
        if label:
            action["label"] = label
        return action

    def config_saved_advisory_logs(self) -> dict[str, list[str]]:
        return {"configSaved": ["save config success"]}

    def merge_advisory_logs(self, *mappings: dict[str, list[str]] | None) -> dict[str, list[str]]:
        merged: dict[str, list[str]] = {}
        for mapping in mappings:
            if not mapping:
                continue
            for key, values in mapping.items():
                bucket = merged.setdefault(str(key).strip(), [])
                items = values if isinstance(values, list) else [values]
                for item in items:
                    text = str(item).strip()
                    if text and text not in bucket:
                        bucket.append(text)
        return merged

    def wakeword_config_values(self) -> list[str]:
        candidates = list(self.behavior_rules.get("wake_word_setting", {}).get("candidate_wake_words", []))
        return ["0"] + [str(index) for index, _ in enumerate(candidates, start=1)]

    def wakeword_log_expectations(self, wake_word: str) -> dict[str, list[str]]:
        normalized = norm_text(wake_word)
        if normalized == self.wake_word:
            return {"wakeup": ["0"]}
        candidates = list(self.behavior_rules.get("wake_word_setting", {}).get("candidate_wake_words", []))
        for index, candidate in enumerate(candidates, start=1):
            if norm_text(candidate) == normalized:
                return {"wakeup": [str(index)]}
        return {}

    def forbid_wakeword_change_logs(self) -> dict[str, list[str]]:
        return {"wakeup": self.wakeword_config_values()}

    def work_mode_log_expectations(self, mode_word: str) -> dict[str, list[str]]:
        mapping = {
            "语音模式": "0",
            "滴答模式": "1",
            "嘀嗒模式": "1",
        }
        value = mapping.get(norm_text(mode_word))
        return {"workMode": [value]} if value is not None else {}

    def forbid_work_mode_change_logs(self) -> dict[str, list[str]]:
        return {"workMode": ["0", "1"]}

    def curtain_mode_log_expectations(self, mode_word: str) -> dict[str, list[str]]:
        mapping = {
            "窗帘模式": "0",
            "布帘模式": "1",
            "纱帘模式": "2",
            "窗纱模式": "2",
        }
        value = mapping.get(norm_text(mode_word))
        return {"curtainMode": [value]} if value is not None else {}

    def forbid_curtain_mode_change_logs(self) -> dict[str, list[str]]:
        return {"curtainMode": ["0", "1", "2"]}

    def factory_entry_action(self) -> tuple[str, str, str, str]:
        entry = self.behavior_rules.get("factory_reset", {}).get("entry", {})
        entry_word = norm_text(entry.get("semantic"), "恢复出厂模式")
        entry_tone_id = norm_text(entry.get("tone_id"), "TONE_ID_1")
        entry_tone_text = norm_text(entry.get("tts_text"), self.tone_text(entry_tone_id) or "好的")
        entry_protocol = norm_hex(entry.get("send_protocol"))
        return entry_word, entry_tone_id, entry_tone_text, entry_protocol

    def restore_default_actions(self, note: str = "恢复默认状态") -> list[dict[str, Any]]:
        entry_word, entry_tone_id, entry_tone_text, entry_protocol = self.factory_entry_action()
        return [
            with_label(
                say_action(
                    entry_word,
                    expect_tone_id=entry_tone_id,
                    expect_tone_text=entry_tone_text,
                    expect_send_protocol=entry_protocol,
                    advisory_log_values=self.config_saved_advisory_logs(),
                    require_protocol_and_tone=False,
                ),
                f"{note}：执行【{entry_word}】",
            ),
            with_label(wait_action(2.0), "等待恢复默认状态稳定"),
        ]

    def tagged_actions(self, actions: list[dict[str, Any]], **extra: Any) -> list[dict[str, Any]]:
        tagged: list[dict[str, Any]] = []
        for action in actions:
            cloned = dict(action)
            cloned.update(extra)
            tagged.append(cloned)
        return tagged

    def cleanup_actions(self, actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return self.tagged_actions(actions, always_run=True, allow_fail_continue=True)

    def isolate_cases(
        self,
        cases: list[dict[str, Any]],
        before_note: str = "用例前恢复默认状态",
        after_note: str = "用例后恢复默认状态",
    ) -> list[dict[str, Any]]:
        isolated: list[dict[str, Any]] = []
        for item in cases:
            cloned = dict(item)
            setup_actions = (
                []
                if cloned.get("skip_default_setup")
                else self.tagged_actions(self.restore_default_actions(note=before_note), setup_action=True)
            )
            cleanup_actions = (
                []
                if cloned.get("skip_default_cleanup")
                else self.cleanup_actions(self.restore_default_actions(note=after_note))
            )
            actions = [
                *setup_actions,
                *cloned["actions"],
                *cleanup_actions,
            ]
            cloned["actions"] = actions
            cloned["steps"] = render_case_steps(actions)
            isolated.append(cloned)
        return isolated

    def wait_for_timeout_exit_action(
        self,
        timeout_s: float,
        *,
        expect_tone_id: str = "TONE_ID_2",
        expect_markers: list[str] | None = None,
        label: str = "",
    ) -> dict[str, Any]:
        max_wait_s = max(float(timeout_s) + 15.0, 35.0)
        action = wait_action(
            max_wait_s,
            expect_tone_id=expect_tone_id,
            expect_markers=expect_markers or ["TIME_OUT"],
            measure_timeout=False,
            validate_duration=False,
        )
        if label:
            return with_label(action, label)
        return action

    def build(self) -> list[dict[str, Any]]:
        cases: list[dict[str, Any]] = []
        for module in MODULE_ORDER:
            builder = getattr(self, f"build_{module.lower()}_cases")
            cases.extend(builder())
        return cases

    def build_boot_cases(self) -> list[dict[str, Any]]:
        boot_tone_id = "TONE_ID_3"
        return [
            case(
                self.next_case_id("BOOT"),
                "BOOT",
                "重启后欢迎语播报",
                "控制串口可用且设备已连接日志串口",
                "执行一次控制串口硬重启并观察上电欢迎语",
                "硬重启上电后播放欢迎语播报",
                "",
                boot_tone_id,
                self.tone_text(boot_tone_id),
                "P1",
                [manual_power_cycle_action(note="执行硬重启并观察上电欢迎语", expect_tone_id=boot_tone_id)],
                self.requirement_ids("BOOT", 1),
                "state_transition",
                manual_only=True,
                manual_reason="需人工确认硬重启上电后是否播放欢迎语播报",
            ),
            case(
                self.next_case_id("BOOT"),
                "BOOT",
                "重启后日志恢复默认等级",
                "设备已连接日志串口",
                "执行 reboot，并人工确认日志等级恢复默认值而非保持 loglevel 4",
                "重启后日志等级恢复默认值，不保持 loglevel 4",
                "",
                "",
                "",
                "P1",
                [{"type": "reboot", "wait_log_recovery_s": 20}],
                self.requirement_ids("BOOT", 2),
                "state_transition",
                manual_only=True,
                manual_reason="需人工确认 reboot 后日志等级恢复默认值，而不是继续保持 loglevel 4",
            ),
            case(
                self.next_case_id("BOOT"),
                "BOOT",
                "重启后默认唤醒词可用",
                "设备已完成启动",
                "重启后使用默认唤醒词进行唤醒",
                "默认唤醒词仍可唤醒设备",
                "",
                "TONE_ID_0",
                self.tone_text("TONE_ID_0"),
                "P1",
                [{"type": "reboot", "wait_log_recovery_s": 20}, wake_action(self.wake_word)],
                self.requirement_ids("BOOT", 3),
                "state_transition",
            ),
        ]


    def build_base_cases(self) -> list[dict[str, Any]]:
        timeout_s = float(self.behavior_rules.get("base", {}).get("wake_timeout_s", 15))
        default_open_word = self.first_word(CURTAIN_INTENT_GROUPS["default_open"], "打开窗帘")
        default_open_row = self.word_row(default_open_word)
        default_open_tone_id = self.row_tone_id(default_open_row, "TONE_ID_1")
        default_open_tone_text = self.row_tone_text(default_open_row, default_open_tone_id, "好的")
        default_open_protocol = norm_hex(default_open_row.get("send_protocol"))
        return [
            case(
                self.next_case_id("BASE"),
                "BASE",
                "默认唤醒词可进入交互态",
                "设备处于默认待命态",
                f"播放默认唤醒词【{self.wake_word}】",
                "设备进入交互态并播报唤醒提示音",
                "",
                "TONE_ID_0",
                self.tone_text("TONE_ID_0"),
                "P1",
                [wake_action(self.wake_word)],
                self.requirement_ids("BASE", 1),
                "positive",
            ),
            case(
                self.next_case_id("BASE"),
                "BASE",
                "未唤醒时直接命令不响应",
                "设备处于默认待命态",
                f"不先唤醒，直接播放普通控制词【{default_open_word}】",
                "未唤醒状态下普通控制词不应响应",
                "",
                "",
                "",
                "P1",
                [say_action(default_open_word, expect_no_response=True, auto_wake=False, observe_s=2.0)],
                self.requirement_ids_direct("DER-BASE-001"),
                "negative",
            ),
            case(
                self.next_case_id("BASE"),
                "BASE",
                "交互态超时退出",
                "设备处于默认待命态",
                f"先唤醒，再静置 {int(timeout_s)}s，观察超时退出",
                "设备在超时后退出交互态并播报超时音",
                "",
                "TONE_ID_2",
                self.tone_text("TONE_ID_2"),
                "P1",
                [
                    wake_action(self.wake_word),
                    wait_action(timeout_s, expect_tone_id="TONE_ID_2", expect_markers=["TIME_OUT"], measure_timeout=True),
                ],
                self.requirement_ids_direct("CFG-001"),
                "boundary",
            ),
            case(
                self.next_case_id("BASE"),
                "BASE",
                "超时退出后直接命令不响应",
                "设备处于默认待命态",
                f"先唤醒，再静置 {int(timeout_s)}s 超时退出，然后不重新唤醒直接说【{default_open_word}】",
                "交互态超时退出后，不重新唤醒的直接命令不应响应",
                "",
                "",
                "",
                "P1",
                [
                    wake_action(self.wake_word),
                    self.wait_for_timeout_exit_action(timeout_s, label="等待设备实际超时退出"),
                    say_action(default_open_word, expect_no_response=True, auto_wake=False, observe_s=2.0),
                ],
                self.requirement_ids("BASE", 2),
                "negative",
            ),
            case(
                self.next_case_id("BASE"),
                "BASE",
                "超时退出后重新唤醒恢复命令响应",
                "设备处于默认待命态",
                f"先唤醒并等待 {int(timeout_s)}s 超时退出，再重新唤醒后执行【{default_open_word}】",
                "超时退出后重新唤醒，普通控制命令恢复正常响应",
                default_open_protocol,
                default_open_tone_id,
                default_open_tone_text,
                "P1",
                [
                    wake_action(self.wake_word),
                    self.wait_for_timeout_exit_action(timeout_s, label="等待设备实际超时退出"),
                    wake_action(self.wake_word),
                    say_action(
                        default_open_word,
                        expect_tone_id=default_open_tone_id,
                        expect_tone_text=default_open_tone_text,
                        expect_send_protocol=default_open_protocol,
                        require_protocol_and_tone=False,
                        auto_wake=False,
                    ),
                ],
                self.requirement_ids_direct("DER-BASE-003"),
                "state_transition",
            ),
        ]


    def build_vol_cases(self) -> list[dict[str, Any]]:
        rules = self.behavior_rules.get("volume", {})
        increase_word = norm_text(rules.get("increase_word"), "大声点")
        decrease_word = norm_text(rules.get("decrease_word"), "小声点")
        step_tone_id = norm_text(rules.get("step_tone_id"), "TONE_ID_1")
        step_tone_text = self.tone_text(step_tone_id)
        upper_boundary_tone = norm_text(rules.get("upper_boundary_tone"), "TONE_ID_4")
        lower_boundary_tone = norm_text(rules.get("lower_boundary_tone"), "TONE_ID_5")
        default_level = int(rules.get("default_level", 3) or 3)
        level_count = int(rules.get("level_count", 4) or 4)
        down_steps = max(default_level - 1, 0)
        up_steps = max(level_count - default_level, 0)

        def repeated_volume_steps(word: str, tone_id: str, tone_text: str, count: int, label_prefix: str) -> list[dict[str, Any]]:
            return [
                with_label(
                    say_action(
                        word,
                        expect_tone_id=tone_id,
                        expect_tone_text=tone_text,
                        require_protocol_and_tone=False,
                    ),
                    f"{label_prefix}{index + 1}",
                )
                for index in range(count)
            ]

        return [
            case(
                self.next_case_id("VOL"),
                "VOL",
                "默认音量档位正确",
                "设备处于默认音量",
                f"从默认音量开始重复执行【{decrease_word}】，统计到最小音量前的步数",
                "到达最小音量前的步数与默认音量档位配置一致",
                "",
                lower_boundary_tone,
                self.tone_text(lower_boundary_tone),
                "P1",
                [
                    {
                        "type": "volume_walk",
                        "command_word": decrease_word,
                        "expected_step_count": down_steps,
                        "step_tone_id": step_tone_id,
                        "boundary_tone_id": lower_boundary_tone,
                        "extra_boundary_repeats": 1,
                    },
                ],
                self.requirement_ids_direct("CFG-003"),
                "boundary",
            ),
            case(
                self.next_case_id("VOL"),
                "VOL",
                "音量增大一步反馈正确",
                "设备处于默认音量",
                f"唤醒后执行【{increase_word}】一次",
                "音量增大一步后播报普通确认音",
                "",
                step_tone_id,
                step_tone_text,
                "P1",
                [say_action(increase_word, expect_tone_id=step_tone_id, expect_tone_text=step_tone_text, require_protocol_and_tone=False)],
                self.requirement_ids_direct("DER-VOL-001"),
                "positive",
            ),
            case(
                self.next_case_id("VOL"),
                "VOL",
                "音量减小一步反馈正确",
                "设备处于默认音量",
                f"唤醒后执行【{decrease_word}】一次",
                "音量减小一步后播报普通确认音",
                "",
                step_tone_id,
                step_tone_text,
                "P1",
                [say_action(decrease_word, expect_tone_id=step_tone_id, expect_tone_text=step_tone_text, require_protocol_and_tone=False)],
                self.requirement_ids_direct("DER-VOL-002"),
                "positive",
            ),
            case(
                self.next_case_id("VOL"),
                "VOL",
                "音量最大边界提示正确",
                "设备处于默认音量",
                f"重复执行【{increase_word}】直到最大音量，再继续执行一次观察边界提示",
                "达到最大音量后继续增大，不再播普通确认音，而是播报最大边界提示",
                "",
                upper_boundary_tone,
                self.tone_text(upper_boundary_tone),
                "P1",
                [
                    *repeated_volume_steps(increase_word, step_tone_id, step_tone_text, up_steps, "执行增大音量步骤"),
                    with_label(
                        say_action(
                            increase_word,
                            expect_tone_id=upper_boundary_tone,
                            expect_tone_text=self.tone_text(upper_boundary_tone),
                            require_protocol_and_tone=False,
                        ),
                        "继续增大音量，验证最大边界提示",
                    ),
                ],
                self.requirement_ids_direct("CFG-005"),
                "boundary",
            ),
            case(
                self.next_case_id("VOL"),
                "VOL",
                "音量最小边界提示正确",
                "设备处于默认音量",
                f"重复执行【{decrease_word}】直到最小音量，再继续执行一次观察边界提示",
                "达到最小音量后继续减小，不再播普通确认音，而是播报最小边界提示",
                "",
                lower_boundary_tone,
                self.tone_text(lower_boundary_tone),
                "P1",
                [
                    *repeated_volume_steps(decrease_word, step_tone_id, step_tone_text, down_steps, "执行减小音量步骤"),
                    with_label(
                        say_action(
                            decrease_word,
                            expect_tone_id=lower_boundary_tone,
                            expect_tone_text=self.tone_text(lower_boundary_tone),
                            require_protocol_and_tone=False,
                        ),
                        "继续减小音量，验证最小边界提示",
                    ),
                ],
                self.requirement_ids_direct("CFG-004"),
                "boundary",
            ),
            case(
                self.next_case_id("VOL"),
                "VOL",
                "从默认音量到最大档位步数正确",
                "设备处于默认音量",
                f"从默认音量开始重复执行【{increase_word}】，统计到最大音量前的步进次数",
                "默认音量到最大档位的步进次数符合档位配置",
                "",
                upper_boundary_tone,
                self.tone_text(upper_boundary_tone),
                "P1",
                [
                    {
                        "type": "volume_walk",
                        "command_word": increase_word,
                        "expected_step_count": up_steps,
                        "step_tone_id": step_tone_id,
                        "boundary_tone_id": upper_boundary_tone,
                        "extra_boundary_repeats": 1,
                    },
                ],
                self.requirement_ids_direct("CFG-002", "CFG-003"),
                "boundary",
            ),
            case(
                self.next_case_id("VOL"),
                "VOL",
                "从默认音量到最小档位步数正确",
                "设备处于默认音量",
                f"从默认音量开始重复执行【{decrease_word}】，统计到最小音量前的步进次数",
                "默认音量到最小档位的步进次数符合档位配置",
                "",
                lower_boundary_tone,
                self.tone_text(lower_boundary_tone),
                "P1",
                [
                    {
                        "type": "volume_walk",
                        "command_word": decrease_word,
                        "expected_step_count": down_steps,
                        "step_tone_id": step_tone_id,
                        "boundary_tone_id": lower_boundary_tone,
                        "extra_boundary_repeats": 1,
                    },
                ],
                self.requirement_ids_direct("CFG-002", "CFG-003"),
                "boundary",
            ),
        ]

    def build_uart_cases(self) -> list[dict[str, Any]]:
        stable_word = self.first_word(CURTAIN_INTENT_GROUPS["default_open"], "打开窗帘")
        stable_row = self.word_row(stable_word)
        protocol = norm_hex(stable_row.get("send_protocol"))
        tone_id = self.row_tone_id(stable_row, "TONE_ID_1")
        tone_text = self.row_tone_text(stable_row, tone_id, "好的")
        default_setup_actions = self.restore_default_actions(note="用例前恢复默认状态")
        clear_row = self.behavior_rules.get("remote_pairing", {}).get("clear", {})
        clear_word = norm_text(clear_row.get("semantic"), "清除遥控器")
        clear_tone_id = norm_text(clear_row.get("tone_id"), "TONE_ID_1")
        clear_tone_text = norm_text(clear_row.get("tts_text"), self.tone_text(clear_tone_id) or "好的")
        clear_protocol = norm_hex(clear_row.get("send_protocol"))
        return [
            case(
                self.next_case_id("UART"),
                "UART",
                "命令词触发协议发送",
                "日志口与协议口已连接",
                f"唤醒后说稳定控制词【{stable_word}】，检查协议口接收到发送协议",
                "命令词执行后协议口观察到目标协议",
                protocol,
                tone_id,
                tone_text,
                "P1",
                [
                    *default_setup_actions,
                    say_action(
                        stable_word,
                        expect_tone_id=tone_id,
                        expect_tone_text=tone_text,
                        expect_send_protocol=protocol,
                        require_protocol_and_tone=False,
                    )
                ],
                self.requirement_ids_direct("CFG-009", "DER-UART-001"),
                "positive",
            ),
            case(
                self.next_case_id("UART"),
                "UART",
                "命令词识别日志字段完整",
                "日志口与协议口已连接",
                f"唤醒后说稳定控制词【{stable_word}】，检查日志口至少出现 asrKw 和播报 playId",
                "日志口出现关键识别字段，能证明语音链路和播报链路完整",
                protocol,
                tone_id,
                tone_text,
                "P1",
                [
                    *default_setup_actions,
                    say_action(
                        stable_word,
                        expect_tone_id=tone_id,
                        expect_tone_text=tone_text,
                        expect_send_protocol=protocol,
                        expect_log_values={"asrKw": [stable_word]},
                        require_protocol_and_tone=False,
                    )
                ],
                self.requirement_ids_direct("CFG-011", "DER-UART-002"),
                "positive",
            ),
            case(
                self.next_case_id("UART"),
                "UART",
                "清除遥控器链路可观测",
                "日志口与协议口已连接",
                f"唤醒后说【{clear_word}】，检查日志口识别结果和协议口输出",
                "清除遥控器命令同时具备识别日志与协议输出",
                clear_protocol,
                clear_tone_id,
                clear_tone_text,
                "P2",
                [
                    *default_setup_actions,
                    say_action(
                        clear_word,
                        expect_tone_id=clear_tone_id,
                        expect_tone_text=clear_tone_text,
                        expect_send_protocol=clear_protocol,
                        expect_log_values={"asrKw": [clear_word]},
                        require_protocol_and_tone=False,
                    )
                ],
                self.requirement_ids_direct("CFG-009", "DER-UART-003"),
                "positive",
            ),
        ]


    def build_power_cases(self) -> list[dict[str, Any]]:
        wake_rules = self.behavior_rules.get("wake_word_setting", {})
        power_rules = self.behavior_rules.get("power", {})
        volume_rules = self.behavior_rules.get("volume", {})
        alternate_wakes = [item for item in wake_rules.get("candidate_wake_words", []) if norm_text(item)]
        changed_wake = alternate_wakes[0] if alternate_wakes else "客厅窗帘"
        other_wake = alternate_wakes[1] if len(alternate_wakes) > 1 else "餐厅窗帘"
        wake_entry = wake_rules.get("entry", {})
        wake_entry_word = norm_text(wake_entry.get("semantic"), "设置唤醒词")
        wake_entry_tone_id = norm_text(wake_entry.get("tone_id"), "TONE_ID_9")
        wake_entry_tone_text = norm_text(wake_entry.get("tts_text"), self.tone_text(wake_entry_tone_id))
        changed_wake_tone_id = self.setting_success_tone_id()
        changed_wake_tone_text = self.setting_success_tone_text()
        decrease_word = norm_text(volume_rules.get("decrease_word"), "小声点")
        step_tone_id = norm_text(volume_rules.get("step_tone_id"), "TONE_ID_1")
        default_level = int(volume_rules.get("default_level", 3) or 3)
        retained_down_steps = max(default_level - 2, 0)
        expected_down_steps = retained_down_steps if power_rules.get("volume_retained_after_power_cycle") else max(default_level - 1, 0)
        lower_boundary_tone = norm_text(volume_rules.get("lower_boundary_tone"), "TONE_ID_5")

        wake_setup_actions = [
            *self.setting_entry_actions(wake_entry_word, wake_entry_tone_id, wake_entry_tone_text),
            say_action(
                changed_wake,
                expect_tone_id=changed_wake_tone_id,
                expect_tone_text=changed_wake_tone_text,
                advisory_log_values=self.merge_advisory_logs(
                    self.wakeword_log_expectations(changed_wake),
                    self.config_saved_advisory_logs(),
                ),
                require_protocol_and_tone=False,
                auto_wake=False,
            ),
        ]

        return [
            case(
                self.next_case_id("POWER"),
                "POWER",
                "上下电后日志恢复",
                "控制串口可用",
                "执行一次上下电循环并检查日志恢复",
                "上下电后日志恢复且设备回到可交互状态",
                "",
                "",
                "",
                "P1",
                [manual_power_cycle_action()],
                self.requirement_ids_direct("DER-POWER-001"),
                "state_transition",
            ),
            case(
                self.next_case_id("POWER"),
                "POWER",
                "掉电后当前有效唤醒词符合配置",
                "设备支持硬断电且存在至少一个非默认候选唤醒词",
                f"先把唤醒词设置为【{changed_wake}】，再执行上下电，检查掉电后当前有效唤醒词是否符合配置",
                "掉电后当前应生效的唤醒词与掉电保持配置一致",
                "",
                "TONE_ID_0",
                self.tone_text("TONE_ID_0"),
                "P1",
                [
                    *wake_setup_actions,
                    manual_power_cycle_action(note="执行上下电循环，校验当前有效唤醒词"),
                    wake_action(changed_wake),
                ],
                self.requirement_ids_direct("CFG-012"),
                "state_transition",
            ),
            case(
                self.next_case_id("POWER"),
                "POWER",
                "掉电后其他候选唤醒词不误生效",
                "设备支持硬断电且存在至少两个候选唤醒词",
                f"先把唤醒词设置为【{changed_wake}】，再执行上下电，随后使用其他候选词【{other_wake}】尝试唤醒",
                "掉电后未被保留/未被设置的其他候选唤醒词不应误生效",
                "",
                "",
                "",
                "P1",
                [
                    *wake_setup_actions,
                    manual_power_cycle_action(note="执行上下电循环，校验其他候选唤醒词边界"),
                    no_wake_action(other_wake),
                ],
                self.requirement_ids_direct("CFG-012"),
                "negative",
            ),
            case(
                self.next_case_id("POWER"),
                "POWER",
                "掉电后音量保持规则正确",
                "设备处于默认音量且支持硬断电",
                f"先执行【{decrease_word}】把音量改成非默认，再执行上下电，最后继续执行【{decrease_word}】直到最小音量，检查剩余步数是否符合掉电配置",
                "掉电后音量保持或恢复默认的行为符合需求配置",
                "",
                lower_boundary_tone,
                self.tone_text(lower_boundary_tone),
                "P1",
                [
                    say_action(
                        decrease_word,
                        expect_tone_id=step_tone_id,
                        expect_tone_text=self.tone_text(step_tone_id),
                        advisory_log_values=self.config_saved_advisory_logs(),
                        observe_s=5.0,
                        require_protocol_and_tone=False,
                    ),
                    manual_power_cycle_action(note="执行上下电循环，校验音量掉电保持行为"),
                    {
                        "type": "volume_walk",
                        "command_word": decrease_word,
                        "expected_step_count": expected_down_steps,
                        "step_tone_id": step_tone_id,
                        "boundary_tone_id": lower_boundary_tone,
                        "extra_boundary_repeats": 1,
                    },
                ],
                self.requirement_ids_direct("CFG-013"),
                "state_transition",
            ),
        ]


    def build_factory_cases(self) -> list[dict[str, Any]]:
        alternate_wakes = [item for item in self.behavior_rules.get("wake_word_setting", {}).get("candidate_wake_words", []) if norm_text(item)]
        changed_wake = alternate_wakes[0] if alternate_wakes else "客厅窗帘"
        changed_wake_tone_id = self.setting_success_tone_id()
        changed_wake_tone_text = self.setting_success_tone_text()
        entry_word, entry_tone_id, entry_tone_text, entry_protocol = self.factory_entry_action()

        wake_entry = self.behavior_rules.get("wake_word_setting", {}).get("entry", {})
        wake_entry_word = norm_text(wake_entry.get("semantic"), "设置唤醒词")
        wake_entry_tone_id = norm_text(wake_entry.get("tone_id"), "TONE_ID_9")
        wake_entry_tone_text = norm_text(wake_entry.get("tts_text"), self.tone_text(wake_entry_tone_id))
        wake_setup_actions = [
            *self.setting_entry_actions(wake_entry_word, wake_entry_tone_id, wake_entry_tone_text),
            say_action(
                changed_wake,
                expect_asr=[changed_wake],
                advisory_tone_id=changed_wake_tone_id,
                advisory_tone_text=changed_wake_tone_text,
                advisory_log_values=self.merge_advisory_logs(
                    self.wakeword_log_expectations(changed_wake),
                    self.config_saved_advisory_logs(),
                ),
                require_protocol_and_tone=False,
                auto_wake=False,
            ),
        ]

        decrease_word = norm_text(self.behavior_rules.get("volume", {}).get("decrease_word"), "小声点")
        step_tone_id = norm_text(self.behavior_rules.get("volume", {}).get("step_tone_id"), "TONE_ID_1")
        default_down_steps = max(int(self.behavior_rules.get("volume", {}).get("default_level", 3) or 3) - 1, 0)
        lower_boundary_tone = norm_text(self.behavior_rules.get("volume", {}).get("lower_boundary_tone"), "TONE_ID_5")

        work_entry = self.behavior_rules.get("work_mode_setting", {}).get("entry", {})
        work_entry_word = norm_text(work_entry.get("semantic"), "设置工作模式")
        work_entry_tone_id = norm_text(work_entry.get("tone_id"), "TONE_ID_9")
        work_entry_tone_text = norm_text(work_entry.get("tts_text"), self.tone_text(work_entry_tone_id))
        tick_word = self.first_word(["滴答模式", "嘀嗒模式"], "滴答模式")
        tick_expect_asr = self.accepted_asr_variants(tick_word)
        tick_tone_id = self.row_tone_id(self.word_row(tick_word), "TONE_ID_15")
        tick_tone_text = self.row_tone_text(self.word_row(tick_word), tick_tone_id, "嘀（音效）")
        work_setup_actions = [
            *self.setting_entry_actions(work_entry_word, work_entry_tone_id, work_entry_tone_text),
            say_action(
                tick_word,
                expect_asr=tick_expect_asr,
                advisory_tone_id=tick_tone_id,
                advisory_tone_text=tick_tone_text,
                advisory_log_values=self.merge_advisory_logs(
                    self.work_mode_log_expectations(tick_word),
                    self.config_saved_advisory_logs(),
                ),
                require_protocol_and_tone=False,
                auto_wake=False,
            ),
        ]

        curtain_entry = self.behavior_rules.get("curtain_mode_setting", {}).get("entry", {})
        curtain_entry_word = norm_text(curtain_entry.get("semantic"), "设置窗帘模式")
        curtain_entry_tone_id = norm_text(curtain_entry.get("tone_id"), "TONE_ID_9")
        curtain_entry_tone_text = norm_text(curtain_entry.get("tts_text"), self.tone_text(curtain_entry_tone_id))
        cloth_mode_word = self.first_word(["布帘模式"], "布帘模式")
        cloth_mode_row = self.word_row(cloth_mode_word)
        cloth_mode_tone_id = self.setting_success_tone_id()
        cloth_mode_tone_text = self.row_tone_text(cloth_mode_row, cloth_mode_tone_id, self.setting_success_tone_text())
        sheer_mode_word = self.first_word(["纱帘模式"], "纱帘模式")
        sheer_mode_row = self.word_row(sheer_mode_word)
        sheer_mode_tone_id = self.setting_success_tone_id()
        sheer_mode_tone_text = self.row_tone_text(sheer_mode_row, sheer_mode_tone_id, self.setting_success_tone_text())
        curtain_setup_actions = [
            *self.setting_entry_actions(
                curtain_entry_word,
                curtain_entry_tone_id,
                curtain_entry_tone_text,
                expect_log_values={"curtainSettingEntry": ["into set curtain type"]},
            ),
            say_action(
                cloth_mode_word,
                expect_asr=self.accepted_asr_variants(cloth_mode_word),
                advisory_tone_id=cloth_mode_tone_id,
                advisory_tone_text=cloth_mode_tone_text,
                advisory_log_values=self.merge_advisory_logs(
                    self.curtain_mode_log_expectations(cloth_mode_word),
                    self.config_saved_advisory_logs(),
                ),
                require_protocol_and_tone=False,
                auto_wake=False,
            ),
        ]
        sheer_curtain_setup_actions = [
            *self.setting_entry_actions(
                curtain_entry_word,
                curtain_entry_tone_id,
                curtain_entry_tone_text,
                expect_log_values={"curtainSettingEntry": ["into set curtain type"]},
            ),
            say_action(
                sheer_mode_word,
                expect_asr=self.accepted_asr_variants(sheer_mode_word),
                advisory_tone_id=sheer_mode_tone_id,
                advisory_tone_text=sheer_mode_tone_text,
                advisory_log_values=self.merge_advisory_logs(
                    self.curtain_mode_log_expectations(sheer_mode_word),
                    self.config_saved_advisory_logs(),
                ),
                require_protocol_and_tone=False,
                auto_wake=False,
            ),
        ]

        default_open_word = self.first_word(CURTAIN_INTENT_GROUPS["default_open"], "打开窗帘")
        default_open_row = self.word_row(default_open_word)
        default_open_tone_id = self.row_tone_id(default_open_row, "TONE_ID_1")
        default_open_tone_text = self.row_tone_text(default_open_row, default_open_tone_id, "好的")
        default_open_protocol = norm_hex(default_open_row.get("send_protocol"))
        cloth_open_word = self.first_word(CURTAIN_INTENT_GROUPS["cloth_open"], "打开布帘")
        sheer_open_word = self.first_word(CURTAIN_INTENT_GROUPS["sheer_open"], "打开纱帘")

        factory_actions = [
            say_action(
                entry_word,
                expect_tone_id=entry_tone_id,
                expect_tone_text=entry_tone_text,
                expect_send_protocol=entry_protocol,
                advisory_log_values=self.config_saved_advisory_logs(),
                require_protocol_and_tone=False,
            ),
            wait_action(2.0),
        ]

        return [
            case(
                self.next_case_id("FACTORY"),
                "FACTORY",
                "恢复出厂入口可达",
                "设备处于默认待命态",
                f"唤醒后说【{entry_word}】",
                "设备进入恢复出厂流程或给出确认播报",
                entry_protocol,
                entry_tone_id,
                entry_tone_text,
                "P1",
                [
                    say_action(
                        entry_word,
                        expect_tone_id=entry_tone_id,
                        expect_tone_text=entry_tone_text,
                        expect_send_protocol=entry_protocol,
                        advisory_log_values=self.config_saved_advisory_logs(),
                        require_protocol_and_tone=False,
                    )
                ],
                self.requirement_ids_direct("DER-FACT-001"),
                "positive",
            ),
            case(
                self.next_case_id("FACTORY"),
                "FACTORY",
                "恢复出厂后默认唤醒词恢复",
                "设备支持至少一个非默认候选唤醒词",
                f"先把唤醒词设置为【{changed_wake}】，再执行恢复出厂，然后验证默认唤醒词【{self.wake_word}】可用",
                "恢复出厂后默认唤醒词恢复可用",
                "",
                "TONE_ID_0",
                self.tone_text("TONE_ID_0"),
                "P1",
                [
                    *wake_setup_actions,
                    wake_action(changed_wake),
                    say_action(
                        entry_word,
                        expect_tone_id=entry_tone_id,
                        expect_tone_text=entry_tone_text,
                        expect_send_protocol=entry_protocol,
                        advisory_log_values=self.config_saved_advisory_logs(),
                        require_protocol_and_tone=False,
                        auto_wake=False,
                    ),
                    wait_action(2.0),
                    wake_action(self.wake_word),
                ],
                self.requirement_ids_direct("FACT-001"),
                "state_transition",
            ),
            case(
                self.next_case_id("FACTORY"),
                "FACTORY",
                "恢复出厂后自定义唤醒词失效",
                "设备支持至少一个非默认候选唤醒词",
                f"先把唤醒词设置为【{changed_wake}】，再执行恢复出厂，然后使用原自定义唤醒词【{changed_wake}】尝试唤醒",
                "恢复出厂后原自定义唤醒词失效",
                "",
                "",
                "",
                "P1",
                [
                    *wake_setup_actions,
                    wake_action(changed_wake),
                    say_action(
                        entry_word,
                        expect_tone_id=entry_tone_id,
                        expect_tone_text=entry_tone_text,
                        expect_send_protocol=entry_protocol,
                        advisory_log_values=self.config_saved_advisory_logs(),
                        require_protocol_and_tone=False,
                        auto_wake=False,
                    ),
                    wait_action(2.0),
                    no_wake_action(changed_wake),
                ],
                self.requirement_ids_direct("FACT-001"),
                "negative",
            ),
            case(
                self.next_case_id("FACTORY"),
                "FACTORY",
                "恢复出厂后默认音量恢复",
                "设备当前音量已被改为非默认值",
                f"先执行【{decrease_word}】把音量改成非默认，再执行恢复出厂，随后继续执行【{decrease_word}】直到最小音量，检查步数是否回到默认档位",
                "恢复出厂后音量恢复默认档位",
                "",
                lower_boundary_tone,
                self.tone_text(lower_boundary_tone),
                "P1",
                [
                    say_action(decrease_word, expect_tone_id=step_tone_id, expect_tone_text=self.tone_text(step_tone_id), require_protocol_and_tone=False),
                    *factory_actions,
                    {
                        "type": "volume_walk",
                        "command_word": decrease_word,
                        "expected_step_count": default_down_steps,
                        "step_tone_id": step_tone_id,
                        "boundary_tone_id": lower_boundary_tone,
                        "extra_boundary_repeats": 1,
                    },
                ],
                self.requirement_ids_direct("FACT-002"),
                "state_transition",
            ),
            case(
                self.next_case_id("FACTORY"),
                "FACTORY",
                "恢复出厂后默认工作模式恢复",
                "设备工作模式已被切到非默认模式",
                f"先切到【{tick_word}】，再执行恢复出厂，然后执行普通窗帘控制命令【{default_open_word}】检查是否恢复语音模式播报",
                "恢复出厂后普通控制命令恢复默认语音模式播报",
                default_open_protocol,
                default_open_tone_id,
                default_open_tone_text,
                "P1",
                [
                    *work_setup_actions,
                    *factory_actions,
                    say_action(default_open_word, expect_tone_id=default_open_tone_id, expect_tone_text=default_open_tone_text, expect_send_protocol=default_open_protocol, require_protocol_and_tone=True),
                ],
                self.requirement_ids_direct("FACT-003"),
                "state_transition",
            ),
            case(
                self.next_case_id("FACTORY"),
                "FACTORY",
                "恢复出厂后布帘模式专有词失效",
                "设备窗帘模式已被切到非默认模式",
                f"先切到【{cloth_mode_word}】，再执行恢复出厂，然后验证布帘控制词【{cloth_open_word}】不再响应",
                "恢复出厂后默认窗帘模式恢复，布帘模式专有词失效",
                "",
                "",
                "",
                "P1",
                [
                    *curtain_setup_actions,
                    *factory_actions,
                    say_action(cloth_open_word, expect_no_response=True, observe_s=2.0),
                ],
                self.requirement_ids_direct("FACT-004"),
                "state_transition",
            ),
            case(
                self.next_case_id("FACTORY"),
                "FACTORY",
                "恢复出厂后纱帘模式专有词失效",
                "设备窗帘模式已被切到非默认模式",
                f"先切到【{sheer_mode_word}】，再执行恢复出厂，然后验证纱帘控制词【{sheer_open_word}】不再响应",
                "恢复出厂后纱帘模式对应的专有控制词失效",
                "",
                "",
                "",
                "P1",
                [
                    *sheer_curtain_setup_actions,
                    *factory_actions,
                    say_action(sheer_open_word, expect_no_response=True, observe_s=2.0),
                ],
                self.requirement_ids_direct("FACT-004"),
                "negative",
            ),
        ]

    def build_wakeword_cases(self) -> list[dict[str, Any]]:
        rules = self.behavior_rules.get("wake_word_setting", {})
        entry = rules.get("entry", {})
        candidates = [item for item in rules.get("candidate_wake_words", []) if norm_text(item)]
        timeout_s = float(self.behavior_rules.get("base", {}).get("wake_timeout_s", 15))
        entry_word = norm_text(entry.get("semantic"), "设置唤醒词")
        entry_tone_id = norm_text(entry.get("tone_id"), "TONE_ID_9")
        entry_tone_text = norm_text(entry.get("tts_text"), self.tone_text(entry_tone_id))
        invalid_word = self.recognized_invalid_word("打开窗帘")
        setting_timeout_tone_id = "TONE_ID_12" if self.tone_text("TONE_ID_12") else "TONE_ID_2"
        probe_candidate = candidates[0] if candidates else ""
        default_open_word = self.first_word(CURTAIN_INTENT_GROUPS["default_open"], "打开窗帘")
        default_open_row = self.word_row(default_open_word)
        default_open_tone_id = self.row_tone_id(default_open_row, "TONE_ID_1")
        default_open_tone_text = self.row_tone_text(default_open_row, default_open_tone_id, "好的")
        default_open_protocol = self.row_protocol(default_open_row)

        def wake_setting_entry_actions() -> list[dict[str, Any]]:
            return [
                with_label(wake_action(self.wake_word), f"播放默认唤醒词【{self.wake_word}】"),
                with_label(
                    say_action(
                        entry_word,
                        expect_asr=self.accepted_asr_variants(entry_word),
                        advisory_tone_id=entry_tone_id,
                        advisory_tone_text=entry_tone_text,
                        require_protocol_and_tone=False,
                        auto_wake=False,
                    ),
                    f"进入唤醒词设置【{entry_word}】",
                ),
            ]

        def candidate_setting_action(candidate: str) -> dict[str, Any]:
            candidate_row = self.word_row(candidate)
            candidate_tone_id = self.setting_success_tone_id()
            candidate_tone_text = self.row_tone_text(candidate_row, candidate_tone_id, self.setting_success_tone_text())
            return with_label(
                say_action(
                    candidate,
                    expect_asr=self.accepted_asr_variants(candidate),
                    advisory_tone_id=candidate_tone_id,
                    advisory_tone_text=candidate_tone_text,
                    advisory_log_values=self.merge_advisory_logs(
                        self.wakeword_log_expectations(candidate),
                        self.config_saved_advisory_logs(),
                    ),
                    require_protocol_and_tone=False,
                    auto_wake=False,
                ),
                f"设置当前唤醒词为【{candidate}】",
            )

        def candidate_setup_actions(candidate: str) -> list[dict[str, Any]]:
            return self.tagged_actions(
                [
                    *wake_setting_entry_actions(),
                    candidate_setting_action(candidate),
                    with_label(wait_action(2.0), f"等待【{candidate}】配置稳定"),
                ],
                setup_action=True,
            )

        cases = [
            case(
                self.next_case_id("WAKEWORD"),
                "WAKEWORD",
                "默认唤醒词可用",
                "设备处于默认待命态",
                f"用默认唤醒词【{self.wake_word}】唤醒设备",
                "默认唤醒词可唤醒设备",
                "",
                "TONE_ID_0",
                self.tone_text("TONE_ID_0"),
                "P1",
                [with_label(wake_action(self.wake_word), f"验证默认唤醒词【{self.wake_word}】可唤醒设备")],
                self.requirement_ids_direct("WAKE-001"),
                "positive",
            )
        ]
        if candidates:
            for candidate in candidates:
                cases.append(
                    case(
                        self.next_case_id("WAKEWORD"),
                        "WAKEWORD",
                        f"候选唤醒词【{candidate}】在设置外不应生效",
                        "设备处于默认待命态",
                        f"不进入设置模式，直接说候选唤醒词【{candidate}】",
                        "未进入设置模式时，候选唤醒词不应直接唤醒设备",
                        "",
                        "",
                        "",
                        "P1",
                        [
                            with_label(
                                no_wake_action(candidate, repeats=2),
                                f"验证候选唤醒词【{candidate}】在设置外不应唤醒设备",
                            )
                        ],
                        self.requirement_ids_direct("WAKE-002"),
                        "negative",
                    )
                )
            cases.append(
                case(
                    self.next_case_id("WAKEWORD"),
                    "WAKEWORD",
                    "设置唤醒词入口后普通控制词不影响后续设置",
                    "设备处于默认待命态",
                    f"进入【{entry_word}】后先执行普通控制词【{invalid_word}】，再继续设置唤醒词",
                    "设置唤醒词窗口内执行普通控制词后，后续设置流程仍可继续完成",
                    "",
                    default_open_tone_id,
                    default_open_tone_text,
                    "P1",
                    [
                        *wake_setting_entry_actions(),
                        with_label(
                            say_action(
                                invalid_word,
                                expect_asr=self.accepted_asr_variants(invalid_word),
                                expect_tone_id=default_open_tone_id,
                                expect_tone_text=default_open_tone_text,
                                expect_send_protocol=default_open_protocol,
                                auto_wake=False,
                                require_protocol_and_tone=False,
                            ),
                            f"在设置窗口内执行普通控制词【{invalid_word}】",
                        ),
                        candidate_setting_action(probe_candidate),
                        with_label(wake_action(probe_candidate), f"验证普通控制词后【{probe_candidate}】仍可设置并立即生效"),
                    ],
                    self.requirement_ids_direct("WAKE-004"),
                    "state_transition",
                )
            )
            for candidate in candidates:
                cases.append(
                    case(
                        self.next_case_id("WAKEWORD"),
                        "WAKEWORD",
                        f"设置为【{candidate}】后新唤醒词立即生效",
                        "设备处于默认待命态",
                        f"进入【{entry_word}】并设置唤醒词为【{candidate}】，随后立即使用该词唤醒设备",
                        "设置成功后，新唤醒词可立即生效",
                        "",
                        "TONE_ID_0",
                        self.tone_text("TONE_ID_0"),
                        "P1",
                        [
                            *wake_setting_entry_actions(),
                            candidate_setting_action(candidate),
                            with_label(wake_action(candidate), f"验证新唤醒词【{candidate}】立即生效"),
                        ],
                        self.requirement_ids_direct("WAKE-003"),
                        "state_transition",
                    )
                )
            cases.append(
                case(
                    self.next_case_id("WAKEWORD"),
                    "WAKEWORD",
                    "设置唤醒词模式超时值符合配置",
                    "设备处于默认待命态",
                    f"先唤醒设备，进入【{entry_word}】后不继续说候选唤醒词，记录超时事件是否在 {int(timeout_s)}s 左右出现",
                    "设置唤醒词模式的超时值符合基础配置",
                    "",
                    setting_timeout_tone_id,
                    self.tone_text(setting_timeout_tone_id),
                    "P2",
                    [
                        *wake_setting_entry_actions(),
                        with_label(
                            wait_action(timeout_s, expect_tone_id=setting_timeout_tone_id, expect_markers=["TIME_OUT"], measure_timeout=True),
                            f"等待 {int(timeout_s)}s 验证设置窗口超时退出",
                        ),
                    ],
                    self.requirement_ids_direct("CFG-001"),
                    "boundary",
                )
            )
            timeout_probe_word = candidates[0]
            cases.append(
                case(
                    self.next_case_id("WAKEWORD"),
                    "WAKEWORD",
                    "设置唤醒词模式超时后退出设置态",
                    "设备处于默认待命态",
                    f"进入【{entry_word}】后等待实际超时退出，再直接说候选唤醒词【{timeout_probe_word}】",
                    "设置唤醒词模式超时后自动退出；超时后候选唤醒词不会再被当作设置项处理",
                    "",
                    "",
                    "",
                    "P2",
                    [
                        *wake_setting_entry_actions(),
                        with_label(
                            self.wait_for_timeout_exit_action(
                                timeout_s,
                                expect_tone_id=setting_timeout_tone_id,
                                label="等待唤醒词设置实际超时退出",
                            ),
                            "等待唤醒词设置实际超时退出",
                        ),
                        with_label(
                            no_wake_action(timeout_probe_word),
                            f"验证超时后【{timeout_probe_word}】不应再触发设置或唤醒",
                        ),
                    ],
                    self.requirement_ids_direct("WAKE-005"),
                    "state_transition",
                )
            )
        return cases


    def build_workmode_cases(self) -> list[dict[str, Any]]:
        rules = self.behavior_rules.get("work_mode_setting", {})
        entry = rules.get("entry", {})
        candidates = rules.get("candidates", [])
        timeout_s = float(self.behavior_rules.get("base", {}).get("wake_timeout_s", 15))
        entry_word = norm_text(entry.get("semantic"), "设置工作模式")
        entry_tone_id = norm_text(entry.get("tone_id"), "TONE_ID_9")
        entry_tone_text = norm_text(entry.get("tts_text"), self.tone_text(entry_tone_id))
        voice_word = next((norm_text(item.get("semantic")) for item in candidates if norm_text(item.get("semantic")) == "语音模式"), "语音模式")
        tick_word = next((norm_text(item.get("semantic")) for item in candidates if norm_text(item.get("semantic")) in {"滴答模式", "嘀嗒模式"}), "滴答模式")
        voice_expect_asr = self.accepted_asr_variants(voice_word)
        tick_expect_asr = self.accepted_asr_variants(tick_word)
        voice_tone_id = self.setting_success_tone_id()
        voice_tone_text = self.setting_success_tone_text()
        tick_row = self.word_row(tick_word)
        tick_tone_id = self.row_tone_id(tick_row, "TONE_ID_15")
        tick_tone_text = self.row_tone_text(tick_row, tick_tone_id, "嘀（音效）")
        invalid_word = self.recognized_invalid_word("打开窗帘")
        default_open_word = self.first_word(CURTAIN_INTENT_GROUPS["default_open"], "打开窗帘")
        default_open_row = self.word_row(default_open_word)
        default_open_tone_id = self.row_tone_id(default_open_row, "TONE_ID_1")
        default_open_tone_text = self.row_tone_text(default_open_row, default_open_tone_id, "好的")
        default_open_protocol = norm_hex(default_open_row.get("send_protocol"))

        def entry_actions() -> list[dict[str, Any]]:
            return self.setting_entry_actions(entry_word, entry_tone_id, entry_tone_text)

        def select_mode_action(mode_word: str) -> dict[str, Any]:
            if norm_text(mode_word) == norm_text(tick_word):
                advisory_tone_id = tick_tone_id
                advisory_tone_text = tick_tone_text
                expect_asr = tick_expect_asr
            else:
                advisory_tone_id = voice_tone_id
                advisory_tone_text = voice_tone_text
                expect_asr = voice_expect_asr
            return say_action(
                mode_word,
                expect_asr=expect_asr,
                advisory_tone_id=advisory_tone_id,
                advisory_tone_text=advisory_tone_text,
                advisory_log_values=self.merge_advisory_logs(
                    self.work_mode_log_expectations(mode_word),
                    self.config_saved_advisory_logs(),
                ),
                require_protocol_and_tone=False,
                auto_wake=False,
            )

        return [
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                "工作模式设置入口可达",
                "设备处于默认待命态",
                f"唤醒后说【{entry_word}】进入工作模式设置",
                "设备进入工作模式设置入口并播报设置提示音",
                "",
                entry_tone_id,
                entry_tone_text,
                "P1",
                [*entry_actions()],
                self.requirement_ids_direct("DER-WORK-001"),
                "positive",
            ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                "默认工作模式为语音模式",
                "设备处于默认待命态",
                f"唤醒后执行普通控制词【{default_open_word}】",
                "默认工作模式下，普通控制词使用语音播报反馈",
                default_open_protocol,
                default_open_tone_id,
            default_open_tone_text,
            "P1",
            [
                *self.restore_default_actions(note="用例前恢复默认状态"),
                say_action(default_open_word, expect_tone_id=default_open_tone_id, expect_tone_text=default_open_tone_text, expect_send_protocol=default_open_protocol, require_protocol_and_tone=False),
            ],
            self.requirement_ids_direct("WORK-001"),
            "positive",
        ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                f"{voice_word}在设置外不响应",
                "设备处于默认待命态",
                f"不进入设置模式，直接说【{voice_word}】",
                "未进入设置工作模式时，语音模式词不应响应",
                "",
                "",
                "",
                "P1",
                [say_action(voice_word, expect_no_response=True, observe_s=2.0, forbid_log_values=self.forbid_work_mode_change_logs())],
                self.requirement_ids_direct("WORK-002"),
                "negative",
            ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                f"{tick_word}在设置外不响应",
                "设备处于默认待命态",
                f"不进入设置模式，直接说【{tick_word}】",
                "未进入设置工作模式时，滴答模式词不应响应",
                "",
                "",
                "",
                "P1",
                [say_action(tick_word, expect_no_response=True, observe_s=2.0, forbid_log_values=self.forbid_work_mode_change_logs())],
                self.requirement_ids_direct("WORK-002"),
                "negative",
            ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                "设置工作模式入口后普通控制词不影响后续设置",
                "设备处于默认待命态",
                f"进入【{entry_word}】后先说普通控制词【{invalid_word}】，再继续设置为【{tick_word}】",
                "设置窗口内普通控制词可响应，且后续仍可继续完成工作模式设置",
                "",
                tick_tone_id,
                tick_tone_text,
                "P1",
                [
                    *entry_actions(),
                    say_action(
                        invalid_word,
                        expect_tone_id=default_open_tone_id,
                        expect_tone_text=default_open_tone_text,
                        expect_send_protocol=default_open_protocol,
                        auto_wake=False,
                        require_protocol_and_tone=False,
                    ),
                    select_mode_action(tick_word),
                ],
                self.requirement_ids_direct("WORK-004"),
                "state_transition",
            ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                f"设置为{tick_word}后立即生效",
                "设备处于默认待命态",
                f"进入【{entry_word}】后切换到【{tick_word}】，随后执行普通控制词【{default_open_word}】",
                "切换到滴答模式后，普通控制词立即使用滴答反馈",
                default_open_protocol,
                tick_tone_id,
                tick_tone_text,
                "P1",
                [
                    *entry_actions(),
                    select_mode_action(tick_word),
                    say_action(default_open_word, expect_tone_id=tick_tone_id, expect_tone_text=tick_tone_text, expect_send_protocol=default_open_protocol, require_protocol_and_tone=True),
                ],
                self.requirement_ids_direct("WORK-003"),
                "state_transition",
            ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                f"设置为{voice_word}后立即生效",
                "设备当前已切到滴答模式",
                f"先切到【{tick_word}】，再重新进入【{entry_word}】切回【{voice_word}】，随后执行普通控制词【{default_open_word}】",
                "切换回语音模式后，普通控制词立即恢复语音播报",
                default_open_protocol,
                default_open_tone_id,
                default_open_tone_text,
                "P1",
                [
                    *entry_actions(),
                    select_mode_action(tick_word),
                    *entry_actions(),
                    select_mode_action(voice_word),
                    say_action(default_open_word, expect_tone_id=default_open_tone_id, expect_tone_text=default_open_tone_text, expect_send_protocol=default_open_protocol, require_protocol_and_tone=True),
                ],
                self.requirement_ids_direct("WORK-003"),
                "state_transition",
            ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                "工作模式设置超时值符合配置",
                "设备处于默认待命态",
                f"进入工作模式设置后不继续选择模式，记录超时事件是否在 {int(timeout_s)}s 左右出现",
                "工作模式设置超时值符合基础配置",
                "",
                "TONE_ID_2",
                self.tone_text("TONE_ID_2"),
                "P1",
                [
                    *entry_actions(),
                    wait_action(timeout_s, expect_tone_id="TONE_ID_2", expect_markers=["TIME_OUT"], measure_timeout=True),
                ],
                self.requirement_ids_direct("CFG-001"),
                "boundary",
            ),
            case(
                self.next_case_id("WORKMODE"),
                "WORKMODE",
                "工作模式设置超时后退出设置态",
                "设备处于默认待命态",
                f"进入工作模式设置后等待实际超时退出，再直接说【{tick_word}】",
                "设置工作模式超时后自动退出，超时后模式词不再被当作设置项处理",
                "",
                "",
                "",
                "P1",
                [
                    *entry_actions(),
                    self.wait_for_timeout_exit_action(timeout_s, label="等待工作模式设置实际超时退出"),
                    say_action(tick_word, expect_no_response=True, observe_s=2.0),
                ],
                self.requirement_ids_direct("WORK-005"),
                "state_transition",
            ),
        ]

    def build_curtainmode_cases(self) -> list[dict[str, Any]]:
        rules = self.behavior_rules.get("curtain_mode_setting", {})
        entry = rules.get("entry", {})
        timeout_s = float(self.behavior_rules.get("base", {}).get("wake_timeout_s", 15))
        entry_word = norm_text(entry.get("semantic"), "设置窗帘模式")
        entry_tone_id = norm_text(entry.get("tone_id"), "TONE_ID_9")
        entry_tone_text = norm_text(entry.get("tts_text"), self.tone_text(entry_tone_id))
        invalid_word = self.recognized_invalid_word("打开窗帘")

        default_open_word = self.first_word(CURTAIN_INTENT_GROUPS["default_open"], "打开窗帘")
        default_open_row = self.word_row(default_open_word)
        default_open_tone_id = self.row_tone_id(default_open_row, "TONE_ID_1")
        default_open_tone_text = self.row_tone_text(default_open_row, default_open_tone_id, "好的")
        default_open_protocol = norm_hex(default_open_row.get("send_protocol"))

        sheer_mode_word = self.first_word(["纱帘模式"], "纱帘模式")
        window_mode_word = self.first_word(["窗纱模式"], "窗纱模式")
        cloth_mode_word = self.first_word(["布帘模式"], "布帘模式")

        cloth_open_word = self.first_word(CURTAIN_INTENT_GROUPS["cloth_open"], "打开布帘")
        cloth_open_row = self.word_row(cloth_open_word)
        cloth_open_tone_id = self.row_tone_id(cloth_open_row, "TONE_ID_1")
        cloth_open_tone_text = self.row_tone_text(cloth_open_row, cloth_open_tone_id, "好的")
        cloth_open_protocol = norm_hex(cloth_open_row.get("send_protocol"))

        sheer_open_word = self.first_word(CURTAIN_INTENT_GROUPS["sheer_open"], "打开纱帘")
        sheer_open_row = self.word_row(sheer_open_word)
        sheer_open_tone_id = self.row_tone_id(sheer_open_row, "TONE_ID_1")
        sheer_open_tone_text = self.row_tone_text(sheer_open_row, sheer_open_tone_id, "好的")
        sheer_open_protocol = norm_hex(sheer_open_row.get("send_protocol"))

        window_open_word = self.first_word(CURTAIN_INTENT_GROUPS["window_sheer_open"], "打开窗纱")
        window_open_row = self.word_row(window_open_word)
        window_open_tone_id = self.row_tone_id(window_open_row, "TONE_ID_1")
        window_open_tone_text = self.row_tone_text(window_open_row, window_open_tone_id, "好的")
        window_open_protocol = norm_hex(window_open_row.get("send_protocol"))

        def entry_actions() -> list[dict[str, Any]]:
            return self.setting_entry_actions(
                entry_word,
                entry_tone_id,
                entry_tone_text,
                expect_log_values={"curtainSettingEntry": ["into set curtain type"]},
            )

        def select_mode_action(mode_word: str) -> dict[str, Any]:
            mode_row = self.word_row(mode_word)
            success_tone_id = self.setting_success_tone_id()
            return say_action(
                mode_word,
                expect_asr=self.accepted_asr_variants(mode_word),
                advisory_tone_id=success_tone_id,
                advisory_tone_text=self.row_tone_text(mode_row, success_tone_id, self.setting_success_tone_text()),
                advisory_log_values=self.merge_advisory_logs(
                    self.curtain_mode_log_expectations(mode_word),
                    self.config_saved_advisory_logs(),
                ),
                require_protocol_and_tone=False,
                auto_wake=False,
            )

        success_tone_id = self.setting_success_tone_id()
        success_tone_text = self.setting_success_tone_text()
        return [
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                "窗帘模式设置入口可达",
                "设备处于默认待命态",
                f"唤醒后说【{entry_word}】进入窗帘模式设置",
                "设备进入窗帘模式设置入口并播报设置提示音",
                "",
                entry_tone_id,
                entry_tone_text,
                "P1",
                [*entry_actions()],
                self.requirement_ids_direct("DER-CURTAINMODE-001"),
                "positive",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                "默认窗帘模式为窗帘模式",
                "设备处于默认待命态",
                f"唤醒后执行默认窗帘控制词【{default_open_word}】",
                "默认窗帘模式下，默认窗帘控制词可正常响应",
                default_open_protocol,
                default_open_tone_id,
            default_open_tone_text,
            "P1",
            [
                *self.restore_default_actions(note="用例前恢复默认状态"),
                say_action(default_open_word, expect_tone_id=default_open_tone_id, expect_tone_text=default_open_tone_text, expect_send_protocol=default_open_protocol, require_protocol_and_tone=False),
            ],
            self.requirement_ids_direct("CURTAIN-001"),
            "positive",
        ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{sheer_mode_word}在设置外不响应",
                "设备处于默认待命态",
                f"不进入设置模式，直接说【{sheer_mode_word}】",
                "未进入设置窗帘模式时，纱帘模式词不应响应",
                "",
                "",
                "",
                "P1",
                [say_action(sheer_mode_word, expect_no_response=True, observe_s=2.0, forbid_log_values=self.forbid_curtain_mode_change_logs())],
                self.requirement_ids_direct("CURTAIN-003"),
                "negative",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{window_mode_word}在设置外不响应",
                "设备处于默认待命态",
                f"不进入设置模式，直接说【{window_mode_word}】",
                "未进入设置窗帘模式时，窗纱模式词不应响应",
                "",
                "",
                "",
                "P1",
                [say_action(window_mode_word, expect_no_response=True, observe_s=2.0, forbid_log_values=self.forbid_curtain_mode_change_logs())],
                self.requirement_ids_direct("CURTAIN-003"),
                "negative",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{cloth_mode_word}在设置外不响应",
                "设备处于默认待命态",
                f"不进入设置模式，直接说【{cloth_mode_word}】",
                "未进入设置窗帘模式时，布帘模式词不应响应",
                "",
                "",
                "",
                "P1",
                [say_action(cloth_mode_word, expect_no_response=True, observe_s=2.0, forbid_log_values=self.forbid_curtain_mode_change_logs())],
                self.requirement_ids_direct("CURTAIN-003"),
                "negative",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                "设置窗帘模式入口后普通控制词不影响后续设置",
                "设备处于默认待命态",
                f"进入【{entry_word}】后先说普通控制词【{invalid_word}】，再继续设置为【{cloth_mode_word}】",
                "设置窗口内普通控制词可响应，且后续仍可继续完成窗帘模式设置",
                "",
                success_tone_id,
                success_tone_text,
                "P1",
                [
                    *entry_actions(),
                    say_action(
                        invalid_word,
                        expect_tone_id=default_open_tone_id,
                        expect_tone_text=default_open_tone_text,
                        expect_send_protocol=default_open_protocol,
                        auto_wake=False,
                        require_protocol_and_tone=False,
                    ),
                    select_mode_action(cloth_mode_word),
                ],
                self.requirement_ids_direct("CURTAIN-005"),
                "state_transition",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"设置为{sheer_mode_word}后立即生效",
                "设备处于默认待命态",
                f"进入【{entry_word}】后切换到【{sheer_mode_word}】",
                "选择纱帘模式后立即完成模式切换",
                "",
                success_tone_id,
                success_tone_text,
                "P1",
                [*entry_actions(), select_mode_action(sheer_mode_word)],
                self.requirement_ids_direct("CURTAIN-004"),
                "state_transition",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{sheer_mode_word}下纱帘控制词可用",
                "设备处于默认待命态",
                f"先切到【{sheer_mode_word}】，再执行纱帘控制词【{sheer_open_word}】",
                "纱帘模式下纱帘控制词可正常响应",
                sheer_open_protocol,
                sheer_open_tone_id,
                sheer_open_tone_text,
                "P1",
                [*entry_actions(), select_mode_action(sheer_mode_word), say_action(sheer_open_word, expect_tone_id=sheer_open_tone_id, expect_tone_text=sheer_open_tone_text, expect_send_protocol=sheer_open_protocol, require_protocol_and_tone=True)],
                self.requirement_ids_direct("CURTAIN-002"),
                "state_transition",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{sheer_mode_word}下布帘控制词不应响应",
                "设备处于默认待命态",
                f"先切到【{sheer_mode_word}】，再执行布帘控制词【{cloth_open_word}】",
                "纱帘模式下布帘控制词不应响应",
                "",
                "",
                "",
                "P1",
                [*entry_actions(), select_mode_action(sheer_mode_word), say_action(cloth_open_word, expect_no_response=True, observe_s=2.0)],
                self.requirement_ids_direct("CURTAIN-002"),
                "negative",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"设置为{window_mode_word}后立即生效",
                "设备处于默认待命态",
                f"进入【{entry_word}】后切换到【{window_mode_word}】",
                "选择窗纱模式后立即完成模式切换",
                "",
                success_tone_id,
                success_tone_text,
                "P1",
                [*entry_actions(), select_mode_action(window_mode_word)],
                self.requirement_ids_direct("CURTAIN-004"),
                "state_transition",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{window_mode_word}下窗纱控制词可用",
                "设备处于默认待命态",
                f"先切到【{window_mode_word}】，再执行窗纱控制词【{window_open_word}】",
                "窗纱模式下窗纱控制词可正常响应",
                window_open_protocol,
                window_open_tone_id,
                window_open_tone_text,
                "P1",
                [*entry_actions(), select_mode_action(window_mode_word), say_action(window_open_word, expect_tone_id=window_open_tone_id, expect_tone_text=window_open_tone_text, expect_send_protocol=window_open_protocol, require_protocol_and_tone=True)],
                self.requirement_ids_direct("CURTAIN-002"),
                "state_transition",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{window_mode_word}下布帘控制词不应响应",
                "设备处于默认待命态",
                f"先切到【{window_mode_word}】，再执行布帘控制词【{cloth_open_word}】",
                "窗纱模式下布帘控制词不应响应",
                "",
                "",
                "",
                "P1",
                [*entry_actions(), select_mode_action(window_mode_word), say_action(cloth_open_word, expect_no_response=True, observe_s=2.0)],
                self.requirement_ids_direct("CURTAIN-002"),
                "negative",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"设置为{cloth_mode_word}后立即生效",
                "设备处于默认待命态",
                f"进入【{entry_word}】后切换到【{cloth_mode_word}】",
                "选择布帘模式后立即完成模式切换",
                "",
                success_tone_id,
                success_tone_text,
                "P1",
                [*entry_actions(), select_mode_action(cloth_mode_word)],
                self.requirement_ids_direct("CURTAIN-004"),
                "state_transition",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{cloth_mode_word}下布帘控制词可用",
                "设备处于默认待命态",
                f"先切到【{cloth_mode_word}】，再执行布帘控制词【{cloth_open_word}】",
                "布帘模式下布帘控制词可正常响应",
                cloth_open_protocol,
                cloth_open_tone_id,
                cloth_open_tone_text,
                "P1",
                [*entry_actions(), select_mode_action(cloth_mode_word), say_action(cloth_open_word, expect_tone_id=cloth_open_tone_id, expect_tone_text=cloth_open_tone_text, expect_send_protocol=cloth_open_protocol, require_protocol_and_tone=True)],
                self.requirement_ids_direct("CURTAIN-002"),
                "state_transition",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                f"{cloth_mode_word}下纱帘控制词不应响应",
                "设备处于默认待命态",
                f"先切到【{cloth_mode_word}】，再执行纱帘控制词【{sheer_open_word}】",
                "布帘模式下纱帘/窗纱控制词不应响应",
                "",
                "",
                "",
                "P1",
                [*entry_actions(), select_mode_action(cloth_mode_word), say_action(sheer_open_word, expect_no_response=True, observe_s=2.0)],
                self.requirement_ids_direct("CURTAIN-002"),
                "negative",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                "窗帘模式设置超时值符合配置",
                "设备处于默认待命态",
                f"进入窗帘模式设置后不继续选择目标模式，记录超时事件是否在 {int(timeout_s)}s 左右出现",
                "窗帘模式设置超时值符合基础配置",
                "",
                "TONE_ID_2",
                self.tone_text("TONE_ID_2"),
                "P1",
                [*entry_actions(), wait_action(timeout_s, expect_tone_id="TONE_ID_2", expect_markers=["TIME_OUT"], measure_timeout=True)],
                self.requirement_ids_direct("CFG-001"),
                "boundary",
            ),
            case(
                self.next_case_id("CURTAINMODE"),
                "CURTAINMODE",
                "窗帘模式设置超时后退出设置态",
                "设备处于默认待命态",
                f"进入窗帘模式设置后等待实际超时退出，再直接说【{cloth_mode_word}】",
                "设置窗帘模式超时后自动退出，超时后模式词不再被当作设置项处理",
                "",
                "",
                "",
                "P1",
                [*entry_actions(), self.wait_for_timeout_exit_action(timeout_s, label="等待窗帘模式设置实际超时退出"), say_action(cloth_mode_word, expect_no_response=True, observe_s=2.0)],
                self.requirement_ids_direct("CURTAIN-006"),
                "state_transition",
            ),
        ]

    def build_curtain_cases(self) -> list[dict[str, Any]]:
        curtain_entry = self.behavior_rules.get("curtain_mode_setting", {}).get("entry", {})
        curtain_entry_word = norm_text(curtain_entry.get("semantic"), "设置窗帘模式")
        curtain_entry_tone_id = norm_text(curtain_entry.get("tone_id"), "TONE_ID_9")
        curtain_entry_tone_text = norm_text(curtain_entry.get("tts_text"), self.tone_text(curtain_entry_tone_id))

        default_open_words = self.curtain_words("default_open", CURTAIN_INTENT_GROUPS["default_open"])
        default_close_words = self.curtain_words("default_close", CURTAIN_INTENT_GROUPS["default_close"])
        default_stop_words = self.curtain_words("default_stop", CURTAIN_INTENT_GROUPS["default_stop"])
        cloth_open_words = self.curtain_words("cloth_open", CURTAIN_INTENT_GROUPS["cloth_open"])
        cloth_close_words = self.curtain_words("cloth_close", CURTAIN_INTENT_GROUPS["cloth_close"])
        cloth_stop_words = self.curtain_words("cloth_stop", CURTAIN_INTENT_GROUPS["cloth_stop"])
        sheer_open_words = self.curtain_words("sheer_open", CURTAIN_INTENT_GROUPS["sheer_open"])
        sheer_close_words = self.curtain_words("sheer_close", CURTAIN_INTENT_GROUPS["sheer_close"])
        sheer_stop_words = self.curtain_words("sheer_stop", CURTAIN_INTENT_GROUPS["sheer_stop"])
        window_open_words = self.curtain_words("window_sheer_open", CURTAIN_INTENT_GROUPS["window_sheer_open"])
        window_close_words = self.curtain_words("window_sheer_close", CURTAIN_INTENT_GROUPS["window_sheer_close"])
        window_stop_words = self.curtain_words("window_sheer_stop", CURTAIN_INTENT_GROUPS["window_sheer_stop"])

        def mode_setup_actions(mode_word: str) -> list[dict[str, Any]]:
            mode_row = self.word_row(mode_word)
            return [
                *self.setting_entry_actions(
                    curtain_entry_word,
                    curtain_entry_tone_id,
                    curtain_entry_tone_text,
                    expect_log_values={"curtainSettingEntry": ["into set curtain type"]},
                ),
                say_action(
                    mode_word,
                    expect_asr=self.accepted_asr_variants(mode_word),
                    advisory_tone_id=self.setting_success_tone_id(),
                    advisory_tone_text=self.row_tone_text(mode_row, self.setting_success_tone_id(), self.setting_success_tone_text()),
                    advisory_log_values=self.merge_advisory_logs(
                        self.curtain_mode_log_expectations(mode_word),
                        self.config_saved_advisory_logs(),
                    ),
                    require_protocol_and_tone=False,
                    auto_wake=False,
                ),
            ]

        def command_case(name: str, precondition: str, setup_actions: list[dict[str, Any]], word: str, requirement_ids: list[str]) -> dict[str, Any]:
            row = self.word_row(word)
            protocol = norm_hex(row.get("send_protocol"))
            tone_id = self.row_tone_id(row, "TONE_ID_1")
            tone_text = self.row_tone_text(row, tone_id, "好的")
            return case(
                self.next_case_id("CURTAIN"),
                "CURTAIN",
                name,
                precondition,
                f"执行控制词【{word}】",
                "控制词被正确识别并触发对应播报与协议",
                protocol,
                tone_id,
                tone_text,
                "P1",
                [*setup_actions, say_action(word, expect_tone_id=tone_id, expect_tone_text=tone_text, expect_send_protocol=protocol, require_protocol_and_tone=False)],
                requirement_ids,
                "positive",
            )

        def synonym_case(name: str, words: list[str], setup_actions: list[dict[str, Any]] | None = None) -> dict[str, Any]:
            primary = words[0]
            row = self.word_row(primary)
            protocol = norm_hex(row.get("send_protocol"))
            tone_id = self.row_tone_id(row, "TONE_ID_1")
            tone_text = self.row_tone_text(row, tone_id, "好的")
            actions = [
                *(setup_actions or []),
                *[
                    say_action(
                        word,
                        expect_tone_id=tone_id,
                        expect_tone_text=tone_text,
                        expect_send_protocol=protocol,
                        require_protocol_and_tone=False,
                    )
                    for word in words[:2]
                ],
            ]
            return case(
                self.next_case_id("CURTAIN"),
                "CURTAIN",
                name,
                "设备处于默认窗帘模式",
                f"依次执行同意图词条【{' / '.join(words[:2])}】",
                "同意图同义词触发的协议和播报保持一致",
                protocol,
                tone_id,
                tone_text,
                "P1",
                actions,
                self.requirement_ids_direct("DER-CURTAIN-002"),
                "positive",
            )

        default_negative_word = cloth_open_words[0]
        default_state_setup_actions = self.restore_default_actions(note="用例前恢复默认状态")
        return [
            command_case("默认窗帘打开命令正确", "设备处于默认窗帘模式", default_state_setup_actions, default_open_words[0], self.requirement_ids_direct("DER-CURTAIN-001")),
            command_case("默认窗帘关闭命令正确", "设备处于默认窗帘模式", default_state_setup_actions, default_close_words[0], self.requirement_ids_direct("DER-CURTAIN-001")),
            command_case("默认窗帘停止命令正确", "设备处于默认窗帘模式", default_state_setup_actions, default_stop_words[0], self.requirement_ids_direct("DER-CURTAIN-001")),
            synonym_case("默认窗帘打开同义词一致", [*default_open_words], default_state_setup_actions),
            synonym_case("默认窗帘关闭同义词一致", [*default_close_words], default_state_setup_actions),
            synonym_case("默认窗帘停止同义词一致", [*default_stop_words], default_state_setup_actions),
            case(
                self.next_case_id("CURTAIN"),
                "CURTAIN",
                "默认窗帘模式下布帘控制词不响应",
                "设备处于默认窗帘模式",
                f"直接执行布帘控制词【{default_negative_word}】",
                "默认窗帘模式下，非当前模式控制词不应响应",
                "",
                "",
                "",
                "P1",
                [
                    *default_state_setup_actions,
                    say_action(default_negative_word, expect_no_response=True, observe_s=2.0),
                ],
                self.requirement_ids_direct("DER-CURTAIN-004"),
                "negative",
            ),
            command_case("布帘模式下打开布帘命令正确", "设备处于布帘模式", mode_setup_actions("布帘模式"), cloth_open_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("布帘模式下关闭布帘命令正确", "设备处于布帘模式", mode_setup_actions("布帘模式"), cloth_close_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("布帘模式下停止布帘命令正确", "设备处于布帘模式", mode_setup_actions("布帘模式"), cloth_stop_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("纱帘模式下打开纱帘命令正确", "设备处于纱帘模式", mode_setup_actions("纱帘模式"), sheer_open_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("纱帘模式下关闭纱帘命令正确", "设备处于纱帘模式", mode_setup_actions("纱帘模式"), sheer_close_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("纱帘模式下停止纱帘命令正确", "设备处于纱帘模式", mode_setup_actions("纱帘模式"), sheer_stop_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("窗纱模式下打开窗纱命令正确", "设备处于窗纱模式", mode_setup_actions("窗纱模式"), window_open_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("窗纱模式下关闭窗纱命令正确", "设备处于窗纱模式", mode_setup_actions("窗纱模式"), window_close_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
            command_case("窗纱模式下停止窗纱命令正确", "设备处于窗纱模式", mode_setup_actions("窗纱模式"), window_stop_words[0], self.requirement_ids_direct("DER-CURTAIN-003")),
        ]

    def build_phrase_cases(self) -> list[dict[str, Any]]:
        curtain_entry = self.behavior_rules.get("curtain_mode_setting", {}).get("entry", {})
        curtain_entry_word = norm_text(curtain_entry.get("semantic"), "设置窗帘模式")
        curtain_entry_tone_id = norm_text(curtain_entry.get("tone_id"), "TONE_ID_9")
        curtain_entry_tone_text = norm_text(curtain_entry.get("tts_text"), self.tone_text(curtain_entry_tone_id))

        def curtain_mode_setup_actions(mode_word: str) -> list[dict[str, Any]]:
            mode_row = self.word_row(mode_word)
            return [
                *self.setting_entry_actions(
                    curtain_entry_word,
                    curtain_entry_tone_id,
                    curtain_entry_tone_text,
                    expect_log_values={"curtainSettingEntry": ["into set curtain type"]},
                ),
                say_action(
                    mode_word,
                    expect_asr=self.accepted_asr_variants(mode_word),
                    advisory_tone_id=self.setting_success_tone_id(),
                    advisory_tone_text=self.row_tone_text(mode_row, self.setting_success_tone_id(), self.setting_success_tone_text()),
                    advisory_log_values=self.merge_advisory_logs(
                        self.curtain_mode_log_expectations(mode_word),
                        self.config_saved_advisory_logs(),
                    ),
                    require_protocol_and_tone=False,
                    auto_wake=False,
                ),
            ]

        def phrase_case(name: str, precondition: str, setup_actions: list[dict[str, Any]], words: list[str]) -> dict[str, Any]:
            actions = [
                *setup_actions,
                *[
                    self.positive_command_action(word, accept_asr=list(words), label=f"验证词条【{word}】")
                    for word in words
                    if norm_text(word)
                ],
            ]
            return case(
                self.next_case_id("PHRASE"),
                "PHRASE",
                name,
                precondition,
                f"依次验证词组【{' / '.join(words)}】的等价识别",
                "同组词条均可被接受为同类合法意图",
                "",
                "",
                "",
                "P2",
                actions,
                self.requirement_ids_direct("DER-PHRASE-001", "DER-PHRASE-002"),
                "positive",
            )

        mixed_phrase_items = [
            {"word": word, "accept_asr": [word]}
            for word in [
                norm_text(self.behavior_rules.get("volume", {}).get("increase_word"), "大声点"),
                norm_text(self.behavior_rules.get("volume", {}).get("decrease_word"), "小声点"),
                "恢复出厂模式",
                "配对遥控器",
                "清除遥控器",
            ]
            if norm_text(word)
        ]
        default_state_setup_actions = self.restore_default_actions(note="用例前恢复默认状态")

        return [
            phrase_case("默认打开窗帘词组等价识别", "设备处于默认窗帘模式", default_state_setup_actions, self.curtain_words("default_open", CURTAIN_INTENT_GROUPS["default_open"])),
            phrase_case("默认关闭窗帘词组等价识别", "设备处于默认窗帘模式", default_state_setup_actions, self.curtain_words("default_close", CURTAIN_INTENT_GROUPS["default_close"])),
            phrase_case("默认停止窗帘词组等价识别", "设备处于默认窗帘模式", default_state_setup_actions, self.curtain_words("default_stop", CURTAIN_INTENT_GROUPS["default_stop"])),
            phrase_case("布帘打开词组等价识别", "设备处于布帘模式", curtain_mode_setup_actions("布帘模式"), self.curtain_words("cloth_open", CURTAIN_INTENT_GROUPS["cloth_open"])),
            phrase_case("布帘关闭词组等价识别", "设备处于布帘模式", curtain_mode_setup_actions("布帘模式"), self.curtain_words("cloth_close", CURTAIN_INTENT_GROUPS["cloth_close"])),
            phrase_case("布帘停止词组等价识别", "设备处于布帘模式", curtain_mode_setup_actions("布帘模式"), self.curtain_words("cloth_stop", CURTAIN_INTENT_GROUPS["cloth_stop"])),
            phrase_case("纱帘打开词组等价识别", "设备处于纱帘模式", curtain_mode_setup_actions("纱帘模式"), self.curtain_words("sheer_open", CURTAIN_INTENT_GROUPS["sheer_open"])),
            phrase_case("纱帘关闭词组等价识别", "设备处于纱帘模式", curtain_mode_setup_actions("纱帘模式"), self.curtain_words("sheer_close", CURTAIN_INTENT_GROUPS["sheer_close"])),
            phrase_case("纱帘停止词组等价识别", "设备处于纱帘模式", curtain_mode_setup_actions("纱帘模式"), self.curtain_words("sheer_stop", CURTAIN_INTENT_GROUPS["sheer_stop"])),
            phrase_case("窗纱打开词组等价识别", "设备处于窗纱模式", curtain_mode_setup_actions("窗纱模式"), self.curtain_words("window_sheer_open", CURTAIN_INTENT_GROUPS["window_sheer_open"])),
            phrase_case("窗纱关闭词组等价识别", "设备处于窗纱模式", curtain_mode_setup_actions("窗纱模式"), self.curtain_words("window_sheer_close", CURTAIN_INTENT_GROUPS["window_sheer_close"])),
            phrase_case("窗纱停止词组等价识别", "设备处于窗纱模式", curtain_mode_setup_actions("窗纱模式"), self.curtain_words("window_sheer_stop", CURTAIN_INTENT_GROUPS["window_sheer_stop"])),
            case(
                self.next_case_id("PHRASE"),
                "PHRASE",
                "连续未识别后可重唤醒并继续剩余词条",
                "设备处于默认待命态",
                "按杂项词组连续检查，若过程中超时或未识别，应可重唤醒后继续后续词条",
                "连续检查发生超时后，可重新唤醒并继续后续词条验证",
                "",
                "",
                "",
                "P2",
                [phrase_check_action(mixed_phrase_items, self.wake_word)],
                self.requirement_ids_direct("DER-PHRASE-003"),
                "state_transition",
            ),
        ]

    def build_selector_cases(self) -> list[dict[str, Any]]:
        candidates = [norm_text(item.get("semantic")) for item in self.behavior_rules.get("selector_mode", {}).get("candidates", []) if norm_text(item.get("semantic"))]
        entry = self.behavior_rules.get("wake_word_setting", {}).get("entry", {})
        timeout_s = float(self.behavior_rules.get("base", {}).get("wake_timeout_s", 15))
        wake_timeout_wait_s = max(timeout_s + 15.0, 35.0)
        entry_word = norm_text(entry.get("semantic"), "设置唤醒词")
        entry_tone_id = norm_text(entry.get("tone_id"), "TONE_ID_9")
        entry_tone_text = norm_text(entry.get("tts_text"), self.tone_text(entry_tone_id))
        if not candidates:
            return []
        first = candidates[0]
        second = candidates[1] if len(candidates) > 1 else "餐厅窗帘"

        def selector_setup_actions(candidate: str) -> list[dict[str, Any]]:
            candidate_row = self.word_row(candidate)
            candidate_tone_id = self.setting_success_tone_id()
            candidate_tone_text = self.row_tone_text(candidate_row, candidate_tone_id, self.setting_success_tone_text())
            return [
                with_label(wake_action(self.wake_word), f"播放默认唤醒词【{self.wake_word}】"),
                with_label(
                    say_action(
                        entry_word,
                        expect_asr=self.accepted_asr_variants(entry_word),
                        advisory_tone_id=entry_tone_id,
                        advisory_tone_text=entry_tone_text,
                        require_protocol_and_tone=False,
                        auto_wake=False,
                    ),
                    f"进入唤醒词设置【{entry_word}】",
                ),
                with_label(
                    say_action(
                        candidate,
                        expect_asr=self.accepted_asr_variants(candidate),
                        advisory_tone_id=candidate_tone_id,
                        advisory_tone_text=candidate_tone_text,
                        advisory_log_values=self.merge_advisory_logs(
                            self.wakeword_log_expectations(candidate),
                            self.config_saved_advisory_logs(),
                        ),
                        require_protocol_and_tone=False,
                        auto_wake=False,
                    ),
                    f"设置当前唤醒词为【{candidate}】",
                ),
                with_label(wait_action(2.0), f"等待【{candidate}】配置稳定"),
            ]

        cases = [
            case(
                self.next_case_id("SELECTOR"),
                "SELECTOR",
                f"定向唤醒词【{first}】连续唤醒稳定",
                f"设备已切换为【{first}】唤醒词",
                f"连续 3 次使用【{first}】唤醒设备",
                "选中的定向唤醒词可以稳定唤醒设备",
                "",
                "TONE_ID_0",
                self.tone_text("TONE_ID_0"),
                "P2",
                [
                    *selector_setup_actions(first),
                    with_label(wake_repeat_action(first, repeats=3, idle_timeout_s=wake_timeout_wait_s), f"连续 3 次使用【{first}】唤醒均成功"),
                ],
                self.requirement_ids_direct("DER-SELECTOR-001"),
                "state_transition",
            ),
            case(
                self.next_case_id("SELECTOR"),
                "SELECTOR",
                f"定向唤醒词【{first}】生效时【{second}】不应误唤醒",
                f"设备已切换为【{first}】唤醒词",
                f"使用未设置的候选唤醒词【{second}】尝试唤醒设备",
                "未选中的候选唤醒词不应误唤醒设备",
                "",
                "",
                "",
                "P2",
                [
                    *selector_setup_actions(first),
                    with_label(no_wake_action(second, repeats=2), f"验证未设置的候选唤醒词【{second}】不应误唤醒"),
                ],
                self.requirement_ids_direct("DER-SELECTOR-002"),
                "negative",
            ),
            case(
                self.next_case_id("SELECTOR"),
                "SELECTOR",
                f"切换到【{second}】后旧定向唤醒词【{first}】失效",
                f"设备已切换为【{second}】唤醒词",
                f"使用旧唤醒词【{first}】尝试唤醒设备",
                "切换到新定向唤醒词后，旧定向唤醒词失效",
                "",
                "",
                "",
                "P2",
                [
                    *selector_setup_actions(second),
                    with_label(no_wake_action(first), f"验证旧定向唤醒词【{first}】不应再唤醒设备"),
                ],
                self.requirement_ids_direct("DER-SELECTOR-003"),
                "negative",
            ),
            case(
                self.next_case_id("SELECTOR"),
                "SELECTOR",
                "默认唤醒词在定向唤醒切换后仍可用",
                f"设备已切换为【{first}】唤醒词",
                f"使用默认唤醒词【{self.wake_word}】唤醒设备",
                "默认唤醒词在定向唤醒切换后仍可用",
                "",
                "TONE_ID_0",
                self.tone_text("TONE_ID_0"),
                "P2",
                [
                    *selector_setup_actions(first),
                    with_label(wake_action(self.wake_word), f"验证默认唤醒词【{self.wake_word}】仍可唤醒设备"),
                ],
                self.requirement_ids_direct("WAKE-001"),
                "positive",
            ),
        ]
        return cases

    def build_ctrl_cases(self) -> list[dict[str, Any]]:
        rules = self.behavior_rules.get("remote_pairing", {})
        entry = rules.get("entry", {})
        success = rules.get("success", {})
        failure = rules.get("failure", {})
        clear = rules.get("clear", {})
        timeout_s = float(rules.get("window_timeout_s", self.behavior_rules.get("base", {}).get("wake_timeout_s", 15)))
        timeout_tone_id = norm_text(rules.get("timeout_tone_id"), "TONE_ID_2")
        entry_word = norm_text(entry.get("semantic"), "配对遥控器")
        entry_tone_id = norm_text(entry.get("tone_id"), "TONE_ID_6")
        entry_tone_text = norm_text(entry.get("tts_text"), self.tone_text(entry_tone_id) or "开始配对")
        success_protocol = norm_hex(success.get("recv_protocol"))
        failure_protocol = norm_hex(failure.get("recv_protocol"))
        clear_word = norm_text(clear.get("semantic"), "清除遥控器")
        clear_tone_id = norm_text(clear.get("tone_id"), "TONE_ID_1")
        clear_tone_text = norm_text(clear.get("tts_text"), self.tone_text(clear_tone_id) or "好的")
        self.case_counters["CTRL"] = max(self.case_counters["CTRL"], 8)
        cases = [
            case(
                "TC_CTRL_001",
                "CTRL",
                entry_word,
                "设备处于默认待命态",
                f"唤醒后说【{entry_word}】进入遥控器配置窗口",
                "设备进入遥控器配置窗口并发送配对协议",
                norm_hex(entry.get("send_protocol")),
                entry_tone_id,
                entry_tone_text,
                "P1",
                [say_action(entry_word, expect_tone_id=entry_tone_id, expect_tone_text=entry_tone_text, expect_send_protocol=norm_hex(entry.get("send_protocol")), require_protocol_and_tone=False)],
                self.requirement_ids_direct("DER-CTRL-001"),
                "state_transition",
            ),
            case(
                "TC_CTRL_002",
                "CTRL",
                "配对成功-被动播报",
                "已通过语音进入遥控器配置窗口",
                "进入配对窗口后注入成功协议，检查 receive msg 与成功播报",
                "窗口内收到 receive msg 并播报配对成功",
                "-",
                norm_text(success.get("tone_id"), "TONE_ID_7"),
                norm_text(success.get("tts_text"), self.tone_text(norm_text(success.get("tone_id"), "TONE_ID_7")) or "配对成功"),
                "P1",
                [
                    say_action(entry_word, expect_tone_id=entry_tone_id, expect_tone_text=entry_tone_text, expect_send_protocol=norm_hex(entry.get("send_protocol")), allow_fail_continue=True, advisory_failure=True, require_protocol_and_tone=False),
                    inject_protocol_action(success_protocol, expect_tone_id=norm_text(success.get("tone_id"), "TONE_ID_7"), expect_tone_text=norm_text(success.get("tts_text"), self.tone_text(norm_text(success.get("tone_id"), "TONE_ID_7")) or "配对成功"), expect_recv_protocol=success_protocol, require_receive_msg=True, requires_active_window=True),
                ],
                self.requirement_ids_direct("DER-CTRL-002"),
                "state_transition",
            ),
            case(
                "TC_CTRL_003",
                "CTRL",
                "配对失败-被动播报",
                "已通过语音进入遥控器配置窗口",
                "进入配对窗口后注入失败协议，检查 receive msg 与失败播报",
                "窗口内收到 receive msg 并播报配对失败",
                "-",
                norm_text(failure.get("tone_id"), "TONE_ID_8"),
                norm_text(failure.get("tts_text"), self.tone_text(norm_text(failure.get("tone_id"), "TONE_ID_8")) or "配对失败"),
                "P1",
                [
                    say_action(entry_word, expect_tone_id=entry_tone_id, expect_tone_text=entry_tone_text, expect_send_protocol=norm_hex(entry.get("send_protocol")), allow_fail_continue=True, advisory_failure=True, require_protocol_and_tone=False),
                    inject_protocol_action(failure_protocol, expect_tone_id=norm_text(failure.get("tone_id"), "TONE_ID_8"), expect_tone_text=norm_text(failure.get("tts_text"), self.tone_text(norm_text(failure.get("tone_id"), "TONE_ID_8")) or "配对失败"), expect_recv_protocol=failure_protocol, require_receive_msg=True, requires_active_window=True),
                ],
                self.requirement_ids_direct("DER-CTRL-003"),
                "state_transition",
            ),
            case(
                "TC_CTRL_004",
                "CTRL",
                "配对窗口超时后注入无效",
                "设备处于默认待命态",
                f"先进入配对窗口，等待 {int(timeout_s)}s 超时退出，再注入配对成功协议",
                "配对窗口超时后关闭，窗口外注入协议不会再触发 receive msg 或播报",
                "-",
                timeout_tone_id,
                self.tone_text(timeout_tone_id),
                "P1",
                [
                    say_action(entry_word, expect_tone_id=entry_tone_id, expect_tone_text=entry_tone_text, expect_send_protocol=norm_hex(entry.get("send_protocol")), require_protocol_and_tone=False),
                    wait_action(timeout_s, expect_tone_id=timeout_tone_id, expect_markers=["TIME_OUT"], measure_timeout=True),
                    inject_protocol_action(success_protocol, observe_s=2.0, expect_no_response=True),
                ],
                self.requirement_ids_direct("DER-CTRL-004"),
                "negative",
            ),
            case(
                "TC_CTRL_005",
                "CTRL",
                "配对成功后窗口关闭",
                "设备处于默认待命态",
                "先进入配对窗口并注入配对成功协议，再次注入协议",
                "配对成功后窗口关闭，再次注入协议不会再出现 receive msg 或播报",
                "-",
                norm_text(success.get("tone_id"), "TONE_ID_7"),
                norm_text(success.get("tts_text"), self.tone_text(norm_text(success.get("tone_id"), "TONE_ID_7")) or "配对成功"),
                "P1",
                [
                    say_action(entry_word, expect_tone_id=entry_tone_id, expect_tone_text=entry_tone_text, expect_send_protocol=norm_hex(entry.get("send_protocol")), allow_fail_continue=True, advisory_failure=True, require_protocol_and_tone=False),
                    inject_protocol_action(success_protocol, expect_tone_id=norm_text(success.get("tone_id"), "TONE_ID_7"), expect_tone_text=norm_text(success.get("tts_text"), self.tone_text(norm_text(success.get("tone_id"), "TONE_ID_7")) or "配对成功"), expect_recv_protocol=success_protocol, require_receive_msg=True, requires_active_window=True),
                    inject_protocol_action(success_protocol, observe_s=2.0, expect_no_response=True),
                ],
                self.requirement_ids_direct("DER-CTRL-002"),
                "state_transition",
            ),
            case(
                "TC_CTRL_006",
                "CTRL",
                "配对失败后窗口关闭",
                "设备处于默认待命态",
                "先进入配对窗口并注入配对失败协议，再次注入协议",
                "配对失败后窗口关闭，再次注入协议不会再出现 receive msg 或播报",
                "-",
                norm_text(failure.get("tone_id"), "TONE_ID_8"),
                norm_text(failure.get("tts_text"), self.tone_text(norm_text(failure.get("tone_id"), "TONE_ID_8")) or "配对失败"),
                "P1",
                [
                    say_action(entry_word, expect_tone_id=entry_tone_id, expect_tone_text=entry_tone_text, expect_send_protocol=norm_hex(entry.get("send_protocol")), allow_fail_continue=True, advisory_failure=True, require_protocol_and_tone=False),
                    inject_protocol_action(failure_protocol, expect_tone_id=norm_text(failure.get("tone_id"), "TONE_ID_8"), expect_tone_text=norm_text(failure.get("tts_text"), self.tone_text(norm_text(failure.get("tone_id"), "TONE_ID_8")) or "配对失败"), expect_recv_protocol=failure_protocol, require_receive_msg=True, requires_active_window=True),
                    inject_protocol_action(failure_protocol, observe_s=2.0, expect_no_response=True),
                ],
                self.requirement_ids_direct("DER-CTRL-003"),
                "state_transition",
            ),
            case(
                "TC_CTRL_007",
                "CTRL",
                clear_word,
                "设备处于默认待命态",
                f"唤醒后说【{clear_word}】清除遥控器配置",
                "设备发送清除协议并返回清除确认",
                norm_hex(clear.get("send_protocol")),
                clear_tone_id,
                clear_tone_text,
                "P1",
                [say_action(clear_word, expect_tone_id=clear_tone_id, expect_tone_text=clear_tone_text, expect_send_protocol=norm_hex(clear.get("send_protocol")), require_protocol_and_tone=False)],
                self.requirement_ids_direct("DER-CTRL-005"),
                "positive",
            ),
        ]
        return cases
def action_text(action: dict[str, Any]) -> str:
    custom = norm_text(action.get("label"))
    if custom:
        return custom
    kind = action["type"]
    if kind == "wake":
        return f"播放唤醒词【{action['word']}】"
    if kind == "say":
        return f"播放命令词【{action['word']}】"
    if kind == "assert_wake_repeats":
        return f"连续唤醒【{action['word']}】{action.get('repeats', 1)}次"
    if kind == "assert_no_wake":
        words = action.get("words") or [action.get("word", "")]
        return f"静默断言【{' / '.join(str(item) for item in words if str(item).strip())}】"
    if kind == "inject_protocol":
        return f"注入协议【{action['protocol']}】"
    if kind == "wait":
        return f"等待 {action['seconds']}s"
    if kind == "volume_walk":
        return f"音量遍历【{action['command_word']}】"
    if kind == "manual_power_cycle":
        return "执行上下电循环"
    if kind == "phrase_check":
        return f"词条组检查【{len(action.get('items', []))}条】"
    if kind == "reboot":
        return "执行 reboot"
    return kind




def render_case_steps(actions: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    for action in actions:
        if action.get("type") == "say" and bool(action.get("auto_wake", True)):
            wake_word = norm_text(action.get("wake_word"), "你好杜亚")
            lines.append(f"{len(lines) + 1}. 播放唤醒词【{wake_word}】")
            say_clone = dict(action)
            say_clone["auto_wake"] = False
            lines.append(f"{len(lines) + 1}. {action_text(say_clone)}")
            continue
        lines.append(f"{len(lines) + 1}. {action_text(action)}")
    return "\n".join(lines)
def write_workbook(cases: list[dict[str, Any]], path: Path, spec: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    headers = [
        "用例编号",
        "功能模块",
        "覆盖需求",
        "阻塞依赖用例",
        "用例类型",
        "用例名称",
        "前置条件",
        "测试步骤",
        "期望结果",
        "期望发送协议(UART1)",
        "期望播报ID",
        "期望播报内容",
        "优先级",
        "测试结果",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row, item in enumerate(cases, 2):
        values = [
            item["case_id"],
            item["module"],
            ", ".join(item.get("requirement_ids", [])),
            ", ".join(item.get("blocked_by_case_ids", [])),
            item.get("case_type", ""),
            item["name"],
            item["precondition"],
            item["steps"],
            item["expected_result"],
            item["expected_protocol"],
            item["expected_tone_id"],
            item["expected_tone_text"],
            item["priority"],
            "",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)

    plan = wb.create_sheet("执行编排")
    for col, header in enumerate(["用例编号", "模块", "覆盖需求", "阻塞依赖用例", "用例类型", "优先级", "执行动作", "动作序列JSON"], 1):
        plan.cell(row=1, column=col, value=header)
    for row, item in enumerate(cases, 2):
        plan.cell(row=row, column=1, value=item["case_id"])
        plan.cell(row=row, column=2, value=item["module"])
        plan.cell(row=row, column=3, value=", ".join(item.get("requirement_ids", [])))
        plan.cell(row=row, column=4, value=", ".join(item.get("blocked_by_case_ids", [])))
        plan.cell(row=row, column=5, value=item.get("case_type", ""))
        plan.cell(row=row, column=6, value=item["priority"])
        plan.cell(row=row, column=7, value="\n".join(action_text(a) for a in item["actions"]))
        plan.cell(row=row, column=8, value=json.dumps(item["actions"], ensure_ascii=False, indent=2))

    coverage = wb.create_sheet("覆盖矩阵")
    for col, header in enumerate(["模块", "需求ID", "需求描述", "策略", "验收层级", "来源类型", "来源章节", "验证目标"], 1):
        coverage.cell(row=1, column=col, value=header)
    for row, item in enumerate(spec.get("coverage_matrix", []), 2):
        coverage.cell(row=row, column=1, value=item.get("module"))
        coverage.cell(row=row, column=2, value=item.get("requirement_id"))
        coverage.cell(row=row, column=3, value=item.get("requirement"))
        coverage.cell(row=row, column=4, value=", ".join(item.get("case_strategy", [])))
        coverage.cell(row=row, column=5, value=item.get("acceptance_level", "main"))
        coverage.cell(row=row, column=6, value=item.get("source_kind", ""))
        coverage.cell(row=row, column=7, value=item.get("source_section", ""))
        coverage.cell(row=row, column=8, value=item.get("validation_target", "runtime_case"))

    tone = wb.create_sheet("播报ID对照表")
    tone.cell(row=1, column=1, value="TONE_ID")
    tone.cell(row=1, column=2, value="ID编号")
    tone.cell(row=1, column=3, value="播报内容")
    for row, item in enumerate(spec.get("tones", {}).get("items", []), 2):
        tone.cell(row=row, column=1, value=item.get("name"))
        tone.cell(row=row, column=2, value=item.get("id"))
        tone.cell(row=row, column=3, value=item.get("text"))

    wb.save(path)


def main() -> None:
    GENERATED_DIR.mkdir(exist_ok=True)
    spec = load_json(WORK_DIR / "normalized_spec.json")
    cases = RequirementCaseBuilder(spec).build()
    write_json(GENERATED_DIR / "cases.json", cases)
    write_workbook(cases, GENERATED_DIR / "CSK5062_杜亚窗帘_测试用例.xlsx", spec)
    print(f"Wrote {GENERATED_DIR / 'cases.json'}")
    print(f"Wrote {GENERATED_DIR / 'CSK5062_杜亚窗帘_测试用例.xlsx'}")
    print(f"case_count={len(cases)}")


if __name__ == "__main__":
    main()
