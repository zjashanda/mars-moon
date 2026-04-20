#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Build a requirement-driven normalized spec for the Dooya curtain project."""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from serial_port_config import get_serial_defaults


ROOT = Path(__file__).resolve().parents[1]
WORK_DIR = ROOT / "work"
REQ_DIR = ROOT / "需求"
REF_DIR = ROOT / "用例skill参考"
ASSET_REF_DIR = ROOT / "assets" / "reference"
WORD_TABLE = REQ_DIR / "词条处理.xlsx"
TONE_FILE = REQ_DIR / "tone.h"
REQ_DOC = REQ_DIR / "需求文档.md"
SUPPLEMENT_FILE = REQ_DIR / "需求和用例补充说明.txt"
REFERENCE_XLSX = REF_DIR / "CSK5062_杜亚窗帘_测试用例v2.xlsx"
RAW_INPUT_EXCLUDES = {
    ".git",
    "__pycache__",
    "assets",
    "generated",
    "result",
    "sample",
    "scripts",
    "tools",
    "work",
}
MODULE_ORDER = [
    "BOOT",
    "BASE",
    "VOL",
    "UART",
    "POWER",
    "FACTORY",
    "WAKEWORD",
    "WORKMODE",
    "CURTAINMODE",
    "CURTAIN",
    "PHRASE",
    "SELECTOR",
    "CTRL",
]
REQUIREMENT_MODULE_ORDER = ["CONFIG", *MODULE_ORDER]


DEFAULT_RUNTIME_PORTS = get_serial_defaults()


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore") if path.is_file() else ""


def path_uses_excluded_parent(path: Path, root: Path, excluded: set[str]) -> bool:
    try:
        relative = path.relative_to(root)
    except ValueError:
        return False
    return any(part in excluded for part in relative.parts[:-1])


def find_latest_matching_file(root: Path, patterns: list[str], excluded_dirs: set[str] | None = None) -> Path | None:
    matches: list[Path] = []
    for pattern in patterns:
        for path in root.rglob(pattern):
            if not path.is_file():
                continue
            if excluded_dirs and path_uses_excluded_parent(path, root, excluded_dirs):
                continue
            matches.append(path)
    if not matches:
        return None
    matches.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return matches[0]


def resolve_preferred_input(preferred: Path, patterns: list[str]) -> tuple[Path | None, str]:
    if preferred.is_file():
        return preferred, "preferred_raw"
    discovered = find_latest_matching_file(ROOT, patterns, excluded_dirs=RAW_INPUT_EXCLUDES)
    if discovered:
        return discovered, "discovered_raw"
    return None, "missing_raw"


def first_existing_path(*candidates: Path) -> Path | None:
    for candidate in candidates:
        if candidate and candidate.is_file():
            return candidate
    return None


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def norm_hex(value: Any) -> str:
    return " ".join(clean_text(value).upper().split())


def safe_int(value: str, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def safe_bool(value: str, default: bool = False) -> bool:
    text = clean_text(value)
    if text in {"是", "true", "True", "1", "yes", "Yes", "Y", "y", "保存", "保留"}:
        return True
    if text in {"否", "false", "False", "0", "no", "No", "N", "n", "不保存", "不保留"}:
        return False
    return default


def extract_baudrate(text: str, labels: list[str], default: int) -> int:
    lines = text.splitlines()
    for raw_line in lines:
        line = clean_text(raw_line)
        if not line:
            continue
        line_lower = line.lower()
        if any(label.lower() in line_lower for label in labels):
            match = re.search(r"波特率\s*[：: ]\s*(\d+)", line, re.IGNORECASE)
            if not match:
                match = re.search(r"(\d{3,})", line)
            if match:
                return safe_int(match.group(1), default)
    return default


def parse_requirement_text(text: str) -> dict[str, Any]:
    patterns = {
        "project_name": (r"项目名称[：:]\s*([^\n]+)", "CSK5062杜亚窗帘"),
        "branch_name": (r"分支名称[：:]\s*([^\n]+)", "unknown"),
        "chip_model": (r"芯片型号[：:]\s*([^\n]+)", "CSK5062"),
        "firmware_version": (r"固件版本[：:]\s*([^\n]+)", "unknown"),
    }
    values: dict[str, Any] = {}
    for key, (pattern, default) in patterns.items():
        match = re.search(pattern, text)
        values[key] = match.group(1).strip() if match else default

    wake_timeout = re.search(r"唤醒时长[：:]\s*(\d+)\s*s", text, re.IGNORECASE)
    volume_levels = re.search(r"音量档位[：:]\s*(\d+)", text)
    default_volume = re.search(r"(?:初始|默认)音量[：:]\s*(\d+)", text)
    protocol_baud = extract_baudrate(text, ["协议串口", "协议口", "uart1"], 9600)
    log_baud = extract_baudrate(text, ["日志串口", "日志口", "uart0"], 115200)
    wake_word_power = re.search(r"唤醒词掉电保存[：:]\s*([^\n]+)", text)
    volume_power = re.search(r"音量掉电保存[：:]\s*([^\n]+)", text)
    default_work_mode = "语音模式" if "默认为语音模式" in text else "语音模式"
    default_curtain_mode = "窗帘模式" if "默认为窗帘模式" in text or "默认的窗帘模式一直生效" in text else "窗帘模式"
    values.update(
        {
            "wake_timeout_s": safe_int(wake_timeout.group(1), 15) if wake_timeout else 15,
            "volume_levels": safe_int(volume_levels.group(1), 4) if volume_levels else 4,
            "default_volume": safe_int(default_volume.group(1), 3) if default_volume else 3,
            "protocol_baudrate": protocol_baud,
            "log_baudrate": log_baud,
            "wake_word_power_retained": safe_bool(wake_word_power.group(1), False) if wake_word_power else False,
            "volume_power_retained": safe_bool(volume_power.group(1), False) if volume_power else False,
            "default_work_mode": default_work_mode,
            "default_curtain_mode": default_curtain_mode,
        }
    )
    return values


def parse_tone_file(path: Path) -> list[dict[str, Any]]:
    text = read_text(path)
    items: list[dict[str, Any]] = []
    pattern = re.compile(r"(TONE_ID_\d+)\s*=\s*(\d+),\s*//\s*(.+)")
    for line in text.splitlines():
        match = pattern.search(line)
        if not match:
            continue
        name, ident, raw = match.groups()
        raw = re.sub(r"\.mp3$", "", raw.strip(), flags=re.IGNORECASE)
        raw = re.sub(r"^\d+_", "", raw)
        items.append({"name": name, "id": int(ident), "text": raw})
    return items


def parse_word_sheet(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[1]]
    rows: list[dict[str, Any]] = []
    wake_words: list[str] = []
    seen_semantics: set[str] = set()
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        semantic = clean_text(row[0] if len(row) > 0 else "")
        category = clean_text(row[1] if len(row) > 1 else "")
        tts_text = clean_text(row[2] if len(row) > 2 else "")
        mode = clean_text(row[3] if len(row) > 3 else "")
        send_protocol = norm_hex(row[4] if len(row) > 4 else "")
        recv_protocol = norm_hex(row[5] if len(row) > 5 else "")
        if not any([semantic, category, tts_text, send_protocol, recv_protocol]):
            continue
        item = {
            "row_index": idx,
            "semantic": semantic,
            "category": category,
            "tts_text": tts_text,
            "mode": mode,
            "send_protocol": send_protocol,
            "recv_protocol": recv_protocol,
        }
        rows.append(item)
        if semantic:
            seen_semantics.add(semantic)
        joined = f"{semantic} {category} {tts_text}"
        if "唤醒" in joined and semantic:
            wake_words.append(semantic)
    return {
        "rows": rows,
        "wake_words": wake_words,
        "semantic_count": len(seen_semantics),
    }


def infer_tone_id(tts_text: str, tones: list[dict[str, Any]]) -> str:
    def normalize_tone_text(value: str) -> str:
        text = clean_text(value)
        text = text.replace("嘀", "滴")
        text = text.replace("音效", "声")
        text = re.sub(r"[（）()\[\]【】,，。.\-_\s]+", "", text)
        return text

    target = clean_text(tts_text)
    if not target:
        return ""
    normalized_target = normalize_tone_text(target)
    for item in tones:
        text = clean_text(item.get("text"))
        normalized_text = normalize_tone_text(text)
        if (
            target == text
            or target in text
            or text in target
            or (normalized_target and normalized_target == normalized_text)
            or (normalized_target and normalized_target in normalized_text)
            or (normalized_text and normalized_text in normalized_target)
        ):
            return clean_text(item.get("name"))
    return ""


def row_lookup(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {clean_text(row.get("semantic")): row for row in rows if clean_text(row.get("semantic"))}


def find_rows(rows: list[dict[str, Any]], *keywords: str) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for row in rows:
        joined = " ".join(
            [
                clean_text(row.get("semantic")),
                clean_text(row.get("category")),
                clean_text(row.get("tts_text")),
                clean_text(row.get("mode")),
            ]
        )
        if all(keyword in joined for keyword in keywords if keyword):
            results.append(row)
    return results


def values_from_rows(rows: list[dict[str, Any]], field: str) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        value = clean_text(row.get(field))
        if value and value not in seen:
            seen.add(value)
            values.append(value)
    return values


def unique_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    for row in rows:
        key = (
            clean_text(row.get("semantic")),
            clean_text(row.get("category")),
            clean_text(row.get("tts_text")),
            norm_hex(row.get("send_protocol")),
            norm_hex(row.get("recv_protocol")),
        )
        if key in seen:
            continue
        seen.add(key)
        ordered.append(row)
    return ordered


def exact_rows(rows: list[dict[str, Any]], semantics: list[str]) -> list[dict[str, Any]]:
    lookup = row_lookup(rows)
    return [lookup[item] for item in semantics if item in lookup]


def first_row_by_category(rows: list[dict[str, Any]], keyword: str) -> dict[str, Any] | None:
    for row in rows:
        if keyword in clean_text(row.get("category")):
            return row
    return None


def infer_wake_word_candidate_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    excluded = {
        "你好杜亚",
        "设置唤醒词",
        "设置工作模式",
        "设置窗帘模式",
        "恢复出厂模式",
        "配对遥控器",
        "清除遥控器",
    }
    candidates: list[dict[str, Any]] = []
    for row in rows:
        semantic = clean_text(row.get("semantic"))
        category = clean_text(row.get("category"))
        tts_text = clean_text(row.get("tts_text"))
        if not semantic or semantic in excluded:
            continue
        if norm_hex(row.get("send_protocol")) or norm_hex(row.get("recv_protocol")):
            continue
        if "命令词" not in category:
            continue
        if "窗帘" not in semantic:
            continue
        if tts_text not in {"在呢", "好的"}:
            continue
        candidates.append(row)
    return unique_rows(candidates)


def curtain_group_rows(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    groups = {
        "default": exact_rows(rows, ["打开窗帘", "窗帘打开", "窗帘关闭", "关闭窗帘", "停止窗帘", "窗帘停止"]),
        "cloth": exact_rows(rows, ["打开布帘", "布帘打开", "关闭布帘", "布帘关闭", "停止布帘", "布帘停止"]),
        "sheer": exact_rows(rows, ["打开纱帘", "纱帘打开", "关闭纱帘", "纱帘关闭", "停止纱帘", "纱帘停止"]),
        "window_sheer": exact_rows(rows, ["打开窗纱", "窗纱打开", "窗纱关闭", "关闭窗纱", "停止窗纱", "窗纱停止"]),
    }
    return {key: unique_rows(value) for key, value in groups.items() if value}


def row_payload(row: dict[str, Any], tones: list[dict[str, Any]]) -> dict[str, Any]:
    tts_text = clean_text(row.get("tts_text"))
    return {
        "semantic": clean_text(row.get("semantic")),
        "category": clean_text(row.get("category")),
        "mode": clean_text(row.get("mode")),
        "send_protocol": norm_hex(row.get("send_protocol")),
        "recv_protocol": norm_hex(row.get("recv_protocol")),
        "tts_text": tts_text,
        "tone_id": infer_tone_id(tts_text, tones),
    }


def requirement_entry(
    module: str,
    requirement_id: str,
    requirement: str,
    *,
    source_kind: str,
    source_section: str,
    source_text: str,
    acceptance_level: str = "main",
    validation_target: str = "runtime_case",
    case_strategy: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "module": module,
        "requirement_id": requirement_id,
        "requirement": requirement,
        "source_kind": source_kind,
        "source_section": source_section,
        "source_text": source_text,
        "acceptance_level": acceptance_level,
        "validation_target": validation_target,
        "case_strategy": case_strategy or ["positive", "boundary", "state_transition"],
    }


def build_semantic_groups(rows: list[dict[str, Any]], tones: list[dict[str, Any]]) -> dict[str, Any]:
    wake_rows = exact_rows(rows, ["你好杜亚"]) + infer_wake_word_candidate_rows(rows)
    work_mode_rows = exact_rows(rows, ["语音模式", "滴答模式", "嘀嗒模式"])
    curtain_mode_rows = exact_rows(rows, ["纱帘模式", "窗纱模式", "布帘模式"])
    groups = {
        "wake_words": [row_payload(row, tones) for row in unique_rows(wake_rows)],
        "work_mode": [row_payload(row, tones) for row in unique_rows(work_mode_rows)],
        "curtain_mode": [row_payload(row, tones) for row in unique_rows(curtain_mode_rows)],
        "selector_mode": [row_payload(row, tones) for row in infer_wake_word_candidate_rows(rows)],
        "volume": [row_payload(row, tones) for row in find_rows(rows, "音量")],
        "remote_pairing": [row_payload(row, tones) for row in find_rows(rows, "遥控") + find_rows(rows, "配对")],
        "factory_reset": [row_payload(row, tones) for row in find_rows(rows, "恢复出厂")],
    }
    groups["phrases"] = [row_payload(row, tones) for row in rows if clean_text(row.get("semantic"))]
    return groups


def rows_from_legacy_spec(legacy_spec: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for item in legacy_spec.get("word_sheet", {}).get("rows", []):
        rows.append(
            {
                "semantic": clean_text(item.get("semantic")),
                "category": clean_text(item.get("func_type")),
                "tts_text": clean_text(item.get("tts_text")),
                "mode": clean_text(item.get("mode")),
                "send_protocol": norm_hex(item.get("send_protocol")),
                "recv_protocol": norm_hex(item.get("recv_protocol")),
            }
        )
    return rows


def load_latest_legacy_bundle(root: Path) -> dict[str, Any]:
    result_root = root / "result"
    empty = {
        "spec": None,
        "cases": None,
        "spec_path": None,
        "cases_path": None,
        "result_dir": None,
    }
    if not result_root.is_dir():
        return empty

    result_dirs = sorted(
        [item for item in result_root.iterdir() if item.is_dir()],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    for result_dir in result_dirs:
        spec_path = first_existing_path(
            result_dir / "artifacts" / "work" / "normalized_spec.json",
            result_dir / "normalized_spec.json",
        )
        cases_path = first_existing_path(
            result_dir / "cases.json",
            result_dir / "artifacts" / "generated" / "cases.json",
        )
        spec = None
        cases = None
        try:
            if spec_path:
                spec = json.loads(spec_path.read_text(encoding="utf-8"))
        except Exception:
            spec = None
        try:
            if cases_path:
                cases = json.loads(cases_path.read_text(encoding="utf-8"))
        except Exception:
            cases = None
        if spec or cases:
            return {
                "spec": spec,
                "cases": cases,
                "spec_path": spec_path,
                "cases_path": cases_path,
                "result_dir": result_dir,
            }
    return empty


def legacy_ctrl_action_protocol(legacy_cases: list[dict[str, Any]] | None, case_id: str) -> str:
    for case_item in legacy_cases or []:
        if clean_text(case_item.get("case_id")) != case_id:
            continue
        for action in case_item.get("actions", []):
            if clean_text(action.get("type")) != "inject_protocol":
                continue
            protocol = norm_hex(action.get("protocol") or action.get("expect_recv_protocol"))
            if protocol:
                return protocol
    return ""


def build_behavior_rules(
    meta: dict[str, Any],
    rows: list[dict[str, Any]],
    tones: list[dict[str, Any]],
    legacy_cases: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    lookup = row_lookup(rows)
    wake_candidate_rows = infer_wake_word_candidate_rows(rows)
    wake_words = values_from_rows(wake_candidate_rows, "semantic")
    if not wake_words:
        wake_words = [clean_text(word) for word in meta.get("fallback_wake_words", []) if clean_text(word)]

    pair_row = lookup.get("配对遥控器", {})
    clear_row = lookup.get("清除遥控器", {})
    pair_success_rows = [
        row for row in rows if "配对成功" in clean_text(row.get("tts_text")) or "配对成功" in clean_text(row.get("semantic"))
    ]
    pair_fail_rows = [
        row for row in rows if "配对失败" in clean_text(row.get("tts_text")) or "配对失败" in clean_text(row.get("semantic"))
    ]
    work_mode_rows = unique_rows(exact_rows(rows, ["语音模式", "滴答模式", "嘀嗒模式"]))
    curtain_mode_rows = unique_rows(exact_rows(rows, ["纱帘模式", "窗纱模式", "布帘模式"]))
    selector_rows = wake_candidate_rows
    increase_row = first_row_by_category(rows, "增大音量") or lookup.get("大声点", {})
    decrease_row = first_row_by_category(rows, "减小音量") or lookup.get("小声点", {})
    curtain_groups = curtain_group_rows(rows)

    success_rule = row_payload(pair_success_rows[0], tones) if pair_success_rows else {
        "semantic": "",
        "category": "control",
        "mode": "",
        "send_protocol": "",
        "recv_protocol": "",
        "tts_text": "配对成功",
        "tone_id": "TONE_ID_7",
    }
    failure_rule = row_payload(pair_fail_rows[0], tones) if pair_fail_rows else {
        "semantic": "",
        "category": "control",
        "mode": "",
        "send_protocol": "",
        "recv_protocol": "",
        "tts_text": "配对失败",
        "tone_id": "TONE_ID_8",
    }
    if not clean_text(success_rule.get("recv_protocol")):
        success_rule["recv_protocol"] = legacy_ctrl_action_protocol(legacy_cases, "TC_CTRL_002")
    if not clean_text(failure_rule.get("recv_protocol")):
        failure_rule["recv_protocol"] = legacy_ctrl_action_protocol(legacy_cases, "TC_CTRL_003")

    selector_candidates = [row_payload(row, tones) for row in selector_rows]

    return {
        "boot": {
            "default_state": "idle",
            "requires_log_ready": True,
            "post_boot_ready_delay_s": 4.0,
        },
        "base": {
            "wake_timeout_s": meta["wake_timeout_s"],
            "default_wake_word": "你好杜亚",
            "alternate_wake_words": wake_words,
            "requires_wake_before_command": True,
        },
        "volume": {
            "level_count": meta["volume_levels"],
            "default_level": meta["default_volume"],
            "step_tone_id": "TONE_ID_1",
            "upper_boundary_tone": "TONE_ID_4",
            "lower_boundary_tone": "TONE_ID_5",
            "increase_word": clean_text((increase_row or {}).get("semantic")) or "大声点",
            "decrease_word": clean_text((decrease_row or {}).get("semantic")) or "小声点",
            "restores_default_after_power_cycle": not bool(meta.get("volume_power_retained", False)),
            "restores_default_after_factory_reset": True,
        },
        "uart": {
            "protocol_port": str(DEFAULT_RUNTIME_PORTS["protocol_port"]),
            "protocol_baudrate": meta["protocol_baudrate"],
            "log_port": str(DEFAULT_RUNTIME_PORTS["log_port"]),
            "log_baudrate": meta["log_baudrate"],
            "frame_header": "55 AA",
            "frame_length": 8,
            "must_validate_links_before_run": True,
        },
        "power": {
            "supports_manual_power_cycle": True,
            "requires_log_recovery_after_power_cycle": True,
            "wake_word_retained_after_power_cycle": bool(meta.get("wake_word_power_retained", False)),
            "volume_retained_after_power_cycle": bool(meta.get("volume_power_retained", False)),
        },
        "factory_reset": {
            "entry": row_payload(lookup.get("恢复出厂模式", {}), tones),
            "must_restore_defaults": [
                "wake_word",
                "volume",
                "work_mode",
                "curtain_mode",
            ],
        },
        "wake_word_setting": {
            "entry": row_payload(lookup.get("设置唤醒词", {}), tones),
            "default_wake_word": "你好杜亚",
            "candidate_wake_words": wake_words,
        },
        "work_mode_setting": {
            "entry": row_payload(lookup.get("设置工作模式", {}), tones),
            "candidates": [row_payload(row, tones) for row in work_mode_rows],
            "default_mode": clean_text(meta.get("default_work_mode")) or "语音模式",
        },
        "curtain_mode_setting": {
            "entry": row_payload(lookup.get("设置窗帘模式", {}), tones),
            "candidates": [row_payload(row, tones) for row in curtain_mode_rows],
            "default_mode": clean_text(meta.get("default_curtain_mode")) or "窗帘模式",
        },
        "selector_mode": {
            "candidates": [row_payload(row, tones) for row in selector_rows],
            "default_selector": values_from_rows(selector_rows, "semantic")[0] if selector_rows else "",
        },
        "remote_pairing": {
            "entry": row_payload(pair_row, tones),
            "clear": row_payload(clear_row, tones),
            "success": success_rule,
            "failure": failure_rule,
            "requires_active_window": True,
            "requires_receive_msg": True,
            "window_timeout_s": meta["wake_timeout_s"],
            "timeout_tone_id": "TONE_ID_2",
            "success_window_closes_after_result": True,
            "failure_window_closes_after_result": True,
        },
        "phrase": {
            "semantic_count": len([row for row in rows if clean_text(row.get("semantic"))]),
            "requires_semantic_grouping": True,
        },
        "curtain_control": {
            "default_mode": clean_text(meta.get("default_curtain_mode")) or "窗帘模式",
            "groups": {
                key: [row_payload(row, tones) for row in value]
                for key, value in curtain_groups.items()
            },
        },
    }


def build_requirement_catalog(meta: dict[str, Any]) -> list[dict[str, Any]]:
    wake_timeout_s = int(meta.get("wake_timeout_s", 15) or 15)
    volume_levels = int(meta.get("volume_levels", 4) or 4)
    default_volume = int(meta.get("default_volume", 3) or 3)
    protocol_baudrate = int(meta.get("protocol_baudrate", 9600) or 9600)
    log_baudrate = int(meta.get("log_baudrate", 115200) or 115200)
    wake_word_power = bool(meta.get("wake_word_power_retained", False))
    volume_power = bool(meta.get("volume_power_retained", False))
    default_work_mode = clean_text(meta.get("default_work_mode")) or "语音模式"
    default_curtain_mode = clean_text(meta.get("default_curtain_mode")) or "窗帘模式"

    return [
        requirement_entry(
            "CONFIG",
            "CFG-001",
            f"唤醒时长为 {wake_timeout_s}s",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 固件配置",
            source_text=f"唤醒时长: {wake_timeout_s}s",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-002",
            f"音量档位数为 {volume_levels}",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 固件配置",
            source_text=f"音量档位: {volume_levels}",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-003",
            f"初始化默认音量为 {default_volume}",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 固件配置",
            source_text=f"初始化默认音量: {default_volume}",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-004",
            "最小音量下溢提示播报为“音量已最小”",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 固件配置",
            source_text="最小音量下溢提示播报: 音量已最小",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-005",
            "最大音量上溢提示播报为“音量已最大”",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 固件配置",
            source_text="最大音量上溢提示播报: 音量已最大",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-006",
            "mic 模拟增益为 32",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 音频输入配置",
            source_text="mic模拟增益: 32",
            validation_target="static_config",
            case_strategy=["static"],
        ),
        requirement_entry(
            "CONFIG",
            "CFG-007",
            "mic 数字增益为 2",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 音频输入配置",
            source_text="mic数字增益: 2",
            validation_target="static_config",
            case_strategy=["static"],
        ),
        requirement_entry(
            "CONFIG",
            "CFG-008",
            "协议串口内部编号为 UART1",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 串口配置",
            source_text="协议串口: UART1",
            validation_target="static_config",
            case_strategy=["static"],
        ),
        requirement_entry(
            "CONFIG",
            "CFG-009",
            f"协议串口波特率为 {protocol_baudrate}",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 串口配置",
            source_text=f"协议串口: UART1、波特率{protocol_baudrate}",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-010",
            "日志串口内部编号为 UART0",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 串口配置",
            source_text="日志串口: UART0",
            validation_target="static_config",
            case_strategy=["static"],
        ),
        requirement_entry(
            "CONFIG",
            "CFG-011",
            f"日志串口波特率为 {log_baudrate}",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 串口配置",
            source_text=f"日志串口: UART0、波特率{log_baudrate}",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-012",
            f"唤醒词掉电保存为 {'是' if wake_word_power else '否'}",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 掉电配置",
            source_text=f"唤醒词掉电保存: {'是' if wake_word_power else '否'}",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-013",
            f"音量掉电保存为 {'是' if volume_power else '否'}",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 掉电配置",
            source_text=f"音量掉电保存: {'是' if volume_power else '否'}",
        ),
        requirement_entry(
            "CONFIG",
            "CFG-014",
            "合成音频发音人为叶子",
            source_kind="requirement_doc",
            source_section="二、基础配置 / 播报配置",
            source_text="合成音频发音人: 叶子",
            validation_target="static_config",
            case_strategy=["static"],
        ),
        requirement_entry(
            "FACTORY",
            "FACT-001",
            "恢复出厂后恢复默认唤醒词",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 恢复出厂设置",
            source_text="需恢复默认唤醒词、默认音量、默认工作模式、默认窗帘类型",
        ),
        requirement_entry(
            "FACTORY",
            "FACT-002",
            "恢复出厂后恢复默认音量",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 恢复出厂设置",
            source_text="需恢复默认唤醒词、默认音量、默认工作模式、默认窗帘类型",
        ),
        requirement_entry(
            "FACTORY",
            "FACT-003",
            f"恢复出厂后恢复默认工作模式（{default_work_mode}）",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 恢复出厂设置",
            source_text="需恢复默认唤醒词、默认音量、默认工作模式、默认窗帘类型",
        ),
        requirement_entry(
            "FACTORY",
            "FACT-004",
            f"恢复出厂后恢复默认窗帘类型（{default_curtain_mode}）",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 恢复出厂设置",
            source_text="需恢复默认唤醒词、默认音量、默认工作模式、默认窗帘类型",
        ),
        requirement_entry(
            "WAKEWORD",
            "WAKE-001",
            "默认唤醒词“你好杜亚”一直生效",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置唤醒词",
            source_text="默认唤醒词为“你好杜亚”，一直生效",
        ),
        requirement_entry(
            "WAKEWORD",
            "WAKE-002",
            "其他候选唤醒词默认不能唤醒，通过设置后才能生效",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置唤醒词",
            source_text="其他所有唤醒词默认不能唤醒，通过触发此功能指令后才能生效",
        ),
        requirement_entry(
            "WAKEWORD",
            "WAKE-003",
            "设置唤醒词模式下说具体唤醒词后立即生效并退出设置模式",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置唤醒词",
            source_text="进入设置唤醒词模式后，若用户说了具体唤醒词，这此唤醒词立即生效并退出设置唤醒词模式",
        ),
        requirement_entry(
            "WAKEWORD",
            "WAKE-004",
            "设置唤醒词模式下说普通控制词可正常响应且不退出设置",
            source_kind="user_confirmed",
            source_section="用户确认 / 2026-04-13 / 设置唤醒词",
            source_text="设置唤醒词模式内，普通控制词本来就会响应；执行控制后仍应保持在设置窗口内，可继续完成唤醒词设置",
        ),
        requirement_entry(
            "WAKEWORD",
            "WAKE-005",
            "设置唤醒词模式超时后自动退出",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置唤醒词",
            source_text="进入设置唤醒词模式后，若超时退出识别模式，则也自动退出设置唤醒词模式",
        ),
        requirement_entry(
            "WORKMODE",
            "WORK-001",
            f"工作模式分为语音模式和嘀嗒模式，默认工作模式为 {default_work_mode}",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置工作模式",
            source_text="工作模式分为语音模式和嘀嗒模式，语音模式所有提示音正常播报，嘀嗒模式所有提示音播报嘀（音效），默认为语音模式",
        ),
        requirement_entry(
            "WORKMODE",
            "WORK-002",
            "未进入设置工作模式时，说语音模式或嘀嗒模式不响应",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置工作模式",
            source_text="未进入设置工作模式，用户说语音模式或嘀嗒模式不响应",
        ),
        requirement_entry(
            "WORKMODE",
            "WORK-003",
            "进入设置工作模式后，说语音模式或嘀嗒模式可切换到对应模式并退出设置",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置工作模式",
            source_text="进入设置工作模式后，若用户说了语音模式或嘀嗒模式，则切换到对应模式并退出设置工作模式",
        ),
        requirement_entry(
            "WORKMODE",
            "WORK-004",
            "进入设置工作模式后，说普通控制词可正常响应且不影响后续模式设置",
            source_kind="user_confirmed",
            source_section="用户确认 / 2026-04-13 / 设置工作模式",
            source_text="设置工作模式窗口内，普通控制词本来就会响应；执行控制后仍应保持可继续完成模式设置",
        ),
        requirement_entry(
            "WORKMODE",
            "WORK-005",
            "设置工作模式超时后自动退出",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置工作模式",
            source_text="进入设置工作模式后，若超时退出识别模式，则也自动退出工作模式",
        ),
        requirement_entry(
            "CURTAINMODE",
            "CURTAIN-001",
            f"窗帘模式分为窗帘模式、纱帘/窗纱模式、布帘模式，默认窗帘模式为 {default_curtain_mode}",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置窗帘模式",
            source_text="窗帘模式分为：窗帘模式、纱帘/窗纱模式、布帘模式三种，默认为窗帘模式一直生效",
        ),
        requirement_entry(
            "CURTAINMODE",
            "CURTAIN-002",
            "只有对应模式下相关的窗帘类型命令词才能响应",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置窗帘模式",
            source_text="只有对应模式下相关的窗帘类型命令词才能响应，例如：窗帘模式下，可响应的命令词有打开窗帘、关闭窗帘和停止窗帘，不可响应的命令词有：打开布帘、关闭纱帘、停止窗纱等",
        ),
        requirement_entry(
            "CURTAINMODE",
            "CURTAIN-003",
            "未进入设置窗帘模式时，说纱帘/窗纱模式、布帘模式不响应",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置窗帘模式",
            source_text="未进入设置窗帘模式，用户说纱帘/窗纱模式、布帘模式不响应",
        ),
        requirement_entry(
            "CURTAINMODE",
            "CURTAIN-004",
            "进入设置窗帘模式后，说纱帘/窗纱模式、布帘模式可切换到对应窗帘类型并退出设置",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置窗帘模式",
            source_text="进入设置窗帘模式后，若用户说纱帘/窗纱模式、布帘模式则切换到对应窗帘类型并退出设置窗帘模式",
        ),
        requirement_entry(
            "CURTAINMODE",
            "CURTAIN-005",
            "进入设置窗帘模式后，说普通控制词可正常响应且不影响后续模式设置",
            source_kind="user_confirmed",
            source_section="用户确认 / 2026-04-13 / 设置窗帘模式",
            source_text="设置窗帘模式窗口内，普通控制词本来就会响应；执行控制后仍应保持可继续完成模式设置",
        ),
        requirement_entry(
            "CURTAINMODE",
            "CURTAIN-006",
            "设置窗帘模式超时后自动退出",
            source_kind="requirement_doc",
            source_section="三、需求补充 / 设置窗帘模式",
            source_text="进入设置窗帘模式后，若超时退出识别模式，则也自动退出设置窗帘模式",
        ),
        requirement_entry(
            "BOOT",
            "DER-BOOT-001",
            "硬上下电后欢迎语正确播报",
            source_kind="tone_reference",
            source_section="tone.h / 历史参考用例",
            source_text="TONE_ID_3 = 你好,欢迎使用杜亚窗帘",
            acceptance_level="derived",
        ),
        requirement_entry(
            "BOOT",
            "DER-BOOT-002",
            "重启后日志链路恢复",
            source_kind="historical_runner",
            source_section="执行基线",
            source_text="重启后日志恢复且设备回到可交互状态",
            acceptance_level="derived",
        ),
        requirement_entry(
            "BOOT",
            "DER-BOOT-003",
            "重启后默认唤醒词仍可用",
            source_kind="historical_runner",
            source_section="执行基线",
            source_text="启动后默认待机态可再次唤醒",
            acceptance_level="derived",
        ),
        requirement_entry(
            "BASE",
            "DER-BASE-001",
            "默认唤醒词可进入交互态",
            source_kind="supplement",
            source_section="补充说明 / 唤醒验证",
            source_text="默认唤醒词一直生效，如果测试命令词或需要设备进入识别模式时，可使用默认唤醒词唤醒设备",
            acceptance_level="derived",
        ),
        requirement_entry(
            "BASE",
            "DER-BASE-002",
            "超时退出后不重新唤醒的直接命令不响应",
            source_kind="supplement",
            source_section="补充说明 / 唤醒验证",
            source_text="通过播报音频唤醒设备后等待超时，结合后续命令验证是否回到待唤醒状态",
            acceptance_level="derived",
        ),
        requirement_entry(
            "BASE",
            "DER-BASE-003",
            "超时退出后重新唤醒可恢复正常命令响应",
            source_kind="supplement",
            source_section="补充说明 / 唤醒验证",
            source_text="通过播报音频唤醒设备后等待超时，重新唤醒后继续交互",
            acceptance_level="derived",
        ),
        requirement_entry(
            "VOL",
            "DER-VOL-001",
            "增大音量命令单步反馈正确",
            source_kind="word_table",
            source_section="词条处理 / 增大音量",
            source_text="大声点 -> 好的",
            acceptance_level="derived",
        ),
        requirement_entry(
            "VOL",
            "DER-VOL-002",
            "减小音量命令单步反馈正确",
            source_kind="word_table",
            source_section="词条处理 / 减小音量",
            source_text="小声点 -> 好的",
            acceptance_level="derived",
        ),
        requirement_entry(
            "UART",
            "DER-UART-001",
            "稳定命令词触发协议发送可被协议口接收",
            source_kind="supplement",
            source_section="补充说明 / 串口配置",
            source_text="协议串口就检测协议输出",
            acceptance_level="derived",
        ),
        requirement_entry(
            "UART",
            "DER-UART-002",
            "日志口包含 asrKw、playId 等关键字段",
            source_kind="supplement",
            source_section="补充说明 / 串口配置",
            source_text="协议串口和日志串口，协议串口就检测协议输出的",
            acceptance_level="derived",
        ),
        requirement_entry(
            "UART",
            "DER-UART-003",
            "清除遥控器链路具备稳定日志与协议可观测性",
            source_kind="word_table",
            source_section="词条处理 / 清除遥控器",
            source_text="清除遥控器 -> 55 AA 05 01 16 00 18 AC",
            acceptance_level="derived",
        ),
        requirement_entry(
            "POWER",
            "DER-POWER-001",
            "上下电后日志恢复且设备回到可交互状态",
            source_kind="supplement",
            source_section="补充说明 / 掉电配置",
            source_text="掉电配置需要结合硬断电",
            acceptance_level="derived",
        ),
        requirement_entry(
            "FACTORY",
            "DER-FACT-001",
            "恢复出厂入口可达",
            source_kind="word_table",
            source_section="词条处理 / 恢复出厂模式",
            source_text="恢复出厂模式",
            acceptance_level="derived",
        ),
        requirement_entry(
            "WORKMODE",
            "DER-WORK-001",
            "工作模式设置入口可达",
            source_kind="word_table",
            source_section="词条处理 / 设置工作模式",
            source_text="设置工作模式",
            acceptance_level="derived",
        ),
        requirement_entry(
            "CURTAINMODE",
            "DER-CURTAINMODE-001",
            "窗帘模式设置入口可达",
            source_kind="word_table",
            source_section="词条处理 / 设置窗帘模式",
            source_text="设置窗帘模式",
            acceptance_level="derived",
        ),
        requirement_entry(
            "CURTAIN",
            "DER-CURTAIN-001",
            "默认窗帘模式下默认窗帘控制词可生效",
            source_kind="supplement",
            source_section="补充说明 / 设置窗帘模式",
            source_text="默认的窗帘模式一直生效，比如窗帘模式是默认的，说打开窗帘可以响应播报",
            acceptance_level="derived",
        ),
        requirement_entry(
            "CURTAIN",
            "DER-CURTAIN-002",
            "同意图同义词触发协议一致",
            source_kind="word_table",
            source_section="词条处理 / 窗帘同义词",
            source_text="打开窗帘 / 窗帘打开 等同义词",
            acceptance_level="derived",
        ),
        requirement_entry(
            "CURTAIN",
            "DER-CURTAIN-003",
            "模式切换后对应模式控制词可生效",
            source_kind="supplement",
            source_section="补充说明 / 设置窗帘模式",
            source_text="使用该模式的对应命令词去交互即可",
            acceptance_level="derived",
        ),
        requirement_entry(
            "CURTAIN",
            "DER-CURTAIN-004",
            "不匹配当前模式的控制词不响应",
            source_kind="supplement",
            source_section="补充说明 / 设置窗帘模式",
            source_text="默认的窗帘模式一直生效，比如窗帘模式是默认的，说打开纱帘则不响应播报",
            acceptance_level="derived",
        ),
        requirement_entry(
            "PHRASE",
            "DER-PHRASE-001",
            "词条按语义组执行，不是逐行平铺",
            source_kind="derived_logic",
            source_section="回归策略",
            source_text="词条按语义组组织执行",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
        requirement_entry(
            "PHRASE",
            "DER-PHRASE-002",
            "同义词按等价意图集合校验",
            source_kind="derived_logic",
            source_section="回归策略",
            source_text="同义词按等价意图集合校验",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
        requirement_entry(
            "PHRASE",
            "DER-PHRASE-003",
            "连续未识别后可重唤醒并继续剩余词条",
            source_kind="supplement",
            source_section="补充说明 / 设置类超时逻辑",
            source_text="三个设置模式均有超时时间，如果超时时间内没有播报对应的设置命令，则设置失败，使用设置前的功能",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
        requirement_entry(
            "SELECTOR",
            "DER-SELECTOR-001",
            "候选唤醒词可被设置并稳定唤醒",
            source_kind="supplement",
            source_section="补充说明 / 设置唤醒词",
            source_text="使用刚设置的唤醒词进行唤醒，连续三次。确认设置的唤醒词生效",
            acceptance_level="derived",
        ),
        requirement_entry(
            "SELECTOR",
            "DER-SELECTOR-002",
            "未选中的候选唤醒词不应误唤醒",
            source_kind="supplement",
            source_section="补充说明 / 设置唤醒词",
            source_text="还要验证除默认唤醒词和设置唤醒词外其他唤醒词是否生效，若生效则失败",
            acceptance_level="derived",
        ),
        requirement_entry(
            "SELECTOR",
            "DER-SELECTOR-003",
            "切换候选唤醒词后旧候选应失效",
            source_kind="derived_logic",
            source_section="回归策略",
            source_text="切换候选唤醒词后旧候选应失效",
            acceptance_level="derived",
        ),
        requirement_entry(
            "CTRL",
            "DER-CTRL-001",
            "遥控器配对入口可达并发送配对协议",
            source_kind="word_table",
            source_section="词条处理 / 配对遥控器",
            source_text="配对遥控器 -> 55 AA 05 01 12 00 80 AF",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
        requirement_entry(
            "CTRL",
            "DER-CTRL-002",
            "窗口内支持配对成功被动播报",
            source_kind="word_table",
            source_section="词条处理 / 配对成功",
            source_text="配对成功 -> 55 AA 05 03 06 01 F8 81",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
        requirement_entry(
            "CTRL",
            "DER-CTRL-003",
            "窗口内支持配对失败被动播报",
            source_kind="word_table",
            source_section="词条处理 / 配对失败",
            source_text="配对失败 -> 55 AA 05 03 06 00 78 84",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
        requirement_entry(
            "CTRL",
            "DER-CTRL-004",
            "配对窗口超时后协议注入无效",
            source_kind="historical_reference",
            source_section="历史参考用例 / 遥控器配对",
            source_text="窗口超时后协议注入无效",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
        requirement_entry(
            "CTRL",
            "DER-CTRL-005",
            "清除遥控器链路完整",
            source_kind="word_table",
            source_section="词条处理 / 清除遥控器",
            source_text="清除遥控器 -> 55 AA 05 01 16 00 18 AC",
            acceptance_level="derived",
            case_strategy=["positive", "negative", "state_transition"],
        ),
    ]


def build_requirement_modules(requirement_catalog: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in requirement_catalog:
        grouped.setdefault(item["module"], []).append(item)
    modules: list[dict[str, Any]] = []
    for module in REQUIREMENT_MODULE_ORDER:
        items = grouped.get(module, [])
        if not items:
            continue
        modules.append(
            {
                "module": module,
                "requirements": [item["requirement"] for item in items],
                "requirement_ids": [item["requirement_id"] for item in items],
            }
        )
    return modules


def build_reference_cases(spec_modules: list[dict[str, Any]], behavior_rules: dict[str, Any]) -> list[dict[str, Any]]:
    pair = behavior_rules["remote_pairing"]
    pair_entry = pair.get("entry", {})
    success = pair.get("success", {})
    failure = pair.get("failure", {})
    return [
        {
            "case_id": "TC_CTRL_001",
            "module": "CTRL",
            "name": clean_text(pair_entry.get("semantic")) or "配对遥控器",
            "precondition": "设备处于默认待命态",
            "steps": "唤醒后说【配对遥控器】进入遥控器配置窗口",
            "expected_result": "进入配置窗口并发送配对协议",
            "expected_protocol": norm_hex(pair_entry.get("send_protocol")),
            "expected_recv_protocol": "",
            "expected_tone_id": clean_text(pair_entry.get("tone_id")) or "TONE_ID_6",
            "expected_tone_text": clean_text(pair_entry.get("tts_text")) or "开始配对",
            "priority": "P1",
            "source": "behavior_rules.remote_pairing.entry",
        },
        {
            "case_id": "TC_CTRL_002",
            "module": "CTRL",
            "name": "配对成功-被动播报",
            "precondition": "已通过语音进入遥控器配置窗口",
            "steps": "在配对窗口内通过协议串口注入配对成功协议并检查 receive msg 与播报",
            "expected_result": "窗口内收到 receive msg 并播报配对成功",
            "expected_protocol": "-",
            "expected_recv_protocol": norm_hex(success.get("recv_protocol")),
            "expected_tone_id": clean_text(success.get("tone_id")) or "TONE_ID_7",
            "expected_tone_text": clean_text(success.get("tts_text")) or "配对成功",
            "priority": "P1",
            "source": "behavior_rules.remote_pairing.success",
        },
        {
            "case_id": "TC_CTRL_003",
            "module": "CTRL",
            "name": "配对失败-被动播报",
            "precondition": "已通过语音进入遥控器配置窗口",
            "steps": "在配对窗口内通过协议串口注入配对失败协议并检查 receive msg 与播报",
            "expected_result": "窗口内收到 receive msg 并播报配对失败",
            "expected_protocol": "-",
            "expected_recv_protocol": norm_hex(failure.get("recv_protocol")),
            "expected_tone_id": clean_text(failure.get("tone_id")) or "TONE_ID_8",
            "expected_tone_text": clean_text(failure.get("tts_text")) or "配对失败",
            "priority": "P1",
            "source": "behavior_rules.remote_pairing.failure",
        },
    ]


def build_coverage_matrix(requirement_catalog: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [dict(item) for item in requirement_catalog]


def write_coverage_markdown(path: Path, matrix: list[dict[str, Any]]) -> None:
    lines = ["# Requirement Coverage Matrix", ""]
    current_module = ""
    for item in matrix:
        module = item["module"]
        if module != current_module:
            current_module = module
            lines.append(f"## {module}")
            lines.append("")
        strategies = ", ".join(item.get("case_strategy", []))
        acceptance = item.get("acceptance_level", "main")
        source_kind = item.get("source_kind", "unknown")
        validation_target = item.get("validation_target", "runtime_case")
        lines.append(
            f"- {item['requirement_id']}: {item['requirement']} | strategy: {strategies} | "
            f"acceptance: {acceptance} | source: {source_kind} | validation: {validation_target}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_spec() -> dict[str, Any]:
    word_table_path, word_source = resolve_preferred_input(
        WORD_TABLE,
        ["词条处理.xlsx", "*词条*.xlsx"],
    )
    tone_file_path, tone_source = resolve_preferred_input(TONE_FILE, ["tone.h", "*tone*.h"])
    req_doc_path, req_source = resolve_preferred_input(
        REQ_DOC,
        ["需求文档.md", "*需求文档*.md"],
    )
    supplement_path, supplement_source = resolve_preferred_input(
        SUPPLEMENT_FILE,
        ["需求和用例补充说明.txt", "*补充说明*.txt"],
    )
    reference_xlsx_path = first_existing_path(
        REFERENCE_XLSX,
        ASSET_REF_DIR / "CSK5062_杜亚窗帘_测试用例v2.xlsx",
    )
    reference_source = "preferred_raw" if REFERENCE_XLSX.is_file() else ("asset_reference" if reference_xlsx_path else "missing")
    if not supplement_path:
        supplement_path = first_existing_path(ASSET_REF_DIR / "需求和用例补充说明.txt")
        if supplement_path:
            supplement_source = "asset_reference"

    legacy_bundle = load_latest_legacy_bundle(ROOT)
    legacy_spec = legacy_bundle.get("spec")
    legacy_cases = legacy_bundle.get("cases") or []
    legacy_spec_path = legacy_bundle.get("spec_path")
    legacy_cases_path = legacy_bundle.get("cases_path")

    requirement_text = read_text(req_doc_path) if req_doc_path else ""
    supplement_text = read_text(supplement_path) if supplement_path else ""
    meta_text = "\n".join(part for part in [requirement_text, supplement_text] if part)
    meta = parse_requirement_text(meta_text)
    tones = parse_tone_file(tone_file_path) if tone_file_path else []

    if word_table_path and word_table_path.is_file():
        word_sheet = parse_word_sheet(word_table_path)
        rows = word_sheet["rows"]
        word_sheet["source"] = "word_table"
    elif legacy_spec:
        rows = rows_from_legacy_spec(legacy_spec)
        legacy_wakes = legacy_spec.get("word_sheet", {}).get("wake_words", []) or legacy_spec.get("behavior_rules", {}).get("wake_word_setting", {}).get("candidate_wake_words", [])
        word_sheet = {
            "rows": rows,
            "wake_words": legacy_wakes,
            "semantic_count": len({clean_text(row.get("semantic")) for row in rows if clean_text(row.get("semantic"))}),
            "source": "legacy_normalized_spec",
        }
        word_source = "legacy_normalized_spec"
    else:
        raise FileNotFoundError("word table not found and no legacy normalized_spec fallback is available")

    if not tones and legacy_spec:
        tones = legacy_spec.get("tones", {}).get("items", [])
        tone_source = "legacy_normalized_spec"

    if legacy_spec:
        project = legacy_spec.get("requirement", {}).get("project", {})
        defaults = legacy_spec.get("requirement", {}).get("defaults", {})
        legacy_meta = {
            "project_name": clean_text(project.get("project_name")) or "CSK5062杜亚窗帘",
            "branch_name": clean_text(project.get("branch_name")) or "unknown",
            "chip_model": clean_text(project.get("chip_model")) or "CSK5062",
            "firmware_version": clean_text(project.get("firmware_version")) or "unknown",
            "wake_timeout_s": int(defaults.get("wake_timeout_s", 15) or 15),
            "volume_levels": int(defaults.get("volume_levels", 4) or 4),
            "default_volume": int(defaults.get("default_volume", 3) or 3),
            "protocol_baudrate": int(defaults.get("protocol_baudrate", 9600) or 9600),
            "log_baudrate": int(defaults.get("log_baudrate", 115200) or 115200),
        }
        if not requirement_text:
            meta = {**legacy_meta, **meta}
            default_numbers = {
                "wake_timeout_s": 15,
                "volume_levels": 4,
                "default_volume": 3,
                "protocol_baudrate": 9600,
                "log_baudrate": 115200,
            }
            for key, default_value in default_numbers.items():
                if meta.get(key) == default_value:
                    meta[key] = legacy_meta[key]
        meta["fallback_wake_words"] = legacy_spec.get("word_sheet", {}).get("wake_words", []) or legacy_spec.get("behavior_rules", {}).get("wake_word_setting", {}).get("candidate_wake_words", [])

    behavior_rules = build_behavior_rules(meta, rows, tones, legacy_cases)
    requirement_catalog = build_requirement_catalog(meta)
    modules = build_requirement_modules(requirement_catalog)
    reference_cases = build_reference_cases(modules, behavior_rules)
    coverage_matrix = build_coverage_matrix(requirement_catalog)
    semantic_groups = build_semantic_groups(rows, tones)

    runtime_send_protocol = norm_hex(behavior_rules["remote_pairing"]["entry"].get("send_protocol"))
    gaps = []
    if runtime_send_protocol:
        gaps.append(
            {
                "type": "runtime_mismatch",
                "case_id": "TC_CTRL_001",
                "field": "ctrl_pair_protocol_runtime",
                "label": "配对遥控器协议",
                "source_value": runtime_send_protocol,
                "runtime_value": "55 AA 04 01 12 B2 17",
                "note": "现网实测协议与词表/生成用例期望不一致，请同步回需求表、词表和结果摘要。",
            }
        )

    selected_requirement_source = req_doc_path or legacy_spec_path
    selected_word_source = word_table_path or legacy_spec_path
    selected_tone_source = tone_file_path or legacy_spec_path
    selected_reference_source = reference_xlsx_path

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "requirement": {
            "project": {
                "project_name": meta["project_name"],
                "branch_name": meta["branch_name"],
                "chip_model": meta["chip_model"],
                "firmware_version": meta["firmware_version"],
            },
            "defaults": {
                "wake_timeout_s": meta["wake_timeout_s"],
                "volume_levels": meta["volume_levels"],
                "default_volume": meta["default_volume"],
                "wake_word_power_retained": bool(meta.get("wake_word_power_retained", False)),
                "volume_power_retained": bool(meta.get("volume_power_retained", False)),
                "default_work_mode": clean_text(meta.get("default_work_mode")) or "语音模式",
                "default_curtain_mode": clean_text(meta.get("default_curtain_mode")) or "窗帘模式",
                "protocol_uart": str(DEFAULT_RUNTIME_PORTS["protocol_port"]),
                "protocol_baudrate": meta["protocol_baudrate"],
                "log_uart": str(DEFAULT_RUNTIME_PORTS["log_port"]),
                "log_baudrate": meta["log_baudrate"],
            },
            "raw_text": requirement_text,
            "supplement_text": supplement_text,
        },
        "runtime": {
            "log_port": {"port": str(DEFAULT_RUNTIME_PORTS["log_port"]), "baudrate": meta["log_baudrate"]},
            "protocol_port": {
                "port": str(DEFAULT_RUNTIME_PORTS["protocol_port"]),
                "baudrate": meta["protocol_baudrate"],
                "baudrate_source": "requirement_doc" if req_doc_path else "legacy_default",
                "verify_in_smoke": True,
            },
            "loglevel_command": "loglevel 4",
            "reboot_command": "reboot",
        },
        "modules": modules,
        "requirement_catalog": requirement_catalog,
        "behavior_rules": behavior_rules,
        "semantic_groups": semantic_groups,
        "tones": {"items": tones},
        "word_sheet": word_sheet,
        "reference_cases": {"case_rows": reference_cases},
        "coverage_matrix": coverage_matrix,
        "sources": {
            "requirement_doc": str(selected_requirement_source) if selected_requirement_source else "",
            "supplement_file": str(supplement_path) if supplement_path else "",
            "word_table": str(selected_word_source) if selected_word_source else "",
            "tone_file": str(selected_tone_source) if selected_tone_source else "",
            "reference_cases": str(selected_reference_source) if selected_reference_source else "",
            "legacy_spec": str(legacy_spec_path) if legacy_spec_path else "",
            "legacy_cases": str(legacy_cases_path) if legacy_cases_path else "",
        },
        "source_resolution": {
            "requirement_doc": req_source if req_doc_path else ("legacy_normalized_spec" if legacy_spec_path else "missing"),
            "supplement_file": supplement_source,
            "word_table": word_source,
            "tone_file": tone_source if tones else "missing",
            "reference_cases": reference_source,
            "legacy_spec": "available" if legacy_spec_path else "missing",
            "legacy_cases": "available" if legacy_cases_path else "missing",
        },
        "coverage": {
            "default_policy": "full",
            "semantic_rows": word_sheet["semantic_count"],
            "module_count": len(modules),
            "main_requirement_count": len([item for item in requirement_catalog if item.get("acceptance_level") == "main"]),
            "derived_requirement_count": len([item for item in requirement_catalog if item.get("acceptance_level") != "main"]),
            "reference_case_count": len(reference_cases),
        },
        "gaps": gaps,
    }


def main() -> None:
    WORK_DIR.mkdir(exist_ok=True)
    spec = build_spec()
    write_json(WORK_DIR / "normalized_spec.json", spec)
    write_json(WORK_DIR / "tone_map.json", {"items": spec["tones"]["items"]})
    sources = spec["sources"]
    resolution = spec.get("source_resolution", {})
    inventory_lines = [
        "# Source Inventory",
        "",
        f"- requirement: {sources['requirement_doc']} | status={resolution.get('requirement_doc', 'unknown')}",
        f"- supplement: {sources['supplement_file']} | status={resolution.get('supplement_file', 'unknown')}",
        f"- word table: {sources['word_table']} | status={resolution.get('word_table', 'unknown')}",
        f"- tone file: {sources['tone_file']} | status={resolution.get('tone_file', 'unknown')}",
        f"- reference xlsx: {sources['reference_cases']} | status={resolution.get('reference_cases', 'unknown')}",
    ]
    if sources.get("legacy_spec"):
        inventory_lines.append(f"- legacy spec: {sources['legacy_spec']} | status={resolution.get('legacy_spec', 'unknown')}")
    if sources.get("legacy_cases"):
        inventory_lines.append(f"- legacy cases: {sources['legacy_cases']} | status={resolution.get('legacy_cases', 'unknown')}")
    (WORK_DIR / "source_inventory.md").write_text("\n".join(inventory_lines) + "\n", encoding="utf-8")
    write_coverage_markdown(WORK_DIR / "coverage_matrix.md", spec["coverage_matrix"])
    print(f"Wrote {WORK_DIR / 'normalized_spec.json'}")
    print(f"Wrote {WORK_DIR / 'tone_map.json'}")
    print(f"Wrote {WORK_DIR / 'source_inventory.md'}")
    print(f"Wrote {WORK_DIR / 'coverage_matrix.md'}")


if __name__ == "__main__":
    main()
